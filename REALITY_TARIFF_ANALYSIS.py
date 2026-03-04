#!/usr/bin/env python3
"""
Hospital Tariff Negotiation Tool
=================================

Compares hospital tariffs against Clearline's reality tariff and generates
negotiated prices based on banding rules.

Usage:
    python hospital_tariff_negotiation.py --reality reality_tariff.csv --hospitals hospital_list.csv --output negotiated_tariffs.xlsx

Author: KLAIRE AI Medical Analyst
Date: January 2026
"""

import pandas as pd
import duckdb
import argparse
import sys
import os
import time
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

try:
    import toml
    TOML_AVAILABLE = True
except ImportError:
    TOML_AVAILABLE = False

try:
    import gspread
    from google.oauth2.service_account import Credentials
    GSPREAD_AVAILABLE = True
except ImportError:
    GSPREAD_AVAILABLE = False
    print("⚠️  Warning: gspread not installed. Google Sheets features will be disabled.")
    print("   Install with: pip install gspread google-auth")

class HospitalTariffNegotiator:
    """
    Handles hospital tariff negotiation based on reality tariff bands
    """
    
    def __init__(self, reality_tariff_path: str, db_path: str = None, use_google_sheets: bool = False, 
                 google_credentials_path: str = None, spreadsheet_name: str = "TARIFF EXISTING"):
        """
        Initialize negotiator with reality tariff and database connection
        
        Parameters:
        -----------
        reality_tariff_path : str
            Path to reality tariff CSV file
        db_path : str, optional
            Path to DuckDB database file. If not provided, will search common locations.
        use_google_sheets : bool, optional
            Whether to use Google Sheets for provider list and output
        google_credentials_path : str, optional
            Path to Google service account JSON credentials file
        spreadsheet_name : str, optional
            Name of the Google Spreadsheet to use (default: "TARIFF EXISTING")
        """
        self.reality_tariff_path = reality_tariff_path
        self.reality_df = None
        self.conn = None
        self.use_google_sheets = use_google_sheets and GSPREAD_AVAILABLE
        self.spreadsheet_name = spreadsheet_name
        self.gs_client = None
        self.gs_spreadsheet = None
        self.gs_main_sheet = None
        self.gs_worksheets_cache = None  # Cache for worksheet names to avoid repeated API calls
        
        # Auto-detect database path if not provided
        if db_path is None:
            db_path = self._find_database()
        
        self.db_path = db_path
        
        self._load_reality_tariff()
        self._connect_database()
        
        # Connect to Google Sheets if enabled
        if self.use_google_sheets:
            self._connect_google_sheets(google_credentials_path)
    
    def _find_database(self):
        """
        Try to find the DuckDB database in common locations
        """
        # Common locations to check
        possible_paths = [
            'ai_driven_data.duckdb',
            '../ai_driven_data.duckdb',
            '/home/claude/ai_driven_data.duckdb',
            '/mnt/project/ai_driven_data.duckdb',
        ]
        
        for path in possible_paths:
            if Path(path).exists():
                print(f"✅ Found database at: {path}")
                return path
        
        # If not found, provide helpful error
        print("❌ Could not find DuckDB database file.")
        print("   Please specify the path using --db parameter")
        print("   Example: --db /path/to/ai_driven_data.duckdb")
        sys.exit(1)
    
    def _load_reality_tariff(self):
        """Load and validate reality tariff CSV"""
        try:
            # Read CSV, skipping first row if it's a header description
            df = pd.read_csv(self.reality_tariff_path)
            
            # Check if header row is actually a description row (contains "USING" or has "Unnamed" columns)
            header_is_description = (
                any('USING' in str(col).upper() for col in df.columns) or
                any('Unnamed' in str(col) for col in df.columns)
            )
            
            if header_is_description:
                df = pd.read_csv(self.reality_tariff_path, skiprows=1)
            
            # Validate required columns
            required_cols = ['procedurecode', 'band_d', 'band_c']
            missing_cols = [col for col in required_cols if col not in df.columns]
            
            if missing_cols:
                raise ValueError(f"Reality tariff missing required columns: {missing_cols}")
            
            # Normalize procedure codes
            df['procedurecode'] = df['procedurecode'].astype(str).str.strip().str.lower()
            
            # Convert band columns to numeric
            for col in ['band_d', 'band_c', 'band_b', 'band_a', 'band_special']:
                if col in df.columns:
                    df[col] = pd.to_numeric(df[col], errors='coerce')
            
            # Get procedure name column (might be 'procedure name' or 'proceduredesc')
            proc_name_col = None
            for col in ['procedure name', 'proceduredesc', 'procedure_name', 'description']:
                if col in df.columns:
                    proc_name_col = col
                    break
            
            if proc_name_col:
                df['procedure_name'] = df[proc_name_col]
            else:
                df['procedure_name'] = df['procedurecode']  # Fallback
            
            self.reality_df = df
            print(f"✅ Loaded {len(df)} procedures from reality tariff")
            
        except Exception as e:
            print(f"❌ Error loading reality tariff: {e}")
            sys.exit(1)
    
    def _connect_database(self):
        """Connect to DuckDB database"""
        try:
            if not Path(self.db_path).exists():
                raise FileNotFoundError(f"Database not found: {self.db_path}")
            
            self.conn = duckdb.connect(self.db_path, read_only=True)
            print(f"✅ Connected to database: {self.db_path}")
            
        except Exception as e:
            print(f"❌ Error connecting to database: {e}")
            sys.exit(1)
    
    def _find_google_credentials(self):
        """Find Google service account credentials file"""
        possible_paths = [
            'CREDENTIALS.json',
            '../CREDENTIALS.json',
            os.path.expanduser('~/CREDENTIALS.json'),
            '.streamlit/CREDENTIALS.json',
        ]
        
        # Also check secrets.toml for credentials path
        if TOML_AVAILABLE:
            secrets_path = Path('.streamlit/secrets.toml')
            if secrets_path.exists():
                try:
                    secrets = toml.load(secrets_path)
                    if 'google' in secrets and 'credentials_path' in secrets['google']:
                        possible_paths.insert(0, secrets['google']['credentials_path'])
                except:
                    pass
        
        for path in possible_paths:
            if Path(path).exists():
                print(f"✅ Found Google credentials at: {path}")
                return path
        
        return None
    
    def _connect_google_sheets(self, credentials_path: str = None):
        """Connect to Google Sheets using service account credentials"""
        if not GSPREAD_AVAILABLE:
            print("❌ gspread not available. Install with: pip install gspread google-auth")
            sys.exit(1)
        
        try:
            # Find credentials file if not provided
            if credentials_path is None:
                credentials_path = self._find_google_credentials()
            
            if credentials_path is None or not Path(credentials_path).exists():
                raise FileNotFoundError(
                    f"Google credentials file not found. Please provide path to service account JSON file.\n"
                    f"   Expected locations: CREDENTIALS.json, .streamlit/CREDENTIALS.json, or specify with --google-creds"
                )
            
            # Define scopes
            scopes = [
                'https://www.googleapis.com/auth/spreadsheets',
                'https://www.googleapis.com/auth/drive'
            ]
            
            # Load credentials
            creds = Credentials.from_service_account_file(credentials_path, scopes=scopes)
            self.gs_client = gspread.authorize(creds)
            
            # Open the spreadsheet - try both with and without .xlsx extension
            spreadsheet_found = False
            spreadsheet_names_to_try = [self.spreadsheet_name]
            
            # If name doesn't end with .xlsx, also try with .xlsx
            if not self.spreadsheet_name.endswith('.xlsx'):
                spreadsheet_names_to_try.append(f"{self.spreadsheet_name}.xlsx")
            # If name ends with .xlsx, also try without it
            elif self.spreadsheet_name.endswith('.xlsx'):
                spreadsheet_names_to_try.insert(0, self.spreadsheet_name.replace('.xlsx', ''))
            
            for name_to_try in spreadsheet_names_to_try:
                try:
                    self.gs_spreadsheet = self.gs_client.open(name_to_try)
                    print(f"✅ Connected to Google Spreadsheet: {name_to_try}")
                    spreadsheet_found = True
                    break
                except gspread.exceptions.SpreadsheetNotFound:
                    continue
            
            if not spreadsheet_found:
                # Try to list available spreadsheets to help user
                try:
                    all_spreadsheets = self.gs_client.list_spreadsheet_files()
                    if all_spreadsheets:
                        print(f"\n📋 Available spreadsheets (first 10):")
                        for sheet in all_spreadsheets[:10]:
                            print(f"   - {sheet['name']}")
                        if len(all_spreadsheets) > 10:
                            print(f"   ... and {len(all_spreadsheets) - 10} more")
                    else:
                        print("\n⚠️  No spreadsheets found. Please share a spreadsheet with your service account.")
                except Exception as e:
                    print(f"\n⚠️  Could not list spreadsheets: {e}")
                
                raise FileNotFoundError(
                    f"\n❌ Spreadsheet '{self.spreadsheet_name}' not found.\n"
                    f"   Tried: {', '.join(spreadsheet_names_to_try)}\n"
                    f"   Please ensure:\n"
                    f"   1. The spreadsheet exists with the exact name\n"
                    f"   2. It is shared with your service account email\n"
                    f"   3. Use --spreadsheet-name to specify the exact name if different"
                )
            
            # Get the main tracking sheet.
            # Your setup: spreadsheet = "TARIFF EXISTING", sheet with providers = "PROVIDERS".
            # Preference order:
            # 1) Worksheet named "PROVIDERS"
            # 2) Worksheet with the same name as the spreadsheet (e.g. "TARIFF EXISTING")
            # 3) First worksheet in the spreadsheet
            try:
                # First try explicit "PROVIDERS" sheet
                self.gs_main_sheet = self.gs_spreadsheet.worksheet("PROVIDERS")
                print("✅ Using main tracking sheet: PROVIDERS")
            except gspread.exceptions.WorksheetNotFound:
                try:
                    # Then try sheet with same name as spreadsheet
                    self.gs_main_sheet = self.gs_spreadsheet.worksheet(self.spreadsheet_name)
                    print(f"✅ Using main tracking sheet: {self.spreadsheet_name}")
                except gspread.exceptions.WorksheetNotFound:
                    # Finally, fall back to the first available worksheet
                    worksheets = self.gs_spreadsheet.worksheets()
                    if worksheets:
                        self.gs_main_sheet = worksheets[0]
                        print(f"✅ Using first available sheet: {self.gs_main_sheet.title}")
                    else:
                        raise FileNotFoundError(f"No worksheets found in spreadsheet '{self.spreadsheet_name}'")
            
            # Cache worksheet names to avoid repeated API calls
            try:
                self.gs_worksheets_cache = {ws.title: ws for ws in self.gs_spreadsheet.worksheets()}
                print(f"✅ Cached {len(self.gs_worksheets_cache)} worksheet names")
            except Exception as e:
                print(f"⚠️  Could not cache worksheets: {e}")
                self.gs_worksheets_cache = {}
            
        except Exception as e:
            print(f"❌ Error connecting to Google Sheets: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def _get_providers_from_google_sheets(self):
        """Get provider list from Google Sheets and check status"""
        if not self.use_google_sheets:
            return None
        
        try:
            # Get all records from the main sheet
            records = self.gs_main_sheet.get_all_records()
            
            if not records:
                print("⚠️  No providers found in Google Sheet")
                return []
            
            # Convert to DataFrame
            df = pd.DataFrame(records)
            
            # Find the provider name column
            name_col = None
            for col in ['NAME OF PROVIDERS', 'name of providers', 'providername', 'name']:
                if col in df.columns:
                    name_col = col
                    break
            
            if name_col is None:
                raise ValueError(f"Could not find provider name column. Available columns: {list(df.columns)}")
            
            # Find analysis and status columns
            analysis_col = None
            status_col = None
            for col in df.columns:
                if 'analysis' in col.lower():
                    analysis_col = col
                if 'status' in col.lower():
                    status_col = col
            
            # Filter providers that should be processed
            # Skip if STATUS is "DONE" (case insensitive) or ANALYSIS is "analysed"
            providers_to_process = []
            provider_row_map = {}  # Map provider name to row number for updating
            
            for idx, row in df.iterrows():
                provider_name = str(row[name_col]).strip()
                
                if not provider_name:
                    continue
                
                # Check status column
                status = ""
                if status_col and status_col in row:
                    status = str(row[status_col]).strip().upper()
                
                # Check analysis column
                analysis = ""
                if analysis_col and analysis_col in row:
                    analysis = str(row[analysis_col]).strip().lower()
                
                # Skip if status is "DONE" or analysis is "analysed"
                if status == "DONE":
                    print(f"   ⏭️  Skipping {provider_name} (Status: DONE)")
                    continue
                
                if analysis == "analysed":
                    print(f"   ⏭️  Skipping {provider_name} (Already analysed)")
                    continue
                
                # Check if sheet already exists using cache (avoids API call)
                sheet_name = provider_name[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
                sheet_exists = False
                
                # Use cache if available
                if self.gs_worksheets_cache and sheet_name in self.gs_worksheets_cache:
                    sheet_exists = True
                    # If sheet exists but not marked as analysed, mark it and skip
                    if analysis != "analysed":
                        print(f"   📝 Found existing sheet for {provider_name}, marking as analysed")
                        self._update_analysis_status(provider_name, "analysed")
                        # Small delay after update
                        time.sleep(0.5)
                    # Skip this provider since it already has a sheet
                    continue
                else:
                    # Not in cache, try to check (but this will trigger API call)
                    try:
                        # Small delay before checking to avoid quota issues
                        time.sleep(0.3)
                        self.gs_spreadsheet.worksheet(sheet_name)
                        sheet_exists = True
                        # Add to cache
                        if self.gs_worksheets_cache is not None:
                            self.gs_worksheets_cache[sheet_name] = self.gs_spreadsheet.worksheet(sheet_name)
                        # If sheet exists but not marked as analysed, mark it and skip
                        if analysis != "analysed":
                            print(f"   📝 Found existing sheet for {provider_name}, marking as analysed")
                            self._update_analysis_status(provider_name, "analysed")
                            # Small delay after update
                            time.sleep(0.5)
                        # Skip this provider since it already has a sheet
                        continue
                    except gspread.exceptions.WorksheetNotFound:
                        pass
                    except gspread.exceptions.APIError as e:
                        # Handle quota errors gracefully
                        if '429' in str(e):
                            print(f"   ⚠️  Quota limit reached while checking for {provider_name}. Will retry later.")
                            # Don't skip - let it process later
                            pass
                        else:
                            raise
                
                # Add to process list
                providers_to_process.append(provider_name)
                provider_row_map[provider_name] = idx + 2  # +2 because: 1 for header, 1 for 0-indexed
            
            self.provider_row_map = provider_row_map
            return providers_to_process
            
        except Exception as e:
            print(f"❌ Error reading from Google Sheets: {e}")
            import traceback
            traceback.print_exc()
            return None
    
    def _update_analysis_status(self, provider_name: str, status: str = "analysed"):
        """Update the ANALYSIS column for a provider in Google Sheets"""
        if not self.use_google_sheets:
            return
        
        try:
            # Get all values from the sheet
            all_values = self.gs_main_sheet.get_all_values()
            
            if not all_values or len(all_values) < 2:
                return
            
            # Find header row
            header_row = all_values[0]
            
            # Find column indices
            name_col_idx = None
            analysis_col_idx = None
            
            for idx, header in enumerate(header_row):
                header_lower = str(header).strip().lower()
                if 'name of providers' in header_lower or 'providername' in header_lower or (header_lower == 'name' and name_col_idx is None):
                    name_col_idx = idx
                if 'analysis' in header_lower:
                    analysis_col_idx = idx
            
            if name_col_idx is None or analysis_col_idx is None:
                print(f"   ⚠️  Could not find required columns in Google Sheet")
                return
            
            # Find the row for this provider
            for row_idx, row in enumerate(all_values[1:], start=2):  # Start from row 2 (skip header)
                if len(row) > name_col_idx:
                    provider_in_row = str(row[name_col_idx]).strip()
                    if provider_in_row == provider_name.strip():
                        # Update the analysis column
                        cell_address = f"{self._get_column_letter(analysis_col_idx + 1)}{row_idx}"
                        # Use named arguments with values as a list
                        self.gs_main_sheet.update(range_name=cell_address, values=[[status]])
                        print(f"   ✅ Updated ANALYSIS status for {provider_name}")
                        return
            
        except Exception as e:
            print(f"   ⚠️  Error updating analysis status for {provider_name}: {e}")
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to letter (1 -> A, 2 -> B, etc.)"""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + (col_num % 26)) + result
            col_num //= 26
        return result
    
    def _create_provider_sheet_in_google_sheets(self, provider_name: str, negotiated_df: pd.DataFrame, max_retries: int = 3):
        """Create a new sheet in Google Sheets for a provider with the negotiated tariff data"""
        if not self.use_google_sheets:
            return False
        
        # Sanitize sheet name (max 31 chars, no special characters)
        sheet_name = provider_name[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
        
        # Check if sheet already exists using cache (avoids API call)
        ws = None
        if self.gs_worksheets_cache and sheet_name in self.gs_worksheets_cache:
            # Sheet exists in cache
            ws = self.gs_worksheets_cache[sheet_name]
            # Clear existing sheet with retry logic
            for attempt in range(max_retries):
                try:
                    ws.clear()
                    break
                except gspread.exceptions.APIError as e:
                    if '429' in str(e) and attempt < max_retries - 1:
                        wait_time = (2 ** attempt) * 2  # Exponential backoff: 2s, 4s, 8s
                        print(f"   ⏳ Quota limit hit, waiting {wait_time}s before retry {attempt + 1}/{max_retries}...")
                        time.sleep(wait_time)
                    else:
                        raise
        else:
            # Sheet not in cache - try to create it directly (avoid checking first to save API calls)
            for attempt in range(max_retries):
                try:
                    # Try to create new sheet directly
                    ws = self.gs_spreadsheet.add_worksheet(
                        title=sheet_name,
                        rows=len(negotiated_df) + 1,
                        cols=5
                    )
                    # Add to cache
                    if self.gs_worksheets_cache is not None:
                        self.gs_worksheets_cache[sheet_name] = ws
                    # Delay after creating sheet to avoid quota issues
                    time.sleep(2.0)
                    break
                except gspread.exceptions.APIError as e:
                    if '429' in str(e) and attempt < max_retries - 1:
                        wait_time = (2 ** attempt) * 2  # Exponential backoff: 2s, 4s, 8s
                        print(f"   ⏳ Quota limit hit, waiting {wait_time}s before retry {attempt + 1}/{max_retries}...")
                        time.sleep(wait_time)
                    else:
                        # If it's a duplicate sheet error, try to get existing sheet
                        if 'already exists' in str(e).lower() or attempt == max_retries - 1:
                            try:
                                ws = self.gs_spreadsheet.worksheet(sheet_name)
                                ws.clear()
                                if self.gs_worksheets_cache is not None:
                                    self.gs_worksheets_cache[sheet_name] = ws
                                break
                            except:
                                raise e
                        else:
                            raise
                except gspread.exceptions.WorksheetNotFound:
                    # This shouldn't happen when creating, but handle it
                    raise
        
        if ws is None:
            return False
        
        try:
            
            # Prepare data for Google Sheets
            # Convert DataFrame to list of lists
            data = [negotiated_df.columns.tolist()]  # Header row
            for _, row in negotiated_df.iterrows():
                data.append(row.tolist())
            
            # Write data to sheet (using named arguments to avoid deprecation warning)
            ws.update(range_name='A1', values=data)
            
            # Small delay to avoid hitting Google API quota limits
            time.sleep(0.5)
            
            # Format header row
            header_range = f'A1:{self._get_column_letter(len(negotiated_df.columns))}1'
            try:
                ws.format(header_range, {
                    'backgroundColor': {'red': 0.12, 'green': 0.31, 'blue': 0.47},
                    'textFormat': {'foregroundColor': {'red': 1.0, 'green': 1.0, 'blue': 1.0}, 'bold': True},
                    'horizontalAlignment': 'CENTER'
                })
            except Exception as e:
                # If formatting fails, continue without it
                print(f"   ⚠️  Could not format header row: {e}")
            
            # Highlight rows that need review (yellow background)
            for idx, row in negotiated_df.iterrows():
                if pd.notna(row.get('Comment')) and "CHECK IF THIS PRICE IS WELL PRICED" in str(row.get('Comment', '')):
                    row_num = idx + 2  # +2 for header and 0-index
                    row_range = f'A{row_num}:{self._get_column_letter(len(negotiated_df.columns))}{row_num}'
                    try:
                        ws.format(row_range, {
                            'backgroundColor': {'red': 1.0, 'green': 0.96, 'blue': 0.8}
                        })
                    except Exception as e:
                        # If formatting fails, continue without it
                        pass
            
            print(f"   ✅ Created/updated sheet '{sheet_name}' in Google Sheets")
            return True
            
        except Exception as e:
            print(f"   ❌ Error creating sheet for {provider_name}: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def _get_hospital_tariff(self, hospital_name: str):
        """
        Query database for hospital tariff using hospital name or ProviderID
        
        Parameters:
        -----------
        hospital_name : str
            Name of the hospital or ProviderID (numeric string)
        
        Returns:
        --------
        DataFrame with columns: procedurecode, tariffamount
        """
        try:
            # Check if input is a numeric ProviderID
            is_provider_id = False
            provider_id = None
            try:
                provider_id = str(hospital_name.strip())
                # Try to match as ProviderID if it's numeric or matches pattern
                if provider_id.isdigit() or provider_id.startswith('LA/') or provider_id.startswith('NHIS-'):
                    is_provider_id = True
            except:
                pass
            
            if is_provider_id:
                # Query by ProviderID
                # First try the standard join, but also check if we need to query by tariffname
                query = """
                SELECT 
                    LOWER(TRIM(t.procedurecode)) as procedurecode,
                    t.tariffamount as hospital_price
                FROM "AI DRIVEN DATA"."PROVIDERS" p
                INNER JOIN "AI DRIVEN DATA"."PROVIDERS_TARIFF" pt 
                    ON p.protariffid = pt.protariffid
                INNER JOIN "AI DRIVEN DATA"."TARIFF" t 
                    ON pt.tariffid = t.tariffid
                WHERE CAST(p.providerid AS VARCHAR) = ?
                    AND (t.tariffamount > 0 OR t.tariffamount IS NULL)
                ORDER BY t.procedurecode
                """
                result = self.conn.execute(query, [provider_id]).fetchdf()
                
                # If no results, try querying by tariffname if provider name suggests a tariff name
                if len(result) == 0:
                    # Get provider name to check for tariff name mapping
                    provider_info = self.conn.execute("""
                        SELECT providername 
                        FROM "AI DRIVEN DATA"."PROVIDERS" 
                        WHERE CAST(providerid AS VARCHAR) = ?
                    """, [provider_id]).fetchone()
                    
                    if provider_info:
                        provider_name = provider_info[0]
                        # Try to find tariff by name pattern (e.g., "Subol Hospital" -> "Subol Tariff")
                        tariff_name_pattern = f"%{provider_name.split()[0]}%Tariff%"
                        alt_query = """
                        SELECT 
                            LOWER(TRIM(t.procedurecode)) as procedurecode,
                            t.tariffamount as hospital_price
                        FROM "AI DRIVEN DATA"."TARIFF" t
                        WHERE LOWER(COALESCE(t.tariffname, '')) LIKE LOWER(?)
                            AND t.tariffamount > 0
                        ORDER BY t.procedurecode
                        """
                        result = self.conn.execute(alt_query, [tariff_name_pattern]).fetchdf()
            else:
                # Query by provider name (use LIKE with wildcards for flexible matching)
                query = """
                SELECT 
                    LOWER(TRIM(t.procedurecode)) as procedurecode,
                    t.tariffamount as hospital_price
                FROM "AI DRIVEN DATA"."PROVIDERS" p
                INNER JOIN "AI DRIVEN DATA"."PROVIDERS_TARIFF" pt 
                    ON p.protariffid = pt.protariffid
                INNER JOIN "AI DRIVEN DATA"."TARIFF" t 
                    ON pt.tariffid = t.tariffid
                WHERE LOWER(TRIM(p.providername)) LIKE LOWER(TRIM(?))
                    AND t.tariffamount > 0
                ORDER BY t.procedurecode
                """
                # Add wildcards for flexible matching
                search_pattern = f"%{hospital_name.strip()}%"
                result = self.conn.execute(query, [search_pattern]).fetchdf()
            
            if len(result) == 0:
                print(f"⚠️  No tariff found for: {hospital_name.strip()}")
                return None
            
            return result
            
        except Exception as e:
            print(f"❌ Error querying tariff for {hospital_name}: {e}")
            return None
    
    def _calculate_proposed_price(self, hospital_price: float, band_d: float, band_c: float):
        """
        Calculate Clearline proposed price based on negotiation rules
        
        Rules:
        1. If hospital_price < band_d: use band_d
        2. If band_d <= hospital_price < band_c: use band_c
        3. If hospital_price >= band_c: use hospital_price (flag for review)
        
        Parameters:
        -----------
        hospital_price : float
            Hospital's current price
        band_d : float
            Reality tariff Band D price
        band_c : float
            Reality tariff Band C price
        
        Returns:
        --------
        tuple: (proposed_price, comment)
        """
        if pd.isna(hospital_price) or pd.isna(band_d) or pd.isna(band_c):
            return (None, "Missing price data")
        
        if hospital_price < band_d:
            # Hospital price is below band D - use band D
            return (band_d, "Increased to Band D minimum")
        
        elif hospital_price < band_c:
            # Hospital price is between band D and C - use band C
            return (band_c, "Adjusted to Band C")
        
        else:
            # Hospital price is at or above band C - keep hospital price but flag
            return (hospital_price, "CHECK IF THIS PRICE IS WELL PRICED")
    
    def negotiate_hospital(self, hospital_name: str):
        """
        Generate negotiated tariff for a single hospital
        
        Parameters:
        -----------
        hospital_name : str
            Name of the hospital
        
        Returns:
        --------
        DataFrame with negotiated tariff or None if failed
        """
        print(f"\n📋 Processing: {hospital_name}")
        
        # Get hospital tariff from database
        hospital_df = self._get_hospital_tariff(hospital_name)
        
        if hospital_df is None or len(hospital_df) == 0:
            return None
        
        print(f"   Found {len(hospital_df)} procedures")
        
        # Merge with reality tariff (inner join - only keep procedures in reality tariff)
        merged = self.reality_df.merge(
            hospital_df,
            on='procedurecode',
            how='inner'
        )
        
        if len(merged) == 0:
            print(f"   ⚠️  No matching procedures with reality tariff")
            return None
        
        print(f"   Matched {len(merged)} procedures with reality tariff")
        
        # Calculate proposed prices
        results = []
        for _, row in merged.iterrows():
            proposed_price, comment = self._calculate_proposed_price(
                row['hospital_price'],
                row['band_d'],
                row['band_c']
            )
            
            results.append({
                'Procedure Name': row['procedure_name'],
                'Procedure Code': row['procedurecode'].upper(),
                'Hospital Tariff Price': row['hospital_price'],
                'Clearline Proposed Price': proposed_price,
                'Comment': comment
            })
        
        result_df = pd.DataFrame(results)
        
        # Sort by procedure code
        result_df = result_df.sort_values('Procedure Code')
        
        print(f"   ✅ Generated negotiated tariff with {len(result_df)} procedures")
        
        return result_df
    
    def process_hospital_list(self, hospital_list_path: str = None, output_path: str = None):
        """
        Process all hospitals in the list and generate output (Excel or Google Sheets)
        
        Parameters:
        -----------
        hospital_list_path : str, optional
            Path to CSV file with hospital names (if not using Google Sheets)
        output_path : str, optional
            Path for output Excel file (if not using Google Sheets)
        """
        try:
            # Get provider list
            if self.use_google_sheets:
                hospital_names = self._get_providers_from_google_sheets()
                if hospital_names is None:
                    print("❌ Failed to get providers from Google Sheets")
                    sys.exit(1)
            else:
                # Load hospital list from CSV
                if hospital_list_path is None:
                    raise ValueError("hospital_list_path is required when not using Google Sheets")
                
                hospitals_df = pd.read_csv(hospital_list_path)
                
                # Find the hospital name column
                name_col = None
                for col in ['NAME OF PROVIDERS', 'name of providers', 'hospital_name', 'providername', 'name']:
                    if col in hospitals_df.columns:
                        name_col = col
                        break
                
                if name_col is None:
                    raise ValueError(f"Could not find hospital name column. Available columns: {list(hospitals_df.columns)}")
                
                hospital_names = hospitals_df[name_col].dropna().unique().tolist()
            
            print("="*80)
            print(f"HOSPITAL TARIFF NEGOTIATION")
            print("="*80)
            print(f"Hospitals to process: {len(hospital_names)}")
            print(f"Reality tariff procedures: {len(self.reality_df)}")
            if self.use_google_sheets:
                print(f"Output: Google Sheets ({self.spreadsheet_name})")
            else:
                print(f"Output file: {output_path}")
            print("="*80)
            
            # Create Excel workbook only if not using Google Sheets
            wb = None
            if not self.use_google_sheets:
                wb = Workbook()
                # Remove default sheet
                if 'Sheet' in wb.sheetnames:
                    wb.remove(wb['Sheet'])
            
            # Process each hospital
            successful = 0
            failed = 0
            
            for hospital_name in hospital_names:
                hospital_name = str(hospital_name).strip()
                
                if not hospital_name:
                    continue
                
                # Generate negotiated tariff
                negotiated_df = self.negotiate_hospital(hospital_name)
                
                if negotiated_df is not None and len(negotiated_df) > 0:
                    if self.use_google_sheets:
                        # Create sheet in Google Sheets
                        try:
                            if self._create_provider_sheet_in_google_sheets(hospital_name, negotiated_df):
                                # Update analysis status
                                self._update_analysis_status(hospital_name, "analysed")
                                # Small delay to avoid hitting Google API quota limits
                                time.sleep(1.0)
                                successful += 1
                            else:
                                failed += 1
                        except gspread.exceptions.APIError as e:
                            if '429' in str(e):
                                print(f"   ⚠️  Quota limit reached for {hospital_name}. Will need to retry later.")
                                failed += 1
                            else:
                                print(f"   ❌ Error creating sheet for {hospital_name}: {e}")
                                failed += 1
                        except Exception as e:
                            print(f"   ❌ Error creating sheet for {hospital_name}: {e}")
                            failed += 1
                    else:
                        # Create worksheet in Excel
                        # Sanitize sheet name (max 31 chars, no special characters)
                        sheet_name = hospital_name[:31].replace('/', '-').replace('\\', '-').replace(':', '-')
                        ws = wb.create_sheet(title=sheet_name)
                        
                        # Write data to worksheet
                        for r_idx, row in enumerate(dataframe_to_rows(negotiated_df, index=False, header=True), 1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                                
                                # Style header row
                                if r_idx == 1:
                                    cell.font = Font(bold=True, color="FFFFFF")
                                    cell.fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
                                    cell.alignment = Alignment(horizontal="center", vertical="center")
                                
                                # Highlight rows that need price review
                                elif r_idx > 1 and c_idx == 5:  # Comment column
                                    if value and "CHECK IF THIS PRICE IS WELL PRICED" in str(value):
                                        for col in range(1, 6):
                                            ws.cell(row=r_idx, column=col).fill = PatternFill(
                                                start_color="FFF4CC", 
                                                end_color="FFF4CC", 
                                                fill_type="solid"
                                            )
                        
                        # Adjust column widths
                        ws.column_dimensions['A'].width = 40  # Procedure Name
                        ws.column_dimensions['B'].width = 15  # Procedure Code
                        ws.column_dimensions['C'].width = 20  # Hospital Price
                        ws.column_dimensions['D'].width = 25  # Proposed Price
                        ws.column_dimensions['E'].width = 35  # Comment
                        
                        successful += 1
                else:
                    failed += 1
            
            # Save workbook (only if using Excel)
            if not self.use_google_sheets and wb is not None:
                if successful > 0:
                    if output_path is None:
                        raise ValueError("output_path is required when not using Google Sheets")
                    wb.save(output_path)
                    print("\n" + "="*80)
                    print(f"✅ SUCCESS!")
                    print("="*80)
                    print(f"Processed: {successful + failed} hospitals")
                    print(f"Successful: {successful} hospitals")
                    print(f"Failed: {failed} hospitals")
                    print(f"Output saved: {output_path}")
                    print("="*80)
                else:
                    print("\n❌ No successful negotiations - no output file created")
            else:
                # Google Sheets mode
                print("\n" + "="*80)
                print(f"✅ SUCCESS!")
                print("="*80)
                print(f"Processed: {successful + failed} hospitals")
                print(f"Successful: {successful} hospitals")
                print(f"Failed: {failed} hospitals")
                print(f"Results saved to Google Sheets: {self.spreadsheet_name}")
                print("="*80)
            
        except Exception as e:
            print(f"❌ Error processing hospital list: {e}")
            import traceback
            traceback.print_exc()
            sys.exit(1)
    
    def close(self):
        """Close database connection"""
        if self.conn:
            self.conn.close()


def main():
    parser = argparse.ArgumentParser(
        description='Hospital Tariff Negotiation Tool - Generate negotiated tariffs based on reality tariff bands',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Example (CSV mode):
  python REALITY_TARIFF_ANALYSIS.py \\
    --reality reality_tariff.csv \\
    --hospitals hospital_list.csv \\
    --output negotiated_tariffs.xlsx

Example (Google Sheets mode):
  python REALITY_TARIFF_ANALYSIS.py \\
    --reality reality_tariff.csv \\
    --google-sheets \\
    --spreadsheet-name "TARIFF EXISTING" \\
    --google-creds CREDENTIALS.json

Hospital List CSV Format:
  - Must contain a column named "NAME OF PROVIDERS" with hospital names
  - One hospital per row
  - Hospital names should match database exactly

Google Sheets Format:
  - Spreadsheet must have a sheet with the same name as the spreadsheet (e.g., "TARIFF EXISTING")
  - Must have columns: NAME OF PROVIDERS, ANALYSIS, STATUS, DOCTOR REVIEWED
  - Providers with STATUS="DONE" or ANALYSIS="analysed" will be skipped
  - Results are written to separate sheets (one per provider) in the same spreadsheet
  - ANALYSIS column is updated to "analysed" after processing

Negotiation Rules:
  1. If hospital price < Band D: Use Band D (increase to minimum)
  2. If hospital price between Band D and C: Use Band C (standard adjustment)
  3. If hospital price >= Band C: Keep hospital price (flag for review)

Output:
  - Excel file with one sheet per hospital (CSV mode)
  - OR Google Sheets with one sheet per provider (Google Sheets mode)
  - Columns: Procedure Name, Procedure Code, Hospital Tariff Price, Clearline Proposed Price, Comment
  - Procedures not in reality tariff are excluded
  - Rows needing review are highlighted in yellow
        """
    )
    
    parser.add_argument(
        '--reality', '-r',
        default=None,
        help='Path to reality tariff CSV file (will auto-detect if not specified)'
    )
    
    parser.add_argument(
        '--hospitals', '-H',
        default=None,
        help='Path to CSV file with hospital list (required if not using --google-sheets)'
    )
    
    parser.add_argument(
        '--output', '-o',
        default=None,
        help='Path for output Excel file (required if not using --google-sheets)'
    )
    
    parser.add_argument(
        '--db',
        default=None,
        help='Path to DuckDB database (will auto-detect if not specified)'
    )
    
    parser.add_argument(
        '--google-sheets',
        action='store_true',
        help='Use Google Sheets for provider list and output (instead of CSV/Excel)'
    )
    
    parser.add_argument(
        '--spreadsheet-name',
        default='TARIFF EXISTING',
        help='Name of the Google Spreadsheet to use. The sheet with the same name will be used for provider tracking. (default: "TARIFF EXISTING")'
    )
    
    parser.add_argument(
        '--google-creds',
        default=None,
        help='Path to Google service account JSON credentials file (will auto-detect if not specified)'
    )
    
    args = parser.parse_args()
    
    # Auto-detect reality tariff file if not provided
    if args.reality is None:
        possible_reality_files = [
            'REALITY TARIFF_Sheet1.csv',
            'REALITY_TARIFF_Sheet1.csv',
            'reality_tariff.csv',
            'reality tariff.csv',
        ]
        for file in possible_reality_files:
            if Path(file).exists():
                args.reality = file
                print(f"✅ Auto-detected reality tariff: {file}")
                break
        
        if args.reality is None:
            print("❌ Could not find reality tariff file. Please specify with --reality")
            print("   Looking for files like: REALITY TARIFF_Sheet1.csv, reality_tariff.csv")
            sys.exit(1)
    
    # Validate input files
    if not Path(args.reality).exists():
        print(f"❌ Reality tariff file not found: {args.reality}")
        sys.exit(1)
    
    # Auto-enable Google Sheets mode if no CSV arguments provided
    if not args.google_sheets and args.hospitals is None and args.output is None:
        if GSPREAD_AVAILABLE:
            args.google_sheets = True
            print("✅ Auto-enabling Google Sheets mode (no CSV arguments provided)")
        else:
            print("❌ No mode specified. Either:")
            print("   1. Use --google-sheets (requires gspread)")
            print("   2. Provide --hospitals and --output for CSV mode")
            sys.exit(1)
    
    # Validate mode-specific requirements
    if args.google_sheets:
        if not GSPREAD_AVAILABLE:
            print("❌ gspread not available. Install with: pip install gspread google-auth")
            sys.exit(1)
    else:
        if args.hospitals is None:
            print("❌ --hospitals is required when not using --google-sheets")
            sys.exit(1)
        if args.output is None:
            print("❌ --output is required when not using --google-sheets")
            sys.exit(1)
        if not Path(args.hospitals).exists():
            print(f"❌ Hospital list file not found: {args.hospitals}")
            sys.exit(1)
    
    # Initialize negotiator
    negotiator = HospitalTariffNegotiator(
        args.reality, 
        args.db,
        use_google_sheets=args.google_sheets,
        google_credentials_path=args.google_creds,
        spreadsheet_name=args.spreadsheet_name
    )
    
    try:
        # Process all hospitals
        negotiator.process_hospital_list(args.hospitals, args.output)
    finally:
        negotiator.close()


if __name__ == '__main__':
    main()