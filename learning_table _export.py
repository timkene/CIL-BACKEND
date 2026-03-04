#!/usr/bin/env python3
"""
Learning Knowledge Export & Backup
====================================

Exports all AI learning tables to a timestamped Excel workbook.
Each table gets its own sheet with formatted headers and summary stats.

Usage:
    python learning_export.py                    # Export to default location
    python learning_export.py --output /path/    # Export to specific folder
    python learning_export.py --db /path/db      # Use specific database

Author: Casey's AI Assistant
Date: February 2026
"""

import duckdb
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import argparse
import os
import sys

# Learning tables and their display names
LEARNING_TABLES = {
    'ai_human_procedure_age': {
        'display_name': 'Procedure Age',
        'description': 'Learned procedure-age validity (e.g., Amlodipine valid for ADULT 18-64)',
        'key_columns': ['procedure_code', 'min_age', 'max_age'],
        'decision_column': 'is_valid_for_age'
    },
    'ai_human_procedure_gender': {
        'display_name': 'Procedure Gender',
        'description': 'Learned procedure-gender validity (e.g., Amlodipine valid for Male)',
        'key_columns': ['procedure_code', 'gender'],
        'decision_column': 'is_valid_for_gender'
    },
    'ai_human_diagnosis_age': {
        'display_name': 'Diagnosis Age',
        'description': 'Learned diagnosis-age validity (e.g., Ovarian cancer invalid for age 0-1)',
        'key_columns': ['diagnosis_code', 'min_age', 'max_age'],
        'decision_column': 'is_valid_for_age'
    },
    'ai_human_diagnosis_gender': {
        'display_name': 'Diagnosis Gender',
        'description': 'Learned diagnosis-gender validity (e.g., Ovarian cancer invalid for Male)',
        'key_columns': ['diagnosis_code', 'gender'],
        'decision_column': 'is_valid_for_gender'
    },
    'ai_human_procedure_diagnosis': {
        'display_name': 'Procedure-Diagnosis',
        'description': 'Learned procedure-diagnosis compatibility (e.g., Amlodipine invalid for cancer)',
        'key_columns': ['procedure_code', 'diagnosis_code'],
        'decision_column': 'is_valid_match'
    },
    'ai_human_procedure_class': {
        'display_name': 'Procedure Class (30-Day)',
        'description': 'Learned therapeutic class relationships for 30-day duplicate checking',
        'key_columns': ['procedure_code_1', 'procedure_code_2'],
        'decision_column': 'same_class'
    }
}

# Styling constants
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(bold=True, color='FFFFFF', size=11, name='Arial')
VALID_FILL = PatternFill('solid', fgColor='C6EFCE')
INVALID_FILL = PatternFill('solid', fgColor='FFC7CE')
DATA_FONT = Font(size=10, name='Arial')
TITLE_FONT = Font(bold=True, size=14, name='Arial', color='1F4E79')
SUBTITLE_FONT = Font(bold=True, size=11, name='Arial', color='333333')
THIN_BORDER = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9')
)


def export_learning_tables(db_path: str = "ai_driven_data.duckdb", output_dir: str = ".") -> str:
    """
    Export all learning tables to a timestamped Excel workbook.
    
    Returns the path to the exported file.
    """
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f"learning_backup_{timestamp}.xlsx"
    filepath = os.path.join(output_dir, filename)
    
    conn = duckdb.connect(db_path, read_only=True)
    
    table_stats = {}
    
    # First pass: export raw data with pandas
    with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
        # Summary sheet placeholder (we'll format it after)
        summary_data = []
        
        for table_name, config in LEARNING_TABLES.items():
            try:
                df = conn.execute(
                    f'SELECT * FROM "PROCEDURE_DIAGNOSIS"."{table_name}" ORDER BY usage_count DESC'
                ).fetchdf()
                
                sheet_name = config['display_name'][:31]  # Excel 31 char limit
                
                # Write data starting at row 4 (leave room for title/description)
                df.to_excel(writer, sheet_name=sheet_name, index=False, startrow=3)
                
                # Collect stats
                decision_col = config['decision_column']
                valid_count = 0
                invalid_count = 0
                if decision_col in df.columns and len(df) > 0:
                    valid_count = int(df[decision_col].sum()) if df[decision_col].dtype == bool else len(df[df[decision_col] == True])
                    invalid_count = len(df) - valid_count
                
                total_usage = int(df['usage_count'].sum()) if 'usage_count' in df.columns and len(df) > 0 else 0
                
                table_stats[table_name] = {
                    'display_name': config['display_name'],
                    'description': config['description'],
                    'total_rows': len(df),
                    'valid_count': valid_count,
                    'invalid_count': invalid_count,
                    'total_usage': total_usage,
                    'has_data': len(df) > 0
                }
                
                summary_data.append({
                    'Table': config['display_name'],
                    'Total Entries': len(df),
                    'Valid/Same': valid_count,
                    'Invalid/Different': invalid_count,
                    'Total Times Used': total_usage,
                    'Description': config['description']
                })
                
            except Exception as e:
                print(f"⚠️  Could not export {table_name}: {e}")
                table_stats[table_name] = {
                    'display_name': config['display_name'],
                    'description': config['description'],
                    'total_rows': 0,
                    'valid_count': 0,
                    'invalid_count': 0,
                    'total_usage': 0,
                    'has_data': False,
                    'error': str(e)
                }
                summary_data.append({
                    'Table': config['display_name'],
                    'Total Entries': 0,
                    'Valid/Same': 0,
                    'Invalid/Different': 0,
                    'Total Times Used': 0,
                    'Description': f'ERROR: {e}'
                })
        
        # Write summary sheet
        summary_df = pd.DataFrame(summary_data)
        summary_df.to_excel(writer, sheet_name='Summary', index=False, startrow=5)
    
    conn.close()
    
    # Second pass: format with openpyxl
    wb = load_workbook(filepath)
    
    # Format Summary sheet
    ws = wb['Summary']
    ws.insert_rows(1, 5)  # Already have 5 rows offset
    
    ws['A1'] = 'CLEARLINE INTERNATIONAL LIMITED'
    ws['A1'].font = TITLE_FONT
    ws['A2'] = 'AI Learning Knowledge Base — Export & Backup'
    ws['A2'].font = SUBTITLE_FONT
    ws['A3'] = f'Exported: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A3'].font = Font(italic=True, size=10, name='Arial', color='666666')
    
    total_entries = sum(s['total_rows'] for s in table_stats.values())
    total_usage = sum(s['total_usage'] for s in table_stats.values())
    ws['A4'] = f'Total learning entries: {total_entries} | Total times used: {total_usage} (AI calls saved)'
    ws['A4'].font = Font(size=10, name='Arial', color='1F4E79')
    
    # Format summary headers (row 6, since we wrote at startrow=5 + inserted 5)
    header_row = 6
    for col_idx in range(1, len(summary_df.columns) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Format summary data rows
    for row_idx in range(header_row + 1, header_row + 1 + len(summary_df)):
        for col_idx in range(1, len(summary_df.columns) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = DATA_FONT
            cell.border = THIN_BORDER
            if col_idx in [2, 3, 4, 5]:  # Numeric columns
                cell.alignment = Alignment(horizontal='center')
    
    # Auto-width for summary
    for col_idx in range(1, len(summary_df.columns) + 1):
        max_len = max(
            len(str(ws.cell(row=r, column=col_idx).value or ''))
            for r in range(header_row, header_row + 1 + len(summary_df))
        )
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)
    
    # Format each data sheet
    for table_name, config in LEARNING_TABLES.items():
        sheet_name = config['display_name'][:31]
        if sheet_name not in wb.sheetnames:
            continue
        
        ws = wb[sheet_name]
        stats = table_stats.get(table_name, {})
        
        # Title and description in rows 1-3
        ws['A1'] = config['display_name']
        ws['A1'].font = TITLE_FONT
        ws['A2'] = config['description']
        ws['A2'].font = Font(italic=True, size=10, name='Arial', color='666666')
        
        row_count = stats.get('total_rows', 0)
        valid = stats.get('valid_count', 0)
        invalid = stats.get('invalid_count', 0)
        usage = stats.get('total_usage', 0)
        ws['A3'] = f'Entries: {row_count} | Valid: {valid} | Invalid: {invalid} | Times used: {usage}'
        ws['A3'].font = Font(size=10, name='Arial', color='1F4E79')
        
        if not stats.get('has_data', False):
            continue
        
        # Format headers (row 4)
        header_row = 4
        for col_idx in range(1, ws.max_column + 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', wrap_text=True)
        
        # Find decision column index
        decision_col_name = config['decision_column']
        decision_col_idx = None
        for col_idx in range(1, ws.max_column + 1):
            if ws.cell(row=header_row, column=col_idx).value == decision_col_name:
                decision_col_idx = col_idx
                break
        
        # Format data rows
        for row_idx in range(header_row + 1, ws.max_row + 1):
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.font = DATA_FONT
                cell.border = THIN_BORDER
            
            # Highlight decision column
            if decision_col_idx:
                cell = ws.cell(row=row_idx, column=decision_col_idx)
                val = cell.value
                if val is True or val == 'True' or val == 1:
                    cell.fill = VALID_FILL
                    cell.value = '✅ VALID'
                elif val is False or val == 'False' or val == 0:
                    cell.fill = INVALID_FILL
                    cell.value = '❌ INVALID'
        
        # Auto-width columns
        for col_idx in range(1, ws.max_column + 1):
            max_len = 0
            for row_idx in range(header_row, min(ws.max_row + 1, header_row + 50)):
                val = ws.cell(row=row_idx, column=col_idx).value
                if val:
                    max_len = max(max_len, len(str(val)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 3, 40)
        
        # Freeze panes (header row)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    
    # Move Summary to first position
    wb.move_sheet('Summary', offset=-len(wb.sheetnames) + 1)
    
    # Freeze panes on summary too
    ws_summary = wb['Summary']
    ws_summary.freeze_panes = ws_summary.cell(row=header_row + 1, column=1)
    
    wb.save(filepath)
    
    return filepath, table_stats


def print_report(filepath: str, stats: dict):
    """Print export summary to console."""
    print("\n" + "=" * 60)
    print("📦 LEARNING KNOWLEDGE EXPORT COMPLETE")
    print("=" * 60)
    print(f"📁 File: {filepath}")
    print(f"📅 Date: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    total_entries = 0
    total_usage = 0
    
    for table_name, s in stats.items():
        icon = "✅" if s.get('has_data') else "⚪"
        rows = s.get('total_rows', 0)
        usage = s.get('total_usage', 0)
        total_entries += rows
        total_usage += usage
        
        print(f"  {icon} {s['display_name']}: {rows} entries ({usage} times used)")
    
    print()
    print(f"  📊 TOTAL: {total_entries} entries | {total_usage} AI calls saved")
    print("=" * 60)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description='Export AI learning tables to Excel')
    parser.add_argument('--db', default='ai_driven_data.duckdb', help='Database path')
    parser.add_argument('--output', default='.', help='Output directory')
    args = parser.parse_args()
    
    if not os.path.exists(args.db):
        print(f"❌ Database not found: {args.db}")
        sys.exit(1)
    
    os.makedirs(args.output, exist_ok=True)
    
    filepath, stats = export_learning_tables(args.db, args.output)
    print_report(filepath, stats)