#!/usr/bin/env python3
"""
Tariff Negotiation Review System
=================================

Automated tariff negotiation with dual pricing sources:
1. Reality Tariff (primary reference)
2. Claims Quartile Clusters (fallback for unlisted procedures)

Author: Casey's AI Assistant
Date: December 2025
Version: 1.0

Features:
- Multi-hospital Excel upload (one sheet per hospital)
- First Counter: Accept if cheaper, use reality/Q1 if expensive
- Second Counter: Accept 0-15% increase, use Q2 for unlisted
- Quartile clustering from DuckDB claims data
- Download reviewed tariffs with changes highlighted
"""


import pandas as pd
import numpy as np
from pathlib import Path
from datetime import datetime
import io
import warnings
warnings.filterwarnings('ignore')

try:
    import duckdb
    DUCKDB_AVAILABLE = True
except ImportError:
    DUCKDB_AVAILABLE = False
    st.error("⚠️ DuckDB not available. Install with: pip install duckdb")

try:
    from openpyxl import load_workbook, Workbook
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    st.error("⚠️ openpyxl not available. Install with: pip install openpyxl")


# ============================================================================
# CONFIGURATION
# ============================================================================

REALITY_TARIFF_PATH = "/Users/kenechukwuchukwuka/Downloads/DLT/REALITY TARIFF_Sheet1.csv"
DUCKDB_PATH = "/Users/kenechukwuchukwuka/Downloads/DLT/ai_driven_data.duckdb"
SECOND_COUNTER_THRESHOLD = 0.15  # 15% increase tolerance


# ============================================================================
# DATA LOADING FUNCTIONS
# ============================================================================

@st.cache_data(ttl=3600)
def load_reality_tariff():
    """Load reality tariff as primary reference"""
    try:
        df = pd.read_csv(REALITY_TARIFF_PATH)

        # Normalize column names
        df.columns = df.columns.str.lower().str.strip()

        # Check if we have band_d column (use as tariffamount)
        if 'procedurecode' not in df.columns:
            st.error(f"Reality tariff must have 'procedurecode' column. Found: {list(df.columns)}")
            return pd.DataFrame()

        if 'band_d' not in df.columns:
            st.error(f"Reality tariff must have 'band_d' column. Found: {list(df.columns)}")
            return pd.DataFrame()

        # Use band_d as the tariffamount reference
        df['tariffamount'] = pd.to_numeric(df['band_d'], errors='coerce')

        # Normalize codes
        df['procedurecode'] = df['procedurecode'].astype(str).str.strip().str.lower()

        # Remove invalid rows
        df = df.dropna(subset=['procedurecode', 'tariffamount'])
        df = df[df['tariffamount'] > 0]
        
        st.success(f"✅ Loaded {len(df)} procedures from Reality Tariff")
        return df
        
    except FileNotFoundError:
        st.error(f"❌ Reality tariff file not found: {REALITY_TARIFF_PATH}")
        return pd.DataFrame()
    except Exception as e:
        st.error(f"❌ Error loading reality tariff: {str(e)}")
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def load_claims_quartiles():
    """
    Load claims data and calculate 4 quartile clusters per procedure
    
    Returns:
    --------
    DataFrame with columns: procedurecode, q1, q2, q3, q4, mean, count
    """
    if not DUCKDB_AVAILABLE:
        st.warning("⚠️ DuckDB not available - claims fallback pricing disabled")
        return pd.DataFrame()
    
    try:
        if not Path(DUCKDB_PATH).exists():
            st.warning(f"⚠️ DuckDB file not found: {DUCKDB_PATH}")
            return pd.DataFrame()
        
        conn = duckdb.connect(DUCKDB_PATH, read_only=True)
        
        query = """
        WITH cleaned_claims AS (
            SELECT 
                LOWER(TRIM(code)) as procedurecode,
                chargeamount
            FROM "AI DRIVEN DATA"."CLAIMS DATA"
            WHERE code IS NOT NULL 
                AND chargeamount > 0
                AND chargeamount IS NOT NULL
        ),
        percentiles AS (
            SELECT 
                procedurecode,
                APPROX_QUANTILE(chargeamount, 0.01) as p01,
                APPROX_QUANTILE(chargeamount, 0.25) as q1,
                APPROX_QUANTILE(chargeamount, 0.50) as q2,
                APPROX_QUANTILE(chargeamount, 0.75) as q3,
                APPROX_QUANTILE(chargeamount, 0.99) as p99,
                AVG(chargeamount) as mean,
                COUNT(*) as count
            FROM cleaned_claims
            GROUP BY procedurecode
        ),
        trimmed_stats AS (
            SELECT 
                c.procedurecode,
                AVG(CASE WHEN c.chargeamount <= p.q1 THEN c.chargeamount END) as q1_cluster,
                AVG(CASE WHEN c.chargeamount > p.q1 AND c.chargeamount <= p.q2 THEN c.chargeamount END) as q2_cluster,
                AVG(CASE WHEN c.chargeamount > p.q2 AND c.chargeamount <= p.q3 THEN c.chargeamount END) as q3_cluster,
                AVG(CASE WHEN c.chargeamount > p.q3 THEN c.chargeamount END) as q4_cluster,
                p.mean,
                p.count
            FROM cleaned_claims c
            JOIN percentiles p ON c.procedurecode = p.procedurecode
            WHERE c.chargeamount >= p.p01 
                AND c.chargeamount <= p.p99
            GROUP BY c.procedurecode, p.mean, p.count
        )
        SELECT 
            procedurecode,
            COALESCE(q1_cluster, mean) as q1,
            COALESCE(q2_cluster, mean) as q2,
            COALESCE(q3_cluster, mean) as q3,
            COALESCE(q4_cluster, mean) as q4,
            mean,
            count
        FROM trimmed_stats
        WHERE count >= 10
        ORDER BY count DESC
        """
        
        df = conn.execute(query).fetchdf()
        conn.close()
        
        st.success(f"✅ Loaded claims quartiles for {len(df)} procedures")
        return df
        
    except Exception as e:
        st.warning(f"⚠️ Could not load claims data: {str(e)}")
        return pd.DataFrame()


# ============================================================================
# NEGOTIATION LOGIC
# ============================================================================

def normalize_code(value) -> str:
    """Normalize procedure codes for matching"""
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def get_reference_price(code: str, stage: str, reality_dict: dict, claims_dict: dict) -> tuple:
    """
    Get reference price based on stage and availability
    
    Parameters:
    -----------
    code : str
        Procedure code (normalized)
    stage : str
        "first" or "second" counter
    reality_dict : dict
        Reality tariff lookup {code: price}
    claims_dict : dict
        Claims quartile lookup {code: {q1, q2, q3, q4}}
    
    Returns:
    --------
    tuple: (reference_price, source)
        source: "reality", "claims_q1", "claims_q2", or "not_found"
    """
    # Try reality tariff first
    if code in reality_dict:
        return reality_dict[code], "reality"
    
    # Fallback to claims quartiles
    if code in claims_dict:
        claims_data = claims_dict[code]
        if stage == "first":
            # Use Q1 cluster for first counter
            return claims_data['q1'], "claims_q1"
        else:  # stage == "second"
            # Use Q2 cluster for second counter
            return claims_data['q2'], "claims_q2"
    
    return None, "not_found"


def process_first_counter(row: pd.Series, reality_dict: dict, claims_dict: dict) -> dict:
    """
    Process CLEARLINE FIRST COUNTER based on HOSPITAL FIRST COUNTER
    
    Rules:
    - If HOSPITAL FIRST COUNTER < reference: Accept hospital counter
    - If HOSPITAL FIRST COUNTER > reference: Use reference price
    - If HOSPITAL FIRST COUNTER is empty: They accepted FIRST PRICE, leave blank
    - Reference = Reality tariff (or Q1 claims if not in reality)
    """
    code = normalize_code(row.get('code', ''))
    hospital_counter = row.get('hospital first counter', None)
    first_price = row.get('first price', None)
    
    # If hospital counter is empty/NaN, they accepted our first price
    if pd.isna(hospital_counter) or hospital_counter == '':
        return {
            'clearline_counter': '',
            'reference_price': None,
            'source': 'accepted',
            'action': 'Hospital accepted FIRST PRICE',
            'difference': 0
        }
    
    # Convert to numeric
    try:
        hospital_counter = float(hospital_counter)
    except (ValueError, TypeError):
        return {
            'clearline_counter': '',
            'reference_price': None,
            'source': 'error',
            'action': 'Invalid hospital counter value',
            'difference': 0
        }
    
    # Get reference price
    reference_price, source = get_reference_price(code, "first", reality_dict, claims_dict)
    
    if reference_price is None:
        # No reference available - accept hospital counter cautiously
        return {
            'clearline_counter': hospital_counter,
            'reference_price': None,
            'source': 'no_reference',
            'action': 'Accepted (no reference data)',
            'difference': 0
        }
    
    # Apply negotiation logic
    if hospital_counter <= reference_price:
        # Hospital counter is cheaper - accept it
        clearline_counter = hospital_counter
        action = f"Accepted (cheaper than {source})"
    else:
        # Hospital counter is expensive - counter with reference
        clearline_counter = reference_price
        action = f"Countered with {source} price"
    
    difference = hospital_counter - clearline_counter
    
    return {
        'clearline_counter': clearline_counter,
        'reference_price': reference_price,
        'source': source,
        'action': action,
        'difference': difference
    }


def process_second_counter(row: pd.Series, reality_dict: dict, claims_dict: dict) -> dict:
    """
    Process CLEARLINE SECOND COUNTER based on HOSPITAL SECOND COUNTER
    
    Rules:
    - Accept prices within 0-15% increase of reference price
    - Reference = Reality tariff (or Q2 claims if not in reality)
    - If hospital counter is empty, they accepted our first counter
    """
    code = normalize_code(row.get('code', ''))
    hospital_counter = row.get('hospital second counter', None)
    clearline_first = row.get('clearline first counter', None)
    
    # If hospital counter is empty/NaN, they accepted our first counter
    if pd.isna(hospital_counter) or hospital_counter == '':
        return {
            'clearline_counter': '',
            'reference_price': None,
            'source': 'accepted',
            'action': 'Hospital accepted CLEARLINE FIRST COUNTER',
            'difference': 0,
            'within_threshold': True
        }
    
    # Convert to numeric
    try:
        hospital_counter = float(hospital_counter)
    except (ValueError, TypeError):
        return {
            'clearline_counter': '',
            'reference_price': None,
            'source': 'error',
            'action': 'Invalid hospital counter value',
            'difference': 0,
            'within_threshold': False
        }
    
    # Get reference price (Q2 for second counter)
    reference_price, source = get_reference_price(code, "second", reality_dict, claims_dict)
    
    if reference_price is None:
        # No reference - accept if reasonable
        return {
            'clearline_counter': hospital_counter,
            'reference_price': None,
            'source': 'no_reference',
            'action': 'Accepted (no reference data)',
            'difference': 0,
            'within_threshold': True
        }
    
    # Calculate percentage increase
    max_acceptable = reference_price * (1 + SECOND_COUNTER_THRESHOLD)
    pct_increase = ((hospital_counter - reference_price) / reference_price) * 100
    
    if hospital_counter <= max_acceptable:
        # Within 15% threshold - accept
        clearline_counter = hospital_counter
        action = f"Accepted (+{pct_increase:.1f}% within threshold)"
        within_threshold = True
    else:
        # Exceeds threshold - counter with max acceptable
        clearline_counter = max_acceptable
        action = f"Countered (+{pct_increase:.1f}% exceeds 15%)"
        within_threshold = False
    
    difference = hospital_counter - clearline_counter
    
    return {
        'clearline_counter': clearline_counter,
        'reference_price': reference_price,
        'source': source,
        'action': action,
        'difference': difference,
        'within_threshold': within_threshold
    }


def detect_negotiation_stage(df: pd.DataFrame) -> str:
    """
    Detect which negotiation stage based on columns present
    
    Returns: "first", "second", or "unknown"
    """
    cols_lower = [c.lower().strip() for c in df.columns]
    
    has_hospital_first = 'hospital first counter' in cols_lower
    has_clearline_first = 'clearline first counter' in cols_lower
    has_hospital_second = 'hospital second counter' in cols_lower
    
    if has_hospital_second and has_clearline_first:
        return "second"
    elif has_hospital_first and not has_clearline_first:
        return "first"
    else:
        return "unknown"


def process_hospital_tariff(df: pd.DataFrame, hospital_name: str, 
                           reality_dict: dict, claims_dict: dict) -> pd.DataFrame:
    """
    Process a single hospital's tariff sheet
    
    Returns processed DataFrame with clearline counter and metadata
    """
    # Normalize column names
    df.columns = df.columns.str.lower().str.strip()
    
    # Detect stage
    stage = detect_negotiation_stage(df)
    
    if stage == "unknown":
        st.error(f"❌ {hospital_name}: Cannot detect negotiation stage. Check column names.")
        return df
    
    # Add normalized code column for matching
    if 'code' not in df.columns:
        st.error(f"❌ {hospital_name}: Missing 'code' column")
        return df
    
    df['code_normalized'] = df['code'].apply(normalize_code)
    
    # Process based on stage
    if stage == "first":
        st.info(f"📝 {hospital_name}: Processing FIRST COUNTER")
        
        results = df.apply(
            lambda row: process_first_counter(row, reality_dict, claims_dict),
            axis=1
        )
        
        df['clearline first counter'] = results.apply(lambda x: x['clearline_counter'])
        df['reference_price'] = results.apply(lambda x: x['reference_price'])
        df['source'] = results.apply(lambda x: x['source'])
        df['action'] = results.apply(lambda x: x['action'])
        df['difference'] = results.apply(lambda x: x['difference'])
        
    else:  # stage == "second"
        st.info(f"📝 {hospital_name}: Processing SECOND COUNTER")
        
        results = df.apply(
            lambda row: process_second_counter(row, reality_dict, claims_dict),
            axis=1
        )
        
        df['clearline second counter'] = results.apply(lambda x: x['clearline_counter'])
        df['reference_price'] = results.apply(lambda x: x['reference_price'])
        df['source'] = results.apply(lambda x: x['source'])
        df['action'] = results.apply(lambda x: x['action'])
        df['difference'] = results.apply(lambda x: x['difference'])
        df['within_threshold'] = results.apply(lambda x: x.get('within_threshold', True))
    
    return df


# ============================================================================
# EXCEL OUTPUT FUNCTIONS
# ============================================================================

def create_styled_excel(hospital_results: dict) -> bytes:
    """
    Create styled Excel workbook with one sheet per hospital
    
    Highlights:
    - Green: Accepted hospital counter
    - Yellow: Countered with reference price
    - Red: No reference data available
    """
    if not OPENPYXL_AVAILABLE:
        st.error("❌ openpyxl not available - cannot create styled Excel")
        return None
    
    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Color definitions
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for hospital_name, df in hospital_results.items():
        # Create sheet
        ws = wb.create_sheet(title=hospital_name[:31])  # Excel sheet name limit
        
        # Determine which columns to export
        stage = detect_negotiation_stage(df)
        
        if stage == "first":
            export_cols = ['procedure', 'code', 'first price', 'hospital first counter', 
                          'clearline first counter', 'action', 'source']
        else:  # second
            export_cols = ['procedure', 'code', 'first price', 'hospital first counter',
                          'clearline first counter', 'hospital second counter',
                          'clearline second counter', 'action', 'source']
        
        # Filter to existing columns
        export_cols = [c for c in export_cols if c in df.columns]
        df_export = df[export_cols].copy()
        
        # Format numeric columns with better error handling
        def safe_format_currency(x):
            """Safely format a value as currency, handling various edge cases"""
            if pd.isna(x) or x == '' or str(x).strip() in ['', '-', 'N/A', 'n/a']:
                return ''
            try:
                return f"₦{float(x):,.2f}"
            except (ValueError, TypeError):
                return str(x)  # Return as-is if can't convert

        for col in ['first price', 'hospital first counter', 'clearline first counter',
                    'hospital second counter', 'clearline second counter']:
            if col in df_export.columns:
                df_export[col] = df_export[col].apply(safe_format_currency)
        
        # Write headers
        for col_idx, col_name in enumerate(df_export.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name.upper())
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Write data with conditional formatting
        for row_idx, (_, row) in enumerate(df_export.iterrows(), 2):
            for col_idx, (col_name, value) in enumerate(row.items(), 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Apply color based on action
                if col_name in ['clearline first counter', 'clearline second counter']:
                    action = row.get('action', '')
                    if 'accepted' in action.lower():
                        cell.fill = green_fill
                    elif 'countered' in action.lower():
                        cell.fill = yellow_fill
                    elif 'no reference' in action.lower():
                        cell.fill = red_fill
        
        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()


# ============================================================================
# STREAMLIT UI
# ============================================================================

def main():
    st.set_page_config(
        page_title="Tariff Negotiation Review",
        page_icon="💰",
        layout="wide"
    )
    
    st.title("💰 Tariff Negotiation Review System")
    st.markdown("""
    **Automated tariff counter generation with dual pricing sources:**
    - 🎯 **Primary**: Reality Tariff (standard reference prices)
    - 📊 **Fallback**: Claims Quartile Clusters (for unlisted procedures)
    """)
    
    st.markdown("---")
    
    # Sidebar - Configuration
    with st.sidebar:
        st.header("⚙️ Configuration")
        
        st.subheader("📁 Data Sources")
        
        # Load reality tariff
        if st.button("🔄 Reload Reality Tariff"):
            st.cache_data.clear()
        
        reality_df = load_reality_tariff()
        
        if not reality_df.empty:
            st.success(f"✅ Reality: {len(reality_df)} procedures")
            reality_dict = dict(zip(reality_df['procedurecode'], reality_df['tariffamount']))
        else:
            st.error("❌ Reality tariff not loaded")
            reality_dict = {}
        
        # Load claims quartiles
        claims_df = load_claims_quartiles()
        
        if not claims_df.empty:
            st.success(f"✅ Claims: {len(claims_df)} procedures")
            claims_dict = claims_df.set_index('procedurecode').to_dict('index')
        else:
            st.warning("⚠️ Claims fallback disabled")
            claims_dict = {}
        
        st.markdown("---")
        st.subheader("📊 Negotiation Rules")
        
        with st.expander("First Counter Rules"):
            st.markdown("""
            - ✅ **Accept** if hospital < reference
            - 🔄 **Counter** if hospital > reference
            - ⏭️ **Skip** if hospital accepts first price
            - 📈 **Reference**: Reality tariff or Q1 claims
            """)
        
        with st.expander("Second Counter Rules"):
            st.markdown(f"""
            - ✅ **Accept** if within 0-{SECOND_COUNTER_THRESHOLD*100:.0f}% increase
            - 🔄 **Counter** if exceeds {SECOND_COUNTER_THRESHOLD*100:.0f}%
            - 📈 **Reference**: Reality tariff or Q2 claims
            """)
    
    # Main area - File upload
    st.header("📤 Upload Hospital Tariffs")
    st.info("💡 Upload an Excel file with one sheet per hospital. Each sheet should have columns: Procedure, Code, First Price, Hospital First Counter (and optionally Clearline First Counter, Hospital Second Counter)")
    
    uploaded_file = st.file_uploader(
        "Choose Excel file",
        type=['xlsx', 'xls'],
        help="Upload Excel file with multiple sheets (one per hospital)"
    )
    
    if uploaded_file is not None:
        if not OPENPYXL_AVAILABLE:
            st.error("❌ openpyxl required. Install with: pip install openpyxl")
            return
        
        if reality_dict is None or len(reality_dict) == 0:
            st.error("❌ Reality tariff must be loaded before processing")
            return
        
        try:
            # Load all sheets
            xl_file = pd.ExcelFile(uploaded_file)
            sheet_names = xl_file.sheet_names
            
            st.success(f"✅ Found {len(sheet_names)} hospital sheet(s): {', '.join(sheet_names)}")
            
            # Process each hospital
            st.header("🔄 Processing Hospitals")
            
            hospital_results = {}
            
            with st.spinner("Processing tariffs..."):
                for sheet_name in sheet_names:
                    with st.expander(f"🏥 {sheet_name}", expanded=True):
                        df = pd.read_excel(uploaded_file, sheet_name=sheet_name)
                        
                        # Show preview
                        st.subheader("📋 Input Preview")
                        st.dataframe(df.head(5), use_container_width=True)
                        
                        # Process
                        processed_df = process_hospital_tariff(
                            df, sheet_name, reality_dict, claims_dict
                        )
                        
                        # Store result
                        hospital_results[sheet_name] = processed_df
                        
                        # Show statistics
                        stage = detect_negotiation_stage(processed_df)
                        
                        if stage == "first":
                            counter_col = 'clearline first counter'
                        elif stage == "second":
                            counter_col = 'clearline second counter'
                        else:
                            counter_col = None
                        
                        if counter_col and counter_col in processed_df.columns:
                            total = len(processed_df)
                            accepted = len(processed_df[
                                processed_df['action'].str.contains('accepted', case=False, na=False)
                            ])
                            countered = len(processed_df[
                                processed_df['action'].str.contains('countered', case=False, na=False)
                            ])
                            no_ref = len(processed_df[
                                processed_df['source'] == 'no_reference'
                            ])
                            
                            col1, col2, col3, col4 = st.columns(4)
                            col1.metric("Total Procedures", total)
                            col2.metric("✅ Accepted", accepted)
                            col3.metric("🔄 Countered", countered)
                            col4.metric("⚠️ No Reference", no_ref)
                            
                            # Show preview of results
                            st.subheader("📊 Output Preview")
                            display_cols = ['procedure', 'code', counter_col, 'action', 'source']
                            display_cols = [c for c in display_cols if c in processed_df.columns]
                            st.dataframe(processed_df[display_cols].head(10), use_container_width=True)
            
            # Generate output file
            st.header("⬇️ Download Reviewed Tariffs")
            
            if hospital_results:
                excel_data = create_styled_excel(hospital_results)
                
                if excel_data:
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                    filename = f"reviewed_tariffs_{timestamp}.xlsx"
                    
                    st.download_button(
                        label="📥 Download Reviewed Tariffs (Excel)",
                        data=excel_data,
                        file_name=filename,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                    
                    st.success("""
                    ✅ **Download ready!** 
                    
                    **Color Legend:**
                    - 🟢 **Green**: Accepted hospital counter (cheaper than reference)
                    - 🟡 **Yellow**: Countered with reference price (hospital too expensive)
                    - 🔴 **Red**: No reference data available (accepted cautiously)
                    """)
                else:
                    st.error("❌ Failed to generate Excel file")
        
        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")
            import traceback
            st.code(traceback.format_exc())


if __name__ == "__main__":
    main()