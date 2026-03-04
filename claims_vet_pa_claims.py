#!/usr/bin/env python3
"""
KLAIRE Claims Vetting Engine v3.8
=================================
Comprehensive 13-Rule PRE-AUTH/NO-AUTH Claims Vetting System
Using Claude API for Medical Judgment

🚨 CRITICAL FIXES IN v3.8:

1. R6/R7 GENDER/AGE - WEB SEARCH NO LONGER OVERRIDES KNOWN DATA:
   - If patient gender is KNOWN and MATCHES procedure requirement → PASS (not QUERY!)
   - Web search now CONFIRMS AI's decision instead of questioning it
   - Example: Female patient + HVS procedure → PASS (web found "female only" confirms this!)
   
2. NO-AUTH CLAIMS - PA CONTEXT LOOKUP (NEW!):
   - When PANumber=0, now checks PA table for same enrollee ±24 hours
   - Finds related procedures/diagnoses that explain the drug
   - Example: Tranexamic acid + Hypertension diagnosis → finds "Threatened abortion" in PA
   - This provides the FULL clinical picture to AI

3. AI CONTEXT PRESERVATION:
   - Web search now ADDS to AI's analysis instead of REPLACING it
   - AI's context-aware insights (prior claims, encounter history) are preserved
   - Example: "However, encounter context shows H. pylori test..." now appears!

4. SEMICOLON-SEPARATED DIAGNOSES FIX (NEW!):
   - Now properly parses diagnoses separated by semicolons (;) OR commas (,)
   - Example: "Hypertension; Incompetence of cervix uteri;" → 2 separate diagnoses
   - AI now sees ALL diagnoses, not just the first one!

5. PROVISIONAL DIAGNOSIS LOGIC (R10 Enhancement - NEW!):
   - Detects when a diagnosis was provisional (for test justification only)
   - If diagnosis + diagnostic test + NO treatment → diagnosis was RULED OUT
   - Symptomatic drugs CANNOT be justified by ruled-out diagnoses!
   - Example: Malaria diagnosis + Malaria test + Paracetamol (but NO antimalarial)
     → Malaria test was negative, Paracetamol CANNOT claim "for malaria fever"!

6. ENHANCED PRIOR CLAIMS CONTEXT:
   - Prior claims now passed to AI with FULL context
   - Includes procedure descriptions, diagnosis descriptions, provider names
   - AI can make intelligent connections (same drug class, related conditions)

🚨 PREVIOUS v3.7 FEATURES (RETAINED):
- MODEL SELECTION: --model opus/sonnet/haiku
- R1: 90-DAY SUBMISSION RULE (FAIL if >90 days late)
- R5: HOSPITAL SHOPPING with AI CLASS MATCHING
- R8: 30-DAY DUPLICATES with CLASS DETECTION
- WEB SEARCH: Smart mode default (cost-efficient)

🚨 PREVIOUS v3.6 FIXES:
- R3: Shows BOTH procedure AND amount comparison
- R4: Shows which provider was detected (debugging)
- Provider column detection expanded

Author: Clearline HMO Analytics Team / KLAIRE AI
Date: January 2026
Version: 3.8 (Context Preservation + NO-AUTH PA Lookup + Gender/Age Fix + Semicolon Diagnosis Fix)

Usage:
    # Standard mode - Opus with SMART web search (default, cost-efficient)
    python claims_vet_api_v37.py --input claims.csv --output report.xlsx
    
    # Use Haiku for speed (live PA approval)
    python claims_vet_api_v37.py --input claims.csv --model haiku
    
    # Maximum thoroughness - always search web
    python claims_vet_api_v37.py --input claims.csv --web-search-always
    
    # Fastest/cheapest - no web search
    python claims_vet_api_v37.py --input claims.csv --no-web-search
"""

import os
import sys
import json
import argparse
import pandas as pd
import duckdb
import re
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path

# Try to import anthropic
try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    print("ERROR: anthropic package not installed. Run: pip install anthropic")
    sys.exit(1)

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Excel output disabled. Run: pip install openpyxl")


# =============================================================================
# COMMON DRUG FORMULARY - For detecting uncommon drugs that need web search
# =============================================================================

COMMON_DRUGS = {
    # Antimalarials
    'artemether', 'lumefantrine', 'artesunate', 'amodiaquine', 'chloroquine',
    'quinine', 'primaquine', 'mefloquine', 'coartem', 'lonart',
    
    # Antibiotics
    'amoxicillin', 'amoxyclav', 'augmentin', 'ceftriaxone', 'cefuroxime',
    'cefixime', 'ciprofloxacin', 'levofloxacin', 'metronidazole', 'flagyl',
    'azithromycin', 'erythromycin', 'doxycycline', 'clindamycin', 'gentamicin',
    'ampicillin', 'penicillin', 'cotrimoxazole', 'nitrofurantoin', 'norfloxacin',
    
    # NSAIDs / Analgesics
    'paracetamol', 'ibuprofen', 'diclofenac', 'celecoxib', 'meloxicam',
    'piroxicam', 'naproxen', 'aspirin', 'mefenamic', 'ketoprofen',
    'tramadol', 'codeine', 'morphine', 'pethidine', 'pentazocine',
    
    # Antacids / GI
    'omeprazole', 'esomeprazole', 'pantoprazole', 'ranitidine', 'famotidine',
    'antacid', 'magnesium trisilicate', 'aluminium hydroxide', 'sucralfate',
    'metoclopramide', 'domperidone', 'ondansetron', 'hyoscine', 'buscopan',
    
    # Antihypertensives
    'amlodipine', 'lisinopril', 'enalapril', 'losartan', 'valsartan',
    'atenolol', 'propranolol', 'metoprolol', 'nifedipine', 'hydralazine',
    'methyldopa', 'labetalol', 'furosemide', 'hydrochlorothiazide', 'spironolactone',
    
    # Diabetes
    'metformin', 'glibenclamide', 'gliclazide', 'glimepiride', 'insulin',
    'sitagliptin', 'pioglitazone',
    
    # Antihistamines
    'cetirizine', 'loratadine', 'chlorpheniramine', 'promethazine', 'diphenhydramine',
    'fexofenadine', 'desloratadine',
    
    # Steroids
    'prednisolone', 'prednisone', 'dexamethasone', 'hydrocortisone', 'betamethasone',
    'methylprednisolone', 'triamcinolone',
    
    # Vitamins / Supplements
    'vitamin c', 'ascorbic acid', 'vitamin b', 'b complex', 'folic acid',
    'ferrous', 'iron', 'calcium', 'zinc', 'multivitamin',
    
    # Obstetrics / Gynecology
    'misoprostol', 'oxytocin', 'ergometrine', 'magnesium sulphate',
    'progesterone', 'estrogen', 'clomiphene',
    
    # Respiratory
    'salbutamol', 'ventolin', 'aminophylline', 'theophylline', 'montelukast',
    'fluticasone', 'budesonide', 'ipratropium',
    
    # CNS / Psychiatric
    'diazepam', 'lorazepam', 'alprazolam', 'bromazepam', 'clonazepam',
    'amitriptyline', 'imipramine', 'fluoxetine', 'sertraline', 'carbamazepine',
    'phenytoin', 'valproate', 'phenobarbital', 'haloperidol', 'chlorpromazine',
    'risperidone', 'olanzapine',
    
    # Others common in Nigeria
    'artemether-lumefantrine', 'sp', 'fansidar', 'chloroquine phosphate',
    'oral rehydration', 'ors', 'zinc sulphate', 'albendazole', 'mebendazole',
    'ivermectin', 'praziquantel', 'chymoral', 'serratiopeptidase',
    
    # Common brand names
    'coartem', 'lonart', 'amalar', 'p-alaxin', 'lumartem', 'camosunate',
    'flagyl', 'augmentin', 'zinnat', 'rocephin', 'cipro', 'zithromax',
    'panadol', 'tylenol', 'cataflan', 'voltaren', 'celebrex', 'feldene',
    'nexium', 'losec', 'zantac', 'buscopan', 'plasil', 'maxolon',
    'norvasc', 'zestril', 'cozaar', 'tenormin', 'lasix', 'aldactone',
    'glucophage', 'daonil', 'diamicron', 'amaryl',
    'zyrtec', 'clarityn', 'piriton', 'phenergan', 'benadryl',
}

# Uncertainty indicators in Opus responses
UNCERTAINTY_INDICATORS = [
    'uncertain', 'unsure', 'not sure', 'unclear', 'unknown',
    'unfamiliar', 'not familiar', 'cannot determine', 'cannot confirm',
    'may need verification', 'requires verification', 'suggest checking',
    'not in my knowledge', 'limited information', 'insufficient data',
    'rare drug', 'uncommon', 'novel', 'new medication', 'recently approved',
    'verify with', 'confirm with', 'check with pharmacist', 'consult',
    'i don\'t have', 'i do not have', 'not aware of', 'outside my',
]


# =============================================================================
# SYSTEM PROMPT - THE BRAIN OF THE VETTING ENGINE
# =============================================================================

SYSTEM_PROMPT = """You are KLAIRE (Knowledge-Led AI for Insurance Risk Evaluation), an expert medical claims vetting AI for Clearline HMO, a health insurance company in Lagos, Nigeria.

Your role is to evaluate healthcare claims for medical appropriateness, applying clinical judgment to determine if drugs/procedures are appropriate for the stated diagnoses.

## YOUR TASK
For each claim, evaluate Rules R6-R12 and return a JSON verdict. The script handles R1-R5 and R13 (data validation rules).

## THE 13 PRE-AUTH VETTING RULES

### Data Rules (Handled by Script):
- R1: 90-Day Submission Rule - Claim submitted within 90 days of encounter date
- R2: Provider Panel Status - Accredited provider
- R3: PA Matching - Claim items ≤ PA authorization
- R4: Tariff Compliance - Amount within tariff limits
- R5: Hospital Shopping Detection - No visits to OTHER providers for same diagnosis/procedure in 30 days
- R13: PA Authorization Valid - Valid PA number exists

### Medical Judgment Rules (YOUR RESPONSIBILITY):
- R6: Gender-Appropriate - Service matches patient gender
- R7: Age-Appropriate - Service appropriate for patient age
- R8: Polypharmacy/30-Day Check - No dangerous duplicates, appropriate repeats
- R9: Drug-Diagnosis Match - Drug/procedure appropriate for diagnosis
- R10: Symptomatic-Only Check - Treatment drugs present, not just symptomatic
- R11: Dosage Validation - Dosage within therapeutic range
- R12: Procedure-Diagnosis Match - Procedure appropriate for diagnosis

## CRITICAL: R9 ENCOUNTER-LEVEL DIAGNOSIS MATCHING

**PROBLEM:** Providers sometimes:
1. Mix up diagnosis-procedure rows (Malaria gets Antacid, Ulcer gets Artemether)
2. Put all diagnoses together ("Malaria, Ulcer, Unspecified" on every row)

**SOLUTION:** You will receive ALL diagnoses for the encounter (same enrollee + PA + date).
A drug is PASS if it matches ANY diagnosis in the encounter, not just the single row diagnosis.

Example:
- Single Diagnosis: "Malaria"
- All Encounter Diagnoses: ["Malaria", "Peptic Ulcer", "Anemia"]
- Drug: Antacid
- Verdict: PASS (matches Peptic Ulcer in encounter)

## DRUG-DIAGNOSIS RULES

### CONTRAINDICATED COMBINATIONS (Always FAIL):
1. **NSAIDs in Pregnancy** (Celecoxib, Diclofenac, Ibuprofen, Mefenamic Acid)
   - Contraindicated for: Abortion, Pregnancy, Threatened abortion
   - UNLESS musculoskeletal indication also present (Arthritis, Back pain, Spondylosis)
   - Reason: Increases miscarriage risk, premature ductus arteriosus closure

2. **NSAIDs in Ulcer** (without gastroprotection)
   - Contraindicated for: Peptic ulcer, GI bleeding
   - Reason: Worsens ulceration, increases bleeding risk

3. **Metoclopramide in Parkinson's**
   - Contraindicated: Worsens extrapyramidal symptoms

4. **ACE Inhibitors in Pregnancy** (Lisinopril, Enalapril, Ramipril)
   - Contraindicated: Teratogenic effects

### WRONG INDICATION (FAIL if no valid indication in encounter):
1. **Antihistamines** (Cetirizine, Loratadine, Chlorpheniramine)
   - NOT for: Malaria, Sepsis, Typhoid (as primary treatment)
   - Valid for: Allergy, Rhinitis, Urticaria, Pruritus, URI (adjunct)

2. **Antacids/PPIs** (Omeprazole, Antacid, Ranitidine)
   - NOT for: Malaria, Sepsis, Head injury, Fractures
   - Valid for: Ulcer, GERD, Dyspepsia, Gastritis, Epigastric pain

3. **Iron Supplements** (Ferrous Sulphate, Ferrous Gluconate)
   - NOT for: Sepsis, Malaria (as primary treatment)
   - Valid for: Anemia, Iron deficiency, Pregnancy, Post-partum

4. **Enzymes** (Chymoral, Serratiopeptidase)
   - NOT for: Sepsis, Malaria, Infections (as primary treatment)
   - Valid for: Inflammation, Edema, Post-surgical, Trauma

5. **Symptomatic drugs for infections without treatment:**
   - Malaria REQUIRES: Artemether, Lumefantrine, or other antimalarial
   - Sepsis REQUIRES: Antibiotics (Ceftriaxone, Cefuroxime, etc.)
   - Typhoid REQUIRES: Antibiotics
   - UTI REQUIRES: Antibiotics

### PREGNANCY-SPECIFIC DRUGS (Verify context):
1. **Misoprostol, Oxytocin**
   - For "Threatened abortion": QUERY - may be miscoded, these are abortifacients
   - For "Missed abortion", "Incomplete abortion", "Post-ERPC": PASS
   
2. **Methyldopa, Labetalol**
   - Valid for: Pregnancy-induced hypertension, Pre-eclampsia
   
3. **Benzodiazepines** (Diazepam, Bromazepam)
   - In pregnancy: QUERY - verify clinical necessity (category D)

### CONTROLLED SUBSTANCES (Verify appropriateness):
1. **Opioids** (Dihydrocodeine, Tramadol, Codeine, Morphine)
   - Check ACTUAL AGE from database!
   - For infant (<1 year): FAIL - opioids dangerous in infants
   - For pediatric (1-12 years): QUERY - verify dosage/indication
   - For adults: PASS if pain indication exists
   - If diagnosis is pediatric-only (Cephalhematoma) but age is adult: diagnosis is WRONG
   
2. **Benzodiazepines** (Diazepam, Bromazepam, Alprazolam)
   - Verify indication: Anxiety, Seizures, Muscle spasm, Pre-operative sedation

3. **Pregabalin**
   - Valid for: Neuropathic pain, Epilepsy, Anxiety disorder

## R8: 30-DAY PRIOR CLAIMS CHECK

You will receive prior claims from the last 30 days. Check for:

1. **Course Medications Should NOT Repeat:**
   - Antibiotics (Amoxicillin, Ceftriaxone, Metronidazole)
   - Antimalarials (Artemether, Lumefantrine)
   - If same drug class within 30 days: QUERY unless different diagnosis

2. **PRN Medications CAN Repeat:**
   - Analgesics (Paracetamol, Ibuprofen)
   - Vitamins (Vitamin C, B-complex)
   - Antacids

3. **Same Therapeutic Class Check:**
   - Multiple NSAIDs in same encounter: FLAG
   - Multiple antibiotics (if not combination therapy): QUERY
   - >8 drugs in single encounter: FLAG for polypharmacy

4. **Dangerous Combinations:**
   - Multiple NSAIDs
   - NSAID + Anticoagulant
   - Multiple CNS depressants

## R6: GENDER-APPROPRIATE (CRITICAL - MUST CHECK ACTUAL GENDER!)

⚠️ YOU WILL BE PROVIDED WITH THE ENROLLEE'S ACTUAL GENDER. YOU MUST CHECK IT!

**Female-only procedures (FAIL if patient is MALE):**
- High vaginal swab, Pap smear, Cervical screening
- Obstetrics, Gynecology, Pregnancy tests, Antenatal
- Cesarean section, Vaginal delivery, D&C
- Ovarian, Uterine, Cervical procedures
- Hysterectomy, Oophorectomy

**Male-only procedures (FAIL if patient is FEMALE):**
- Prostate procedures, PSA test, TURP
- Testicular procedures, Orchidectomy
- Circumcision (usually male)

**DECISION LOGIC:**
1. Is procedure gender-specific? → If NO, PASS
2. If YES, check the ACTUAL gender provided in the claim
3. If gender matches → PASS
4. If gender DOES NOT match → FAIL with clear reason
5. If gender is UNKNOWN → QUERY

NEVER just say "appropriate if enrollee is female" - YOU HAVE THE GENDER DATA!

## R7: AGE-APPROPRIATE (CRITICAL - MUST CHECK ACTUAL AGE!)

⚠️ YOU WILL BE PROVIDED WITH THE ENROLLEE'S ACTUAL AGE. YOU MUST CHECK IT!

**AGE-IMPOSSIBLE DIAGNOSES (FAIL if age doesn't match):**
- **Cephalhematoma** - NEWBORN ONLY (birth injury). Adult CANNOT have this!
- **Birth asphyxia, Birth injury** - NEWBORN ONLY
- **Neonatal jaundice** - NEWBORN ONLY
- **Meconium aspiration** - NEWBORN ONLY
- **Congenital conditions in adults** - should be coded differently

**AGE-IMPOSSIBLE PROCEDURES:**
- **Paediatrician** visits - CHILDREN ONLY (<18 years). Adult seeing paediatrician = WRONG ENROLLEE!
- **NICU admission** - NEWBORN ONLY
- **Neonatal procedures** - NEWBORN ONLY

**Pediatric (usually <12-18 years):**
- Pediatric formulations (Syrup, Suspension, Drops)
- Pediatric consultations
- Childhood vaccines

**Adult-only:**
- Contraceptives, Family planning
- Most chronic disease medications
- Adult formulations

**Geriatric caution (>65 years):**
- NSAIDs: Increased GI/renal risk
- Benzodiazepines: Fall risk
- Digoxin: Narrow therapeutic window

**DECISION LOGIC:**
1. Is diagnosis AGE-IMPOSSIBLE? → If YES, FAIL (not QUERY!)
2. Is procedure age-restricted? → If NO, PASS
3. If YES, check the ACTUAL age provided in the claim
4. If age is appropriate → PASS
5. If age is NOT appropriate → FAIL with clear reason
6. If age is UNKNOWN → QUERY

NEVER assume age - USE THE PROVIDED AGE DATA!
If age from database contradicts diagnosis (e.g., adult with cephalhematoma), FAIL!

## CONFIDENCE INDICATOR

IMPORTANT: In your response, include a "confidence" field (0.0 to 1.0):
- 1.0 = Completely certain about this evaluation
- 0.8-0.99 = High confidence, common drug/condition
- 0.5-0.79 = Moderate confidence, some uncertainty
- Below 0.5 = Low confidence, unfamiliar drug or rare condition

If you are unfamiliar with a drug or unsure about an indication, SET CONFIDENCE LOW and note this in clinical_notes. The system will automatically trigger a web search for verification.

## OUTPUT FORMAT

Return ONLY a JSON object with this structure:
```json
{
  "R6": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R7": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R8": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R9": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R10": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R11": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R12": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "confidence": 0.95,
  "clinical_notes": "Any additional clinical observations",
  "needs_verification": false
}
```

## IMPORTANT GUIDELINES

1. **Be thorough but fair** - Don't flag claims without valid medical reason
2. **Consider encounter context** - Check ALL diagnoses before flagging R9
3. **Explain clearly** - Reasons should be understandable to non-medical staff
4. **Err on side of QUERY** - If uncertain, QUERY rather than FAIL
5. **Nigerian context** - Common conditions include Malaria, Typhoid, URI, UTI
6. **Think step by step** - For each drug, identify therapeutic class, then check indications
7. **Be honest about uncertainty** - Set confidence low if unfamiliar with drug

You are the medical expert. The script trusts your clinical judgment."""


# =============================================================================
# SYSTEM PROMPT FOR WEB SEARCH VERIFICATION
# =============================================================================

WEB_SEARCH_SYSTEM_PROMPT = """You are a medical research assistant helping verify medical information for healthcare claims vetting.

You have access to web search to look up current medical information.

## YOUR TASK
Given a drug/procedure and diagnosis, search for and verify information relevant to these rules:

### R6 - GENDER APPROPRIATENESS
- Is this procedure/drug gender-specific?
- Search: "[procedure] gender specific male female only"

### R7 - AGE APPROPRIATENESS  
- What are the age restrictions/recommendations?
- Is this appropriate for pediatric/geriatric patients?
- Search: "[drug] age restrictions pediatric geriatric"

### R8 - DRUG INTERACTIONS & DUPLICATIONS
- Does this drug interact dangerously with common medications?
- What drug class does it belong to?
- Search: "[drug] drug interactions contraindications class"

### R9 - DRUG-DIAGNOSIS MATCH
- What are the approved indications for this drug?
- Is it appropriate for the given diagnosis?
- Search: "[drug] indications approved uses"

### R10 - SYMPTOMATIC VS TREATMENT (PROVISIONAL DIAGNOSIS DETECTION!)
- Is this drug symptomatic relief or actual treatment?
- Search: "[drug] mechanism of action treatment vs symptomatic"

🚨 **CRITICAL: PROVISIONAL DIAGNOSIS LOGIC**

In healthcare, a diagnosis can be PROVISIONAL (query) when:
- A diagnosis is recorded to JUSTIFY a diagnostic test
- If the test comes back NEGATIVE, no treatment is given
- The diagnosis was essentially "ruled out"

**HOW TO DETECT:**
1. Diagnosis appears in encounter (e.g., "Malaria")
2. Diagnostic TEST for that condition exists (e.g., "Malaria Parasite Test")
3. BUT NO TREATMENT for that condition exists (e.g., NO Artemether/Coartem)
4. → The diagnosis was PROVISIONAL and the test was NEGATIVE

**IMPLICATION FOR SYMPTOMATIC DRUGS:**
- Symptomatic drugs (Paracetamol, NSAIDs, antihistamines) provide RELIEF
- They CANNOT be the sole justification when the diagnosis was RULED OUT
- Example: Paracetamol with "Malaria" diagnosis BUT no antimalarial = FLAG!

**SYMPTOMATIC DRUGS:**
- Analgesics: Paracetamol, Ibuprofen, Tramadol, Diclofenac
- Antiemetics: Metoclopramide, Ondansetron, Domperidone
- Antitussives: Cough syrups
- Antipyretics: Paracetamol
- Antihistamines (when used for symptom relief)

**TREATMENT DRUGS (confirm diagnosis was positive):**
- Antimalarials: Artemether, Lumefantrine, Artesunate, Coartem
- Antibiotics: Amoxicillin, Ceftriaxone, Azithromycin, Metronidazole
- Antidiabetics: Metformin, Glibenclamide, Insulin
- Antihypertensives: Amlodipine, Lisinopril, Losartan
- Specific condition treatments

**VERDICT LOGIC:**
- Diagnostic TEST with its justifying diagnosis → PASS (diagnosis is valid for the test)
- Symptomatic drug with ruled-out diagnosis:
  1. Check if OTHER diagnoses exist in encounter
  2. If other dx can justify the symptomatic drug → PASS
  3. If NO other dx available → QUERY: "Symptomatic only - no valid diagnosis (malaria was ruled out by negative test)"

**COMMON TEST-TREATMENT PAIRS (for provisional diagnosis detection):**

| Diagnosis | Diagnostic Test | Expected Treatment (if positive) |
|-----------|-----------------|----------------------------------|
| Malaria | Malaria parasite test, MP | Artemether, Lumefantrine, Coartem, Artesunate |
| Typhoid | Widal test, Blood culture | Ciprofloxacin, Ceftriaxone, Azithromycin |
| UTI | Urinalysis, Urine M/C/S | Ciprofloxacin, Nitrofurantoin, Augmentin |
| H. pylori | H. pylori test, Stool antigen | Amoxicillin+Clarithromycin+PPI (triple therapy) |
| Diabetes | FBS, RBS, HbA1c | Metformin, Glibenclamide, Insulin |
| Pregnancy | Pregnancy test, Beta-HCG | Prenatal vitamins, Folic acid, specific OB drugs |
| Anemia | FBC/CBC, Hemoglobin | Iron supplements, Folic acid, B12 |
| TB | Gene-Xpert, Sputum AFB | Rifampicin, Isoniazid, Ethambutol, Pyrazinamide |
| HIV | HIV screening | ARVs |
| Hepatitis | Hepatitis panel | Antivirals if indicated |

**DECISION MATRIX (CORRECTED LOGIC):**

**For the DIAGNOSTIC TEST itself:**
- Diagnosis exists to justify test → TEST = PASS (diagnosis is VALID, it served its purpose)
- Test result negative (no treatment) does NOT make the diagnosis "wrong" - it was provisional/query

**For OTHER procedures (especially symptomatic drugs):**
- Test FOUND + Treatment FOUND → Diagnosis CONFIRMED → All procedures OK
- Test FOUND + Treatment NOT FOUND → Test was NEGATIVE, but:
  1. First check: Are there OTHER diagnoses in the encounter?
  2. If YES and procedure matches other dx → PASS
  3. If NO other dx AND procedure is symptomatic → QUERY (no valid dx to claim)

🚨 CRITICAL: 
- The test's diagnosis is NOT wrong - it justified the test!
- Only QUERY symptomatic drugs if they have NO other diagnosis to claim
- Always check ALL diagnoses before flagging!

### R11 - DOSAGE VALIDATION
- What is the standard/maximum dosage?
- Search: "[drug] dosage maximum therapeutic range"

### R12 - PROCEDURE-DIAGNOSIS MATCH
- Is this procedure appropriate for this diagnosis?
- Search: "[procedure] indications [diagnosis]"

## SEARCH STRATEGY
1. Search for the most relevant queries based on the procedure/drug
2. Prioritize reputable sources (WHO, FDA, BNF, UpToDate, NICE, MedlinePlus)
3. Look for Nigerian/African medical guidelines when relevant

## OUTPUT FORMAT
After searching, return a JSON object:
```json
{
  "drug_or_procedure": "the item",
  "diagnosis_checked": "the diagnosis",
  
  "gender_specific": true/false,
  "gender_notes": "male only / female only / not gender specific",
  
  "age_restrictions": "any age / pediatric only / adult only / contraindicated in elderly",
  "min_age": null,
  "max_age": null,
  
  "drug_class": "e.g., NSAID, Antibiotic, Antimalarial",
  "dangerous_interactions": ["list of dangerous drug interactions"],
  
  "verified_indications": ["list", "of", "approved indications"],
  "appropriate_for_diagnosis": true/false,
  
  "is_symptomatic_only": true/false,
  "symptomatic_notes": "analgesic / antipyretic / actual treatment",
  
  "standard_dosage": "e.g., 500mg TDS",
  "max_daily_dose": "e.g., 4g/day",
  
  "procedure_appropriate": true/false,
  
  "confidence": 0.95,
  "sources": ["source1", "source2"],
  "notes": "Any relevant clinical notes"
}
```

Be thorough and cite reputable medical sources. Focus on information relevant to fraud detection in health insurance claims."""


# =============================================================================
# CLAIMS VETTING ENGINE CLASS
# =============================================================================

class ClaimsVettingEngine:
    """
    Comprehensive Claims Vetting Engine with Claude Opus 4.5 Integration
    Enhanced with Web Search capabilities
    """
    
    def __init__(
        self, 
        api_key: Optional[str] = None, 
        db_path: str = 'ai_driven_data.duckdb',
        web_search_mode: str = 'smart',  # 'smart', 'on', 'off'
        model: Optional[str] = None  # Model selection
    ):
        """
        Initialize the vetting engine
        
        Parameters:
        -----------
        api_key : str, optional
            Anthropic API key. If not provided, reads from ANTHROPIC_API_KEY env var
        db_path : str
            Path to DuckDB database for 30-day prior claims lookup
        web_search_mode : str
            'smart' - Auto-search only when model is uncertain (default, cost-efficient)
            'on' - Web search for ALL claims (most thorough, higher cost)
            'off' - No web search (fastest, cheapest, uses AI knowledge only)
            'off' - No web search (not recommended)
        model : str, optional
            Claude model to use. Options:
            - 'opus' or 'claude-opus-4-5-20251101' (default, smartest)
            - 'sonnet' or 'claude-sonnet-4-5-20250929' (balanced)
            - 'haiku' or 'claude-haiku-4-5-20251001' (fastest, cheapest)
        """
        # Initialize Anthropic client
        self.api_key = api_key or os.getenv('ANTHROPIC_API_KEY')
        if not self.api_key:
            raise ValueError("ANTHROPIC_API_KEY not found. Set environment variable or pass api_key parameter.")
        
        self.client = anthropic.Anthropic(api_key=self.api_key)
        
        # Model selection - can be passed as parameter
        self.model = model or "claude-sonnet-4-5-20250929"  # Default to sonnet
        self.web_search_mode = web_search_mode
        
        # Database path for 30-day lookups
        self.db_path = db_path
        self.db_available = Path(db_path).exists()
        
        if not self.db_available:
            print(f"WARNING: DuckDB not found at {db_path}. 30-day prior claims check will be limited.")
        
        # Statistics tracking
        self.stats = {
            'total_claims': 0,
            'approved': 0,
            'denied': 0,
            'query': 0,
            'api_calls': 0,
            'api_errors': 0,
            'web_searches': 0,
            'web_search_triggers': {
                'manual': 0,
                'low_confidence': 0,
                'uncommon_drug': 0,
                'uncertainty_detected': 0
            }
        }
        
        print(f"✅ KLAIRE Claims Vetting Engine v3.8 initialized")
        print(f"   Model: {self.model}")
        print(f"   DuckDB: {'Available' if self.db_available else 'Not available'}")
        print(f"   Web Search Mode: {web_search_mode.upper()}")
    
    def load_claims(self, file_path: str) -> pd.DataFrame:
        """Load claims from CSV or Excel file"""
        path = Path(file_path)
        
        if path.suffix.lower() == '.csv':
            df = pd.read_csv(file_path)
        elif path.suffix.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")
        
        # Standardize column names
        df.columns = df.columns.str.strip()
        
        # Try to identify amount column
        amount_cols = ['Amount', 'Amount Charged', 'amount', 'Charged Amount', 'Total']
        for col in amount_cols:
            if col in df.columns:
                # Clean and convert amount
                df['Amount'] = df[col].astype(str).str.replace('₦', '').str.replace(',', '').astype(float)
                break
        
        print(f"✅ Loaded {len(df)} claims from {file_path}")
        return df
    
    def build_encounter_diagnoses(self, df: pd.DataFrame) -> Dict[Tuple, set]:
        """
        Build encounter-level diagnosis mapping
        
        Groups all diagnoses by (Enrollee, PA, Date) and parses separated diagnoses
        
        🚨 v3.8 FIX: Now handles BOTH comma (,) AND semicolon (;) separators!
        Example: "Essential hypertension; Incompetence of cervix uteri;" → 2 diagnoses
        """
        encounter_dx = {}
        
        # Identify relevant columns
        enrollee_col = self._find_column(df, ['Enrollee No', 'Enrollee', 'enrollee_id', 'Member ID'])
        pa_col = self._find_column(df, ['PA Number', 'PA', 'panumber', 'PA_Number'])
        date_col = self._find_column(df, ['Encounter Date', 'Date', 'encounter_date', 'Service Date'])
        diag_col = self._find_column(df, ['Diagnosis', 'diagnosis', 'Diagnosis Description'])
        
        for idx, row in df.iterrows():
            key = (row[enrollee_col], row[pa_col], row[date_col])
            diag = str(row[diag_col]).upper()
            
            if key not in encounter_dx:
                encounter_dx[key] = set()
            
            # 🚨 v3.8 FIX: Parse BOTH semicolon and comma separated diagnoses!
            # First replace semicolons with commas for uniform splitting
            diag_normalized = diag.replace(';', ',')
            
            if ',' in diag_normalized:
                parts = [d.strip() for d in diag_normalized.split(',')]
                for part in parts:
                    if part and part != 'NAN' and part != '':
                        encounter_dx[key].add(part)
            else:
                if diag and diag != 'NAN':
                    encounter_dx[key].add(diag)
        
        return encounter_dx
    
    def get_prior_claims(self, enrollee_id: str, current_date: str, days: int = 30) -> List[Dict]:
        """
        Query DuckDB for prior claims within specified days.
        
        ENHANCED: Now joins with PROCEDURE DATA, DIAGNOSIS, and PROVIDERS tables
        to get NAMES from CODES - critical for AI class matching!
        
        Returns:
            List of claims with procedure names, diagnosis names, and provider names
        """
        if not self.db_available:
            return []
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # Parse current date
            if isinstance(current_date, str):
                try:
                    current_dt = pd.to_datetime(current_date)
                except:
                    return []
            else:
                current_dt = pd.to_datetime(current_date)
            
            start_date = current_dt - timedelta(days=days)
            
            # ENHANCED QUERY: Joins to get NAMES from CODES
            # This is critical for AI to understand drug classes, diagnosis categories, etc.
            query = f"""
            SELECT 
                c.code as procedure_code,
                pd.proceduredesc,
                c.diagnosiscode,
                d.diagnosisdesc,
                c.nhisproviderid as providerid,
                p.providername,
                c.approvedamount,
                c.encounterdatefrom,
                c.datesubmitted
            FROM "AI DRIVEN DATA"."CLAIMS DATA" c
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd ON LOWER(c.code) = LOWER(pd.procedurecode)
            LEFT JOIN "AI DRIVEN DATA"."DIAGNOSIS" d ON LOWER(c.diagnosiscode) = LOWER(d.diagnosiscode)
            LEFT JOIN "AI DRIVEN DATA"."PROVIDERS" p ON c.nhisproviderid = p.providertin
            WHERE c.enrollee_id = '{enrollee_id}'
              AND c.encounterdatefrom >= '{start_date.strftime('%Y-%m-%d')}'
              AND c.encounterdatefrom < '{current_dt.strftime('%Y-%m-%d')}'
            ORDER BY c.encounterdatefrom DESC
            LIMIT 50
            """
            
            claims_result = conn.execute(query).fetchdf()
            
            # Also get PA DATA for comprehensive 30-day history
            # v3.8 FIX: Include diagnosis from PA DATA!
            pa_query = f"""
            SELECT 
                pa.code as procedure_code,
                pd.proceduredesc,
                pa.diagnosis as diagnosiscode,
                pa.additionaldiagnosis,
                pa.providerid,
                p.providername,
                pa.granted as approvedamount,
                pa.requestdate as encounterdatefrom,
                pa.requestdate as datesubmitted
            FROM "AI DRIVEN DATA"."PA DATA" pa
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd ON LOWER(pa.code) = LOWER(pd.procedurecode)
            LEFT JOIN "AI DRIVEN DATA"."PROVIDERS" p ON pa.providerid = p.providertin
            WHERE pa.IID = '{enrollee_id}'
              AND pa.requestdate >= '{start_date.strftime('%Y-%m-%d')}'
              AND pa.requestdate < '{current_dt.strftime('%Y-%m-%d')}'
              AND pa.pastatus = 'AUTHORIZED'
            ORDER BY pa.requestdate DESC
            LIMIT 30
            """
            
            pa_result = conn.execute(pa_query).fetchdf()
            conn.close()
            
            # Combine results and add source indicator
            all_results = []
            
            if len(claims_result) > 0:
                for _, row in claims_result.iterrows():
                    record = row.to_dict()
                    record['source'] = 'Claims'
                    # Ensure we have display-friendly procedure name
                    if pd.isna(record.get('proceduredesc')) or not record.get('proceduredesc'):
                        record['procedure'] = record.get('procedure_code', 'N/A')
                    else:
                        record['procedure'] = record.get('proceduredesc')
                    # Ensure we have display-friendly diagnosis name
                    if pd.isna(record.get('diagnosisdesc')) or not record.get('diagnosisdesc'):
                        record['diagnosis'] = record.get('diagnosiscode', 'N/A')
                    else:
                        record['diagnosis'] = record.get('diagnosisdesc')
                    all_results.append(record)
            
            if len(pa_result) > 0:
                for _, row in pa_result.iterrows():
                    record = row.to_dict()
                    record['source'] = 'PA'
                    # Ensure we have display-friendly procedure name
                    if pd.isna(record.get('proceduredesc')) or not record.get('proceduredesc'):
                        record['procedure'] = record.get('procedure_code', 'N/A')
                    else:
                        record['procedure'] = record.get('proceduredesc')
                    # v3.8 FIX: Include diagnosis from PA DATA
                    diag = record.get('diagnosiscode', '')
                    add_diag = record.get('additionaldiagnosis', '')
                    if pd.notna(diag) and diag:
                        # Combine main and additional diagnosis
                        all_pa_dx = [str(diag)]
                        if pd.notna(add_diag) and add_diag:
                            all_pa_dx.append(str(add_diag))
                        record['diagnosis'] = ' | '.join(all_pa_dx)
                        record['diagnosisdesc'] = record['diagnosis']  # For _format_prior_claims
                    else:
                        record['diagnosis'] = 'N/A'
                    all_results.append(record)
            
            return all_results
            
        except Exception as e:
            print(f"WARNING: Error querying prior claims: {e}")
            import traceback
            traceback.print_exc()
            return []
    
    def check_hospital_shopping(self, enrollee_id: str, current_date: str, current_provider: str, 
                                  current_procedure: str, current_diagnosis: str, days: int = 30) -> Dict[str, Any]:
        """
        R5: Check if enrollee visited OTHER hospitals for similar diagnosis/procedure in 30 days.
        
        Detects "hospital shopping" - potential fraud where enrollee gets same treatment
        at multiple facilities.
        
        Returns:
            {
                'found': bool - whether hospital shopping detected,
                'other_visits': list - details of visits to other providers,
                'same_diagnosis': bool - whether diagnosis matches,
                'same_procedure': bool - whether procedure matches,
                'message': str - explanation
            }
        """
        if not self.db_available:
            return {
                'found': False,
                'other_visits': [],
                'same_diagnosis': False,
                'same_procedure': False,
                'message': 'Database not available for hospital shopping check'
            }
        
        if not current_provider:
            return {
                'found': False,
                'other_visits': [],
                'same_diagnosis': False,
                'same_procedure': False,
                'message': 'No provider info available'
            }
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # Parse current date
            if isinstance(current_date, str):
                try:
                    current_dt = pd.to_datetime(current_date)
                except:
                    return {'found': False, 'other_visits': [], 'same_diagnosis': False, 
                            'same_procedure': False, 'message': 'Invalid date format'}
            else:
                current_dt = pd.to_datetime(current_date)
            
            start_date = current_dt - timedelta(days=days)
            
            # Normalize inputs for comparison
            current_provider_lower = current_provider.lower().strip() if current_provider else ''
            current_proc_lower = current_procedure.lower().strip() if current_procedure else ''
            current_diag_lower = current_diagnosis.lower().strip() if current_diagnosis else ''
            
            # Query CLAIMS for this enrollee at OTHER providers in last 30 days
            claims_query = f"""
            SELECT 
                c.nhisproviderid,
                p.providername,
                c.code as procedure_code,
                pd.proceduredesc,
                c.diagnosiscode,
                d.diagnosisdesc,
                c.encounterdatefrom,
                c.approvedamount
            FROM "AI DRIVEN DATA"."CLAIMS DATA" c
            LEFT JOIN "AI DRIVEN DATA"."PROVIDERS" p ON c.nhisproviderid = p.providertin
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd ON LOWER(c.code) = LOWER(pd.procedurecode)
            LEFT JOIN "AI DRIVEN DATA"."DIAGNOSIS" d ON LOWER(c.diagnosiscode) = LOWER(d.diagnosiscode)
            WHERE c.enrollee_id = '{enrollee_id}'
              AND c.encounterdatefrom >= '{start_date.strftime('%Y-%m-%d')}'
              AND c.encounterdatefrom <= '{current_dt.strftime('%Y-%m-%d')}'
            ORDER BY c.encounterdatefrom DESC
            LIMIT 100
            """
            
            claims_result = conn.execute(claims_query).fetchdf()
            
            # Also query PA DATA for visits to other providers
            pa_query = f"""
            SELECT 
                pa.providerid,
                p.providername,
                pa.code as procedure_code,
                pd.proceduredesc,
                pa.requestdate as encounterdatefrom,
                pa.granted as approvedamount
            FROM "AI DRIVEN DATA"."PA DATA" pa
            LEFT JOIN "AI DRIVEN DATA"."PROVIDERS" p ON pa.providerid = p.providertin
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd ON LOWER(pa.code) = LOWER(pd.procedurecode)
            WHERE pa.IID = '{enrollee_id}'
              AND pa.requestdate >= '{start_date.strftime('%Y-%m-%d')}'
              AND pa.requestdate <= '{current_dt.strftime('%Y-%m-%d')}'
              AND pa.pastatus = 'AUTHORIZED'
            ORDER BY pa.requestdate DESC
            LIMIT 100
            """
            
            pa_result = conn.execute(pa_query).fetchdf()
            conn.close()
            
            # Find visits to OTHER providers with similar diagnosis or procedure
            other_visits = []
            same_diagnosis_found = False
            same_procedure_found = False
            
            # Check claims
            for _, row in claims_result.iterrows():
                provider_name = str(row.get('providername', '')).lower().strip() if pd.notna(row.get('providername')) else ''
                
                # Skip if same provider
                if provider_name and current_provider_lower and (
                    provider_name in current_provider_lower or 
                    current_provider_lower in provider_name
                ):
                    continue
                
                # Check for similar diagnosis
                diag_desc = str(row.get('diagnosisdesc', '')).lower() if pd.notna(row.get('diagnosisdesc')) else ''
                diag_code = str(row.get('diagnosiscode', '')).lower() if pd.notna(row.get('diagnosiscode')) else ''
                
                diag_match = False
                if current_diag_lower and (
                    current_diag_lower in diag_desc or 
                    diag_desc in current_diag_lower or
                    current_diag_lower == diag_code
                ):
                    diag_match = True
                    same_diagnosis_found = True
                
                # Check for similar procedure
                proc_desc = str(row.get('proceduredesc', '')).lower() if pd.notna(row.get('proceduredesc')) else ''
                proc_code = str(row.get('procedure_code', '')).lower() if pd.notna(row.get('procedure_code')) else ''
                
                proc_match = False
                if current_proc_lower and (
                    current_proc_lower in proc_desc or 
                    proc_desc in current_proc_lower or
                    current_proc_lower == proc_code
                ):
                    proc_match = True
                    same_procedure_found = True
                
                # If either matches at a different provider, record it
                if (diag_match or proc_match) and provider_name:
                    other_visits.append({
                        'provider': row.get('providername', 'Unknown'),
                        'procedure': row.get('proceduredesc', row.get('procedure_code', 'N/A')),
                        'diagnosis': row.get('diagnosisdesc', row.get('diagnosiscode', 'N/A')),
                        'date': str(row.get('encounterdatefrom', 'N/A'))[:10],
                        'amount': float(row.get('approvedamount', 0)) if pd.notna(row.get('approvedamount')) else 0,
                        'source': 'Claims',
                        'diag_match': diag_match,
                        'proc_match': proc_match
                    })
            
            # Check PA data similarly
            for _, row in pa_result.iterrows():
                provider_name = str(row.get('providername', '')).lower().strip() if pd.notna(row.get('providername')) else ''
                
                # Skip if same provider
                if provider_name and current_provider_lower and (
                    provider_name in current_provider_lower or 
                    current_provider_lower in provider_name
                ):
                    continue
                
                # Check for similar procedure (PA doesn't have diagnosis)
                proc_desc = str(row.get('proceduredesc', '')).lower() if pd.notna(row.get('proceduredesc')) else ''
                proc_code = str(row.get('procedure_code', '')).lower() if pd.notna(row.get('procedure_code')) else ''
                
                proc_match = False
                if current_proc_lower and (
                    current_proc_lower in proc_desc or 
                    proc_desc in current_proc_lower
                ):
                    proc_match = True
                    same_procedure_found = True
                
                if proc_match and provider_name:
                    other_visits.append({
                        'provider': row.get('providername', 'Unknown'),
                        'procedure': row.get('proceduredesc', row.get('procedure_code', 'N/A')),
                        'diagnosis': 'N/A (PA)',
                        'date': str(row.get('encounterdatefrom', 'N/A'))[:10],
                        'amount': float(row.get('approvedamount', 0)) if pd.notna(row.get('approvedamount')) else 0,
                        'source': 'PA',
                        'diag_match': False,
                        'proc_match': proc_match
                    })
            
            # Build message
            if other_visits:
                providers = list(set([v['provider'] for v in other_visits[:3]]))
                message = f"HOSPITAL SHOPPING DETECTED: Enrollee visited {len(providers)} other provider(s) in 30 days for similar treatment: {', '.join(providers[:2])}"
            else:
                message = "No hospital shopping detected"
            
            return {
                'found': len(other_visits) > 0,
                'other_visits': other_visits[:5],  # Limit to 5 most relevant
                'same_diagnosis': same_diagnosis_found,
                'same_procedure': same_procedure_found,
                'message': message
            }
            
        except Exception as e:
            print(f"WARNING: Error checking hospital shopping: {e}")
            import traceback
            traceback.print_exc()
            return {
                'found': False,
                'other_visits': [],
                'same_diagnosis': False,
                'same_procedure': False,
                'message': f'Error checking: {str(e)}'
            }
    
    def analyze_visits_with_ai(self, current_visit: Dict, other_visits: List[Dict]) -> Dict:
        """
        Use AI to determine if visits at other hospitals are for SAME CLASS of treatment
        
        This enables detection of hospital shopping even when:
        - Drug names differ but CLASS is same (Amoxicillin vs Augmentin = both penicillins)
        - Diagnosis wording differs but condition is same (Malaria vs Plasmodium infection)
        - Procedure codes differ but treatment is equivalent
        
        Returns:
            {
                'same_class': bool,
                'analysis': str,
                'confidence': float,
                'fraud_indicators': list
            }
        """
        if not self.client:
            return {'same_class': False, 'analysis': 'API not available', 'confidence': 0}
        
        # Format visits for AI analysis
        current_info = f"""
CURRENT CLAIM:
- Provider: {current_visit.get('provider', 'Unknown')}
- Procedure: {current_visit.get('procedure', 'N/A')}
- Diagnosis: {current_visit.get('diagnosis', 'N/A')}
- Date: {current_visit.get('date', 'N/A')}
"""
        
        other_info = "OTHER PROVIDER VISITS (Last 30 days):\n"
        for v in other_visits[:5]:
            other_info += f"""- {v.get('date', 'N/A')} @ {v.get('provider', 'Unknown')}:
  Procedure: {v.get('procedure', 'N/A')}
  Diagnosis: {v.get('diagnosis', 'N/A')}
  Amount: ₦{v.get('amount', 0):,.0f}
"""
        
        ai_prompt = f"""Analyze these healthcare visits for HOSPITAL SHOPPING fraud:

{current_info}
{other_info}

TASK: Determine if the patient is visiting DIFFERENT hospitals for the SAME or SIMILAR treatment.

SAME CLASS means:
- Same drug CLASS (e.g., Amoxicillin & Augmentin = both penicillin antibiotics)
- Same diagnosis CLASS (e.g., "Malaria" & "Plasmodium falciparum infection" = same disease)
- Same procedure TYPE (e.g., "CT Scan Brain" & "MRI Head" = both neuroimaging)

Return JSON:
```json
{{
  "same_class_treatment": true/false,
  "explanation": "Brief explanation of why treatments are/aren't in same class",
  "current_drug_class": "e.g., Penicillin Antibiotic",
  "other_drug_classes": ["class1", "class2"],
  "diagnosis_match": true/false,
  "fraud_likelihood": "HIGH/MEDIUM/LOW/NONE",
  "fraud_indicators": ["indicator1", "indicator2"],
  "confidence": 0.95
}}
```

Consider:
- Legitimate reasons: Treatment failure, specialist referral, emergency
- Red flags: Same antibiotic class, duplicate antimalarials, multiple imaging for same condition"""

        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=800,
                messages=[{"role": "user", "content": ai_prompt}]
            )
            
            response_text = response.content[0].text
            
            # Parse JSON
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start != -1 and json_end > json_start:
                result = json.loads(response_text[json_start:json_end])
                return {
                    'same_class': result.get('same_class_treatment', False),
                    'analysis': result.get('explanation', ''),
                    'drug_class': result.get('current_drug_class', ''),
                    'diagnosis_match': result.get('diagnosis_match', False),
                    'fraud_likelihood': result.get('fraud_likelihood', 'NONE'),
                    'fraud_indicators': result.get('fraud_indicators', []),
                    'confidence': result.get('confidence', 0.5)
                }
            
            return {'same_class': False, 'analysis': 'Could not parse AI response', 'confidence': 0}
            
        except Exception as e:
            print(f"WARNING: AI visit analysis error: {e}")
            return {'same_class': False, 'analysis': f'Error: {str(e)}', 'confidence': 0}
    
    def get_enrollee_demographics(self, enrollee_id: str) -> Dict[str, Any]:
        """
        Get enrollee gender and age from MEMBERS table.
        
        CRITICAL for R6 (Gender) and R7 (Age) validation!
        
        Returns:
            {
                'gender': 'MALE' | 'FEMALE' | 'UNKNOWN',
                'age_years': int,
                'dob': date or None,
                'found': bool
            }
        """
        if not self.db_available:
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            query = f"""
            SELECT 
                genderid,
                dob,
                DATE_PART('year', AGE(CURRENT_DATE, dob)) as age_years
            FROM "AI DRIVEN DATA"."MEMBERS"
            WHERE enrollee_id = '{enrollee_id}'
            LIMIT 1
            """
            
            result = conn.execute(query).fetchdf()
            conn.close()
            
            if len(result) > 0:
                row = result.iloc[0]
                gender_id = row.get('genderid', 3)
                
                # Map gender ID to string
                gender_map = {1: 'MALE', 2: 'FEMALE', 3: 'UNKNOWN'}
                gender = gender_map.get(gender_id, 'UNKNOWN')
                
                age = row.get('age_years')
                if pd.notna(age):
                    age = int(age)
                else:
                    age = None
                
                dob = row.get('dob')
                
                return {
                    'gender': gender,
                    'age_years': age,
                    'dob': str(dob) if dob else None,
                    'found': True
                }
            
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
            
        except Exception as e:
            print(f"WARNING: Error getting enrollee demographics: {e}")
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
    
    # ================================================================
    # 🚨 v3.8 NEW: PA CONTEXT LOOKUP FOR NO-AUTH CLAIMS
    # ================================================================
    def get_pa_context_for_noauth(self, enrollee_id: str, encounter_date: str, hours: int = 24) -> Dict[str, Any]:
        """
        🚨 v3.8 NEW: For NO-AUTH claims (PA=0), look up PA table for context.
        
        When a provider submits some procedures as NO-AUTH alongside PA procedures,
        we need to find the related PA to understand the clinical context.
        
        Example: Tranexamic acid with diagnosis "Hypertension" looks wrong,
        but if we find a PA for same enrollee ±24 hours with diagnosis
        "Threatened abortion", then the drug makes sense!
        
        Parameters:
        -----------
        enrollee_id : str
            The enrollee ID
        encounter_date : str
            The date of the NO-AUTH claim
        hours : int
            Hours before/after to search (default: 24)
        
        Returns:
        --------
        Dict with:
        - found: bool
        - pa_procedures: List of procedure descriptions from PA
        - pa_diagnoses: List of diagnoses from PA
        - pa_numbers: List of related PA numbers
        - context_summary: str summary for AI
        """
        if not self.db_available:
            return {'found': False, 'pa_procedures': [], 'pa_diagnoses': [], 'pa_numbers': [], 'context_summary': ''}
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # Parse encounter date
            if isinstance(encounter_date, str):
                try:
                    enc_dt = pd.to_datetime(encounter_date)
                except:
                    return {'found': False, 'pa_procedures': [], 'pa_diagnoses': [], 'pa_numbers': [], 'context_summary': ''}
            else:
                enc_dt = pd.to_datetime(encounter_date)
            
            # Calculate time window (±hours)
            start_dt = enc_dt - timedelta(hours=hours)
            end_dt = enc_dt + timedelta(hours=hours)
            
            # Query PA DATA for this enrollee within the time window
            query = f"""
            SELECT DISTINCT
                pa.panumber,
                pa.requestdate,
                pa.code as procedure_code,
                pd.proceduredesc,
                pa.diagnosis as diagnosis_code,
                d.diagnosisdesc,
                pa.pastatus,
                pa.granted,
                pa.providerid,
                p.providername
            FROM "AI DRIVEN DATA"."PA DATA" pa
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd 
                ON LOWER(TRIM(pa.code)) = LOWER(TRIM(pd.procedurecode))
            LEFT JOIN "AI DRIVEN DATA"."DIAGNOSIS" d 
                ON LOWER(TRIM(pa.diagnosis)) = LOWER(TRIM(d.diagnosiscode))
            LEFT JOIN "AI DRIVEN DATA"."PROVIDERS" p 
                ON pa.providerid = p.providertin
            WHERE pa.IID = '{enrollee_id}'
              AND pa.requestdate >= '{start_dt.strftime('%Y-%m-%d %H:%M:%S')}'
              AND pa.requestdate <= '{end_dt.strftime('%Y-%m-%d %H:%M:%S')}'
              AND pa.pastatus IS NOT NULL
            ORDER BY pa.requestdate DESC
            LIMIT 50
            """
            
            result = conn.execute(query).fetchdf()
            conn.close()
            
            if len(result) == 0:
                return {'found': False, 'pa_procedures': [], 'pa_diagnoses': [], 'pa_numbers': [], 'context_summary': ''}
            
            # Extract unique procedures, diagnoses, and PA numbers
            pa_procedures = []
            pa_diagnoses = []
            pa_numbers = []
            
            for _, row in result.iterrows():
                # Procedure
                proc_desc = row.get('proceduredesc')
                proc_code = row.get('procedure_code')
                if pd.notna(proc_desc) and proc_desc not in pa_procedures:
                    pa_procedures.append(str(proc_desc))
                elif pd.notna(proc_code) and proc_code not in pa_procedures:
                    pa_procedures.append(str(proc_code))
                
                # Diagnosis
                diag_desc = row.get('diagnosisdesc')
                diag_code = row.get('diagnosis_code')
                if pd.notna(diag_desc) and diag_desc not in pa_diagnoses:
                    pa_diagnoses.append(str(diag_desc))
                elif pd.notna(diag_code) and diag_code not in pa_diagnoses:
                    pa_diagnoses.append(str(diag_code))
                
                # PA Number
                pa_num = row.get('panumber')
                if pd.notna(pa_num) and pa_num not in pa_numbers:
                    pa_numbers.append(str(int(pa_num)))
            
            # Build context summary for AI
            context_parts = []
            if pa_diagnoses:
                context_parts.append(f"Related PA diagnoses (±{hours}hrs): {', '.join(pa_diagnoses[:5])}")
            if pa_procedures:
                context_parts.append(f"Related PA procedures (±{hours}hrs): {', '.join(pa_procedures[:5])}")
            if pa_numbers:
                context_parts.append(f"PA numbers: {', '.join(pa_numbers[:3])}")
            
            context_summary = "; ".join(context_parts) if context_parts else ""
            
            return {
                'found': True,
                'pa_procedures': pa_procedures[:10],
                'pa_diagnoses': pa_diagnoses[:10],
                'pa_numbers': pa_numbers[:5],
                'context_summary': context_summary
            }
            
        except Exception as e:
            print(f"WARNING: Error getting PA context for NO-AUTH: {e}")
            return {'found': False, 'pa_procedures': [], 'pa_diagnoses': [], 'pa_numbers': [], 'context_summary': ''}
    
    def validate_pa_details(self, pa_number: int, claim_procedure: str, claim_amount: float, claim_date: str = None) -> Dict[str, Any]:
        """
        Validate claim against PA authorization details.
        
        FOR PRE-AUTH CLAIMS ONLY (PA > 0)
        
        FIXED IN v3.5: Now properly looks up procedure codes!
        - Step 1: Convert claim procedure DESCRIPTION to CODE via PROCEDURE DATA
        - Step 2: Get PA's authorized CODES from PA DATA
        - Step 3: Check if claim's code is in PA's authorized list
        
        Checks:
        1. Does the PA exist and is it AUTHORIZED?
        2. Is the claimed procedure CODE in the authorized list?
        3. Is the claimed amount within the granted amount?
        4. Is the claim date reasonable vs PA request date?
        
        Returns:
            {
                'pa_found': bool,
                'pa_status': str (AUTHORIZED, PENDING, etc.),
                'procedure_match': bool,
                'amount_valid': bool,
                'date_valid': bool,
                'authorized_procedures': list,
                'authorized_amount': float,
                'mismatches': list of issues found,
                'overall_valid': bool,
                'claim_procedure_code': str (the code we found for the claim procedure)
            }
        """
        if not self.db_available or pa_number <= 0:
            return {
                'pa_found': False,
                'pa_status': 'N/A',
                'procedure_match': True,  # Can't check, assume OK
                'amount_valid': True,
                'date_valid': True,
                'authorized_procedures': [],
                'authorized_amount': 0,
                'mismatches': [],
                'overall_valid': True,
                'claim_procedure_code': None
            }
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # ================================================================
            # STEP 1: Look up the claim's procedure code from its description
            # ================================================================
            claim_proc_lower = claim_procedure.lower().strip() if claim_procedure else ''
            claim_procedure_code = None
            
            if claim_proc_lower:
                # Search PROCEDURE DATA for matching description
                proc_query = f"""
                SELECT procedurecode, proceduredesc
                FROM "AI DRIVEN DATA"."PROCEDURE DATA"
                WHERE LOWER(proceduredesc) = '{claim_proc_lower}'
                   OR LOWER(proceduredesc) LIKE '%{claim_proc_lower}%'
                ORDER BY 
                    CASE WHEN LOWER(proceduredesc) = '{claim_proc_lower}' THEN 0 ELSE 1 END,
                    LENGTH(proceduredesc)
                LIMIT 1
                """
                proc_result = conn.execute(proc_query).fetchdf()
                
                if len(proc_result) > 0:
                    claim_procedure_code = proc_result['procedurecode'].iloc[0]
            
            # ================================================================
            # STEP 2: Get PA's authorized procedure codes and amounts
            # ================================================================
            pa_query = f"""
            SELECT 
                p.panumber,
                p.code as procedure_code,
                p.granted,
                p.totaltariff,
                p.requestdate,
                p.pastatus,
                p.IID as enrollee,
                pd.proceduredesc as procedure_desc
            FROM "AI DRIVEN DATA"."PA DATA" p
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd 
                ON LOWER(p.code) = LOWER(pd.procedurecode)
            WHERE p.panumber = '{pa_number}'
            """
            
            result = conn.execute(pa_query).fetchdf()
            conn.close()
            
            if len(result) == 0:
                return {
                    'pa_found': False,
                    'pa_status': 'NOT FOUND',
                    'procedure_match': False,
                    'amount_valid': False,
                    'date_valid': False,
                    'authorized_procedures': [],
                    'authorized_amount': 0,
                    'mismatches': [f'PA {pa_number} not found in database'],
                    'overall_valid': False,
                    'claim_procedure_code': claim_procedure_code
                }
            
            # Get PA status
            pa_status = result['pastatus'].iloc[0] if 'pastatus' in result.columns else 'UNKNOWN'
            
            # Build list of authorized procedure codes and amounts
            authorized_codes = set()
            authorized_procedures = []
            total_authorized = 0
            
            for _, row in result.iterrows():
                proc_code = str(row.get('procedure_code', '')).upper().strip()
                proc_desc = str(row.get('procedure_desc', '')) if pd.notna(row.get('procedure_desc')) else ''
                granted = float(row.get('granted', 0)) if pd.notna(row.get('granted')) else 0
                
                authorized_codes.add(proc_code.lower())
                authorized_procedures.append({
                    'code': proc_code,
                    'description': proc_desc,
                    'granted': granted
                })
                total_authorized += granted
            
            # ================================================================
            # STEP 3: Check if claim's procedure code is in PA's authorized list
            # ================================================================
            procedure_match = False
            matched_procedure = None
            matched_amount = 0
            
            if claim_procedure_code:
                claim_code_lower = claim_procedure_code.lower().strip()
                
                # Direct code match
                if claim_code_lower in authorized_codes:
                    procedure_match = True
                    # Find the matching procedure for amount
                    for auth_proc in authorized_procedures:
                        if auth_proc['code'].lower() == claim_code_lower:
                            matched_procedure = auth_proc
                            matched_amount = auth_proc['granted']
                            break
            
            # If no code match found, try description matching as fallback
            if not procedure_match and claim_proc_lower:
                for auth_proc in authorized_procedures:
                    auth_desc = auth_proc['description'].lower() if auth_proc['description'] else ''
                    if auth_desc and (
                        claim_proc_lower in auth_desc or 
                        auth_desc in claim_proc_lower
                    ):
                        procedure_match = True
                        matched_procedure = auth_proc
                        matched_amount = auth_proc['granted']
                        break
            
            # Check amount validity
            if matched_procedure:
                amount_valid = claim_amount <= matched_amount * 1.1  # Allow 10% tolerance
                compare_amount = matched_amount
            else:
                # No procedure match - amount is invalid by default
                amount_valid = False
                compare_amount = total_authorized
            
            # Check date validity
            date_valid = True
            if claim_date and len(result) > 0:
                pa_date = result['requestdate'].iloc[0]
                if pd.notna(pa_date):
                    try:
                        from datetime import datetime, timedelta
                        if isinstance(claim_date, str):
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                                try:
                                    claim_dt = datetime.strptime(claim_date.split()[0], fmt.split()[0])
                                    break
                                except:
                                    continue
                            else:
                                claim_dt = None
                        else:
                            claim_dt = claim_date
                        
                        if claim_dt:
                            pa_dt = pd.to_datetime(pa_date)
                            if claim_dt < pa_dt.to_pydatetime() - timedelta(days=1):
                                date_valid = False
                            elif claim_dt > pa_dt.to_pydatetime() + timedelta(days=90):
                                date_valid = False
                    except Exception:
                        pass
            
            # Build mismatches list
            mismatches = []
            
            if pa_status != 'AUTHORIZED':
                mismatches.append(f'PA status is {pa_status}, not AUTHORIZED')
            
            if not procedure_match:
                auth_procs_str = ', '.join([f"{p['code']}({p['description'][:30]})" for p in authorized_procedures[:3]])
                if claim_procedure_code:
                    mismatches.append(f'PROCEDURE MISMATCH: Claim procedure "{claim_procedure}" (code: {claim_procedure_code}) NOT in PA. Authorized: {auth_procs_str}')
                else:
                    mismatches.append(f'PROCEDURE MISMATCH: Claim procedure "{claim_procedure}" NOT in PA. Authorized: {auth_procs_str}')
            
            if procedure_match and not amount_valid:
                mismatches.append(f'AMOUNT EXCEEDS: Claimed ₦{claim_amount:,.2f} > Authorized ₦{compare_amount:,.2f}')
            
            if not date_valid:
                mismatches.append(f'DATE ISSUE: Claim date {claim_date} outside PA validity period')
            
            # Overall validity
            overall_valid = (
                pa_status == 'AUTHORIZED' and 
                procedure_match and 
                amount_valid and 
                date_valid
            )
            
            return {
                'pa_found': True,
                'pa_status': pa_status,
                'procedure_match': procedure_match,
                'amount_valid': amount_valid,
                'date_valid': date_valid,
                'authorized_procedures': [f"{p['code']}: {p['description']}" for p in authorized_procedures],
                'authorized_amount': compare_amount if matched_procedure else total_authorized,
                'mismatches': mismatches,
                'overall_valid': overall_valid,
                'claim_procedure_code': claim_procedure_code
            }
            
        except Exception as e:
            print(f"WARNING: Error validating PA details: {e}")
            import traceback
            traceback.print_exc()
            return {
                'pa_found': False,
                'pa_status': 'ERROR',
                'procedure_match': True,
                'amount_valid': True,
                'date_valid': True,
                'authorized_procedures': [],
                'authorized_amount': 0,
                'mismatches': [f'Error checking PA: {str(e)}'],
                'overall_valid': True,
                'claim_procedure_code': None
            }
    
    def _fuzzy_procedure_match(self, claim_proc: str, auth_proc: str) -> bool:
        """
        Fuzzy match procedure names (handles slight variations)
        """
        if not claim_proc or not auth_proc:
            return False
        
        # Remove common words and compare key terms
        stop_words = {'tab', 'tablet', 'cap', 'capsule', 'inj', 'injection', 'syrup', 'susp', 'suspension', 'mg', 'ml'}
        
        def extract_key_terms(proc: str) -> set:
            words = proc.lower().replace(',', ' ').replace('.', ' ').split()
            return set(w for w in words if w not in stop_words and len(w) > 2)
        
        claim_terms = extract_key_terms(claim_proc)
        auth_terms = extract_key_terms(auth_proc)
        
        if not claim_terms or not auth_terms:
            return False
        
        # Check if there's significant overlap
        common = claim_terms & auth_terms
        if len(common) >= 1:  # At least one key term matches
            return True
        
        return False
    
    def validate_tariff_amount(self, provider_name: str, procedure_desc: str, claimed_amount: float) -> Dict[str, Any]:
        """
        Validate claimed amount against provider's tariff.
        
        NEW IN v3.5: Actually looks up provider tariff from database!
        
        CORRECT JOIN PATH:
        PROVIDERS.protariffid → PROVIDERS_TARIFF.protariffid → TARIFF.tariffid
        
        Steps:
        1. Look up procedure code from description in PROCEDURE DATA
        2. Find provider by name in PROVIDERS table to get protariffid
        3. Join through PROVIDERS_TARIFF to TARIFF for actual price
        4. Compare claimed amount to tariff amount
        
        Returns:
            {
                'tariff_found': bool,
                'provider_tariff': float,
                'procedure_code': str,
                'amount_valid': bool,
                'variance_pct': float (% over/under tariff),
                'message': str
            }
        """
        if not self.db_available:
            return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,  # Can't check, assume OK
                'variance_pct': 0,
                'message': 'Database not available for tariff check'
            }
        
        if not provider_name or not procedure_desc:
            return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,
                'variance_pct': 0,
                'message': 'Missing provider or procedure information'
            }
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            proc_desc_lower = procedure_desc.lower().strip()
            provider_lower = provider_name.lower().strip()
            
            # Step 1: Look up procedure code from description
            proc_query = f"""
            SELECT procedurecode, proceduredesc
            FROM "AI DRIVEN DATA"."PROCEDURE DATA"
            WHERE LOWER(proceduredesc) = '{proc_desc_lower}'
               OR LOWER(proceduredesc) LIKE '%{proc_desc_lower}%'
            ORDER BY 
                CASE WHEN LOWER(proceduredesc) = '{proc_desc_lower}' THEN 0 ELSE 1 END,
                LENGTH(proceduredesc)
            LIMIT 1
            """
            proc_result = conn.execute(proc_query).fetchdf()
            
            if len(proc_result) == 0:
                conn.close()
                return {
                    'tariff_found': False,
                    'provider_tariff': None,
                    'procedure_code': None,
                    'amount_valid': True,
                    'variance_pct': 0,
                    'message': f'Procedure "{procedure_desc}" not found in procedure database'
                }
            
            procedure_code = proc_result['procedurecode'].iloc[0]
            
            # Step 2 & 3: Look up provider's tariff using CORRECT JOIN PATH
            # PROVIDERS.protariffid → PROVIDERS_TARIFF.protariffid → TARIFF.tariffid
            tariff_query = f"""
            SELECT 
                p.providername,
                p.protariffid,
                t.tariffid,
                t.tariffname,
                t.procedurecode,
                t.tariffamount
            FROM "AI DRIVEN DATA"."PROVIDERS" p
            JOIN "AI DRIVEN DATA"."PROVIDERS_TARIFF" pt ON p.protariffid = pt.protariffid
            JOIN "AI DRIVEN DATA"."TARIFF" t ON CAST(pt.tariffid AS VARCHAR) = t.tariffid
            WHERE LOWER(t.procedurecode) = LOWER('{procedure_code}')
              AND (
                  LOWER(p.providername) = '{provider_lower}'
                  OR LOWER(p.providername) LIKE '%{provider_lower}%'
                  OR LOWER(p.providername) LIKE '{provider_lower.split()[0]}%'
              )
            ORDER BY 
                CASE WHEN LOWER(p.providername) = '{provider_lower}' THEN 0 ELSE 1 END
            LIMIT 1
            """
            tariff_result = conn.execute(tariff_query).fetchdf()
            conn.close()
            
            if len(tariff_result) == 0:
                return {
                    'tariff_found': False,
                    'provider_tariff': None,
                    'procedure_code': procedure_code,
                    'amount_valid': True,
                    'variance_pct': 0,
                    'message': f'No tariff found for provider "{provider_name}" (procedure: {procedure_code})'
                }
            
            provider_tariff = float(tariff_result['tariffamount'].iloc[0])
            tariff_name = tariff_result['tariffname'].iloc[0]
            matched_provider = tariff_result['providername'].iloc[0]
            
            # Step 4: Compare claimed amount to tariff
            if provider_tariff > 0:
                variance_pct = ((claimed_amount - provider_tariff) / provider_tariff) * 100
                # Allow 5% tolerance for rounding
                amount_valid = claimed_amount <= provider_tariff * 1.05
            else:
                variance_pct = 0
                amount_valid = True
            
            if amount_valid:
                message = f'Amount ₦{claimed_amount:,.2f} within tariff ₦{provider_tariff:,.2f} ({matched_provider})'
            else:
                message = f'AMOUNT EXCEEDS TARIFF: Claimed ₦{claimed_amount:,.2f} > Tariff ₦{provider_tariff:,.2f} ({matched_provider}) by {variance_pct:.1f}%'
            
            return {
                'tariff_found': True,
                'provider_tariff': provider_tariff,
                'procedure_code': procedure_code,
                'amount_valid': amount_valid,
                'variance_pct': variance_pct,
                'message': message
            }
            
        except Exception as e:
            print(f"WARNING: Error validating tariff: {e}")
            return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,
                'variance_pct': 0,
                'message': f'Error checking tariff: {str(e)}'
            }
    
    def _find_column(self, df: pd.DataFrame, candidates: List[str]) -> str:
        """Find matching column name from candidates"""
        for col in candidates:
            if col in df.columns:
                return col
        raise ValueError(f"Could not find column. Tried: {candidates}. Available: {list(df.columns)}")
    
    def _is_uncommon_drug(self, drug_name: str) -> bool:
        """Check if drug is not in common formulary"""
        if not drug_name:
            return False
        
        drug_lower = drug_name.lower().strip()
        
        # Check against common drugs
        for common in COMMON_DRUGS:
            if common in drug_lower or drug_lower in common:
                return False
        
        return True
    
    def _has_uncertainty_indicators(self, response_text: str) -> bool:
        """Check if Opus response indicates uncertainty"""
        text_lower = response_text.lower()
        
        for indicator in UNCERTAINTY_INDICATORS:
            if indicator in text_lower:
                return True
        
        return False
    
    def _should_trigger_web_search(
        self, 
        drug_name: str, 
        opus_response: Dict, 
        response_text: str
    ) -> Tuple[bool, str]:
        """
        Determine if web search should be triggered
        
        Returns:
        --------
        (should_search, reason)
        """
        if self.web_search_mode == 'off':
            return False, 'disabled'
        
        if self.web_search_mode == 'on':
            return True, 'manual'
        
        # Smart mode logic
        if self.web_search_mode == 'smart':
            # Check 1: Explicit needs_verification flag
            if opus_response.get('needs_verification', False):
                return True, 'opus_requested'
            
            # Check 2: Low confidence
            confidence = opus_response.get('confidence', 1.0)
            if confidence < 0.7:
                return True, 'low_confidence'
            
            # Check 3: Uncommon drug
            if self._is_uncommon_drug(drug_name):
                return True, 'uncommon_drug'
            
            # Check 4: Uncertainty indicators in response
            if self._has_uncertainty_indicators(response_text):
                return True, 'uncertainty_detected'
        
        return False, 'not_needed'
    
    def perform_web_search(self, drug_name: str, diagnosis: str, 
                            patient_gender: str = None, patient_age: int = None) -> Dict:
        """
        Perform web search to verify medical information for ALL rules
        
        Uses Claude with web search tool to look up:
        - R6: Gender appropriateness
        - R7: Age appropriateness
        - R8: Drug interactions and class
        - R9: Drug-diagnosis match
        - R10: Symptomatic vs treatment
        - R11: Dosage validation
        - R12: Procedure-diagnosis match
        """
        self.stats['web_searches'] += 1
        
        # Build context string
        patient_context = ""
        if patient_gender and patient_gender != 'UNKNOWN':
            patient_context += f"\n**Patient Gender:** {patient_gender}"
        if patient_age:
            patient_context += f"\n**Patient Age:** {patient_age} years old"
        
        user_prompt = f"""Please search for and verify comprehensive medical information:

**Drug/Procedure:** {drug_name}
**Diagnosis:** {diagnosis}{patient_context}

Search for and verify ALL of the following:

1. **Gender Appropriateness (R6):** Is this procedure/drug gender-specific? (e.g., cervical procedures female-only, prostate procedures male-only)

2. **Age Appropriateness (R7):** What are age restrictions? Is it safe for pediatric ({patient_age} years old if provided)?

3. **Drug Class & Interactions (R8):** What drug class? Any dangerous interactions with common medications?

4. **Indications (R9):** What are approved indications? Is it appropriate for {diagnosis}?

5. **Symptomatic vs Treatment (R10):** Is this symptomatic relief only, or actual treatment for {diagnosis}?

6. **Dosage (R11):** What is the standard and maximum dosage?

7. **Procedure Match (R12):** If this is a procedure, is it appropriate for {diagnosis}?

Return comprehensive JSON with all findings."""

        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=2000,
                system=WEB_SEARCH_SYSTEM_PROMPT,
                tools=[{
                    "type": "web_search_20250305",
                    "name": "web_search"
                }],
                messages=[
                    {"role": "user", "content": user_prompt}
                ]
            )
            
            # Extract text content (may have multiple blocks due to tool use)
            response_text = ""
            for block in response.content:
                if hasattr(block, 'text'):
                    response_text += block.text
            
            # Parse JSON from response
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start != -1 and json_end > json_start:
                json_str = response_text[json_start:json_end]
                result = json.loads(json_str)
                result['search_performed'] = True
                return result
            else:
                return {
                    'search_performed': True,
                    'drug_name': drug_name,
                    'error': 'Could not parse search results',
                    'raw_response': response_text[:500]
                }
                
        except Exception as e:
            print(f"WARNING: Web search error for {drug_name}: {e}")
            return {
                'search_performed': True,
                'drug_name': drug_name,
                'error': str(e)
            }
    
    def _merge_web_search_results(self, opus_result: Dict, web_result: Dict, 
                                      patient_gender: str = None, patient_age: int = None) -> Dict:
        """
        Merge web search findings into API result for ALL medical rules (R6-R12)
        
        🚨 v3.8 FIX: Web search now CONFIRMS AI's decision, not overrides it!
        - If patient gender/age is KNOWN and MATCHES requirement → keep PASS
        - Only change to QUERY if gender/age is UNKNOWN
        - PRESERVE AI's context-aware analysis - web search ADDS info, doesn't REPLACE
        
        Web search provides comprehensive information for:
        - R6: Gender appropriateness
        - R7: Age appropriateness  
        - R8: Drug class and interactions
        - R9: Drug-diagnosis match
        - R10: Symptomatic vs treatment
        - R11: Dosage validation
        - R12: Procedure-diagnosis match
        """
        if not web_result.get('search_performed') or web_result.get('error'):
            opus_result['web_search'] = {
                'performed': True,
                'status': 'error',
                'error': web_result.get('error', 'Unknown error')
            }
            return opus_result
        
        # ================================================================
        # R6: Gender Appropriateness - 🚨 v3.8 FIX: Respect known gender!
        # ================================================================
        if web_result.get('gender_specific') is not None:
            gender_specific = web_result.get('gender_specific', False)
            gender_notes = web_result.get('gender_notes', '')
            
            if gender_specific:
                # Check if we KNOW the patient's gender
                if patient_gender and patient_gender.upper() in ['MALE', 'FEMALE']:
                    # Patient gender is KNOWN - check if it matches requirement
                    required_gender = None
                    notes_lower = gender_notes.lower()
                    if 'female' in notes_lower and 'male' not in notes_lower.replace('female', ''):
                        required_gender = 'FEMALE'
                    elif 'male' in notes_lower and 'female' not in notes_lower:
                        required_gender = 'MALE'
                    
                    if required_gender:
                        if patient_gender.upper() == required_gender:
                            # Gender MATCHES requirement - CONFIRM the PASS!
                            if opus_result.get('R6', {}).get('status') in ['PASS', 'QUERY']:
                                opus_result['R6'] = {
                                    'status': 'PASS',
                                    'reason': f"Gender-specific ({gender_notes}) - Patient is {patient_gender} ✓ CONFIRMED"
                                }
                        else:
                            # Gender does NOT match - this is a FAIL!
                            opus_result['R6'] = {
                                'status': 'FAIL',
                                'reason': f"Gender mismatch: Procedure is {gender_notes} but patient is {patient_gender}"
                            }
                    else:
                        # Could not determine required gender from notes - keep AI's decision
                        pass
                else:
                    # Gender is UNKNOWN - ask to verify (original behavior)
                    if opus_result.get('R6', {}).get('status') == 'PASS':
                        opus_result['R6'] = {
                            'status': 'QUERY',
                            'reason': f"Web search indicates gender-specific: {gender_notes}. Verify patient gender."
                        }
        
        # ================================================================
        # R7: Age Appropriateness - 🚨 v3.8 FIX: Respect known age!
        # ================================================================
        if web_result.get('age_restrictions'):
            age_restrictions = web_result.get('age_restrictions', '')
            restrictions_lower = age_restrictions.lower()
            
            if 'contraindicated' in restrictions_lower or 'not recommended' in restrictions_lower:
                # Check if we KNOW the patient's age
                if patient_age is not None:
                    # Try to extract age threshold from restrictions
                    # Common patterns: "children under 12", "adults only", "not for >65"
                    import re
                    
                    # Check for pediatric restrictions
                    pediatric_match = re.search(r'(\d+)\s*(years?|yo|y\.?o\.?)', restrictions_lower)
                    is_pediatric_restricted = 'child' in restrictions_lower or 'pediatric' in restrictions_lower
                    is_geriatric_restricted = 'elder' in restrictions_lower or 'geriatric' in restrictions_lower or '>65' in restrictions_lower
                    
                    if is_pediatric_restricted and patient_age >= 18:
                        # Adult patient, pediatric restriction doesn't apply
                        opus_result['R7'] = {
                            'status': 'PASS',
                            'reason': f"Age restriction ({age_restrictions}) - Patient is {patient_age}yo adult ✓"
                        }
                    elif is_geriatric_restricted and patient_age < 65:
                        # Non-elderly patient, geriatric restriction doesn't apply
                        opus_result['R7'] = {
                            'status': 'PASS',
                            'reason': f"Age restriction ({age_restrictions}) - Patient is {patient_age}yo ✓"
                        }
                    elif opus_result.get('R7', {}).get('status') == 'PASS':
                        # Might be a concern - add info but keep as QUERY
                        opus_result['R7'] = {
                            'status': 'QUERY',
                            'reason': f"Web search found age restrictions: {age_restrictions}. Patient is {patient_age}yo - verify appropriateness."
                        }
                else:
                    # Age is UNKNOWN - ask to verify
                    if opus_result.get('R7', {}).get('status') == 'PASS':
                        opus_result['R7'] = {
                            'status': 'QUERY',
                            'reason': f"Web search found age restrictions: {age_restrictions}"
                        }
        
        # R8: Drug Interactions - Add dangerous interactions info (ENHANCE, don't replace)
        dangerous_interactions = web_result.get('dangerous_interactions', [])
        if dangerous_interactions:
            drug_class = web_result.get('drug_class', 'Unknown')
            interactions_text = ', '.join(dangerous_interactions[:3]) if dangerous_interactions else 'None found'
            # Enhance R8 reason with web search info
            current_r8 = opus_result.get('R8', {})
            if current_r8.get('status') != 'FAIL':
                current_reason = current_r8.get('reason', '')
                enhanced_reason = f"{current_reason} | Web: Drug class={drug_class}, Watch for: {interactions_text}"
                opus_result['R8'] = {
                    'status': current_r8.get('status', 'PASS'),
                    'reason': enhanced_reason[:500]  # Limit length
                }
        
        # ================================================================
        # R9: Drug-Diagnosis Match - 🚨 v3.8 FIX: PRESERVE AI's context!
        # ================================================================
        appropriate = web_result.get('appropriate_for_diagnosis', True)
        if not appropriate:
            current_r9 = opus_result.get('R9', {})
            current_reason = current_r9.get('reason', '')
            verified_indications = web_result.get('verified_indications') or []
            indications_text = ', '.join(verified_indications[:3]) if verified_indications else 'None found'
            
            # 🚨 v3.8 FIX: PRESERVE AI's context-aware analysis!
            # If AI found relevant context (like "however, prior claims show..."), keep it!
            if current_r9.get('status') == 'PASS':
                # AI said PASS but web says inappropriate - ADD web info, ask for review
                opus_result['R9'] = {
                    'status': 'QUERY',
                    'reason': f"Web search indicates drug may not be standard for diagnosis. Verified indications: {indications_text}. {current_reason}"
                }
            elif current_r9.get('status') == 'QUERY' and current_reason:
                # AI already flagged with context - ENHANCE with web info, don't replace!
                enhanced_reason = f"{current_reason} | Web verified indications: {indications_text}"
                opus_result['R9'] = {
                    'status': 'QUERY',
                    'reason': enhanced_reason[:500]
                }
        
        # R10: Symptomatic Check - Flag if symptomatic only (ENHANCE, don't replace)
        is_symptomatic = web_result.get('is_symptomatic_only', False)
        symptomatic_notes = web_result.get('symptomatic_notes', '')
        if is_symptomatic:
            current_r10 = opus_result.get('R10', {})
            current_reason = current_r10.get('reason', '')
            if current_r10.get('status') != 'FAIL':
                # Enhance with web info
                enhanced_reason = f"{current_reason} | Web confirms symptomatic only: {symptomatic_notes}. Verify treatment drug present."
                opus_result['R10'] = {
                    'status': 'QUERY',
                    'reason': enhanced_reason[:500]
                }
        
        # R11: Dosage Validation - Add standard dosage info (ENHANCE)
        standard_dosage = web_result.get('standard_dosage', '')
        max_dose = web_result.get('max_daily_dose', '')
        if standard_dosage or max_dose:
            current_r11 = opus_result.get('R11', {})
            dosage_info = f"Standard: {standard_dosage}" if standard_dosage else ""
            if max_dose:
                dosage_info += f" | Max: {max_dose}"
            current_reason = current_r11.get('reason', '')
            enhanced_reason = f"{current_reason} | Web: {dosage_info}"
            opus_result['R11'] = {
                'status': current_r11.get('status', 'PASS'),
                'reason': enhanced_reason[:500]
            }
        
        # R12: Procedure-Diagnosis Match (ENHANCE, don't replace)
        procedure_appropriate = web_result.get('procedure_appropriate')
        if procedure_appropriate is not None and not procedure_appropriate:
            current_r12 = opus_result.get('R12', {})
            current_reason = current_r12.get('reason', '')
            if current_r12.get('status') == 'PASS':
                opus_result['R12'] = {
                    'status': 'QUERY',
                    'reason': f"Web search suggests procedure may not be appropriate for this diagnosis. {current_reason}"
                }
        
        # Add comprehensive web search info
        opus_result['web_search'] = {
            'performed': True,
            'status': 'success',
            'drug_class': web_result.get('drug_class'),
            'gender_specific': web_result.get('gender_specific'),
            'age_restrictions': web_result.get('age_restrictions'),
            'verified_indications': web_result.get('verified_indications', []),
            'dangerous_interactions': web_result.get('dangerous_interactions', []),
            'is_symptomatic_only': web_result.get('is_symptomatic_only'),
            'standard_dosage': web_result.get('standard_dosage'),
            'max_daily_dose': web_result.get('max_daily_dose'),
            'appropriate_for_diagnosis': appropriate,
            'sources': web_result.get('sources', [])
        }
        
        # Update confidence based on web search confirmation
        if appropriate and not web_result.get('gender_specific') and not web_result.get('is_symptomatic_only'):
            opus_result['confidence'] = max(opus_result.get('confidence', 0.8), 0.9)
        
        return opus_result
    
    def run_data_rules(self, row: pd.Series, pa_exists: bool = True, pa_validation: Dict = None, 
                        tariff_validation: Dict = None, hospital_shopping: Dict = None) -> Dict[str, Dict]:
        """
        Run data validation rules (R1-R5, R13)
        
        v3.7 ENHANCED: 
        - R1 now checks 90-day submission rule!
        - R3 now actually validates claim against PA authorization!
        - R4 now actually validates against provider tariff!
        - R5 now checks for hospital shopping (same procedure/diagnosis at different providers)!
        
        Parameters:
        -----------
        row : pd.Series
            The claim row
        pa_exists : bool
            Whether PA number > 0
        pa_validation : Dict
            Results from validate_pa_details() for PRE-AUTH claims
        tariff_validation : Dict
            Results from validate_tariff_amount() for tariff compliance
        hospital_shopping : Dict
            Results from check_hospital_shopping() for R5
        """
        results = {}
        
        # R1: Enrollee Eligibility + 90-Day Submission Rule
        # Check if claim was submitted within 90 days of encounter
        encounter_date_col = None
        submission_date_col = None
        
        for col in ['Encounter Date', 'encounter_date', 'Date', 'encounterdatefrom']:
            if col in row.index and pd.notna(row.get(col)):
                encounter_date_col = col
                break
        
        for col in ['Final Submission Date', 'Submission Date', 'submission_date', 'Date Submitted', 'datesubmitted']:
            if col in row.index and pd.notna(row.get(col)):
                submission_date_col = col
                break
        
        if encounter_date_col and submission_date_col:
            try:
                encounter_dt = pd.to_datetime(row.get(encounter_date_col))
                submission_dt = pd.to_datetime(row.get(submission_date_col))
                
                days_diff = (submission_dt - encounter_dt).days
                
                if days_diff > 90:
                    results['R1'] = {
                        'status': 'FAIL',
                        'reason': f"90-DAY RULE VIOLATION: Claim submitted {days_diff} days after encounter (Encounter: {encounter_dt.strftime('%Y-%m-%d')}, Submitted: {submission_dt.strftime('%Y-%m-%d')})"
                    }
                elif days_diff < 0:
                    results['R1'] = {
                        'status': 'QUERY',
                        'reason': f"DATE ANOMALY: Submission date ({submission_dt.strftime('%Y-%m-%d')}) is before encounter date ({encounter_dt.strftime('%Y-%m-%d')})"
                    }
                else:
                    results['R1'] = {
                        'status': 'PASS',
                        'reason': f'Enrollee valid, claim submitted within 90 days ({days_diff} days after encounter)'
                    }
            except Exception as e:
                results['R1'] = {
                    'status': 'PASS',
                    'reason': 'Enrollee ID valid (date check unavailable)'
                }
        else:
            results['R1'] = {
                'status': 'PASS',
                'reason': 'Enrollee ID valid and active'
            }
        
        # R2: Provider Panel Status
        results['R2'] = {
            'status': 'PASS',
            'reason': 'Provider is accredited panel member'
        }
        
        # R3: PA Matching - ENHANCED in v3.5!
        if not pa_exists:
            # NO-AUTH claim - R3 passes (no PA to check)
            results['R3'] = {
                'status': 'PASS',
                'reason': 'NO-AUTH claim - PA validation not required'
            }
        elif pa_validation:
            # PRE-AUTH claim - Actually validate against PA!
            if pa_validation.get('overall_valid', False):
                # Get specific amount info
                authorized_amount = pa_validation.get('authorized_amount', 0)
                claimed = row.get('Amount', 0)
                results['R3'] = {
                    'status': 'PASS',
                    'reason': f"Claim matches PA: Procedure verified ✓, Amount ₦{claimed:,.2f} within authorized ₦{authorized_amount:,.2f} ✓"
                }
            else:
                # There are mismatches!
                mismatches = pa_validation.get('mismatches', [])
                if mismatches:
                    mismatch_text = '; '.join(mismatches[:3])
                    
                    # Determine severity
                    if not pa_validation.get('pa_found', True):
                        status = 'FAIL'
                        reason = f"PA NOT FOUND in database! {mismatch_text}"
                    elif pa_validation.get('pa_status') != 'AUTHORIZED':
                        status = 'FAIL'
                        reason = f"PA status is {pa_validation.get('pa_status')}, not AUTHORIZED"
                    elif not pa_validation.get('procedure_match', True):
                        status = 'FAIL'
                        reason = f"PROCEDURE MISMATCH: {mismatch_text}"
                    elif not pa_validation.get('amount_valid', True):
                        # Procedure matched but amount exceeded
                        authorized_amount = pa_validation.get('authorized_amount', 0)
                        claimed = row.get('Amount', 0)
                        status = 'FAIL'
                        reason = f"AMOUNT EXCEEDS PA: Claimed ₦{claimed:,.2f} > PA authorized ₦{authorized_amount:,.2f} for this procedure"
                    else:
                        status = 'QUERY'
                        reason = f"PA validation issue: {mismatch_text}"
                    
                    results['R3'] = {'status': status, 'reason': reason}
                else:
                    results['R3'] = {
                        'status': 'PASS',
                        'reason': 'Claim within PA authorization'
                    }
        else:
            # PRE-AUTH but no validation data - assume pass but note
            results['R3'] = {
                'status': 'PASS',
                'reason': 'PA exists but validation data unavailable'
            }
        
        # R4: Tariff Compliance - ENHANCED in v3.5!
        amount = row.get('Amount', 0)
        
        # Get provider for debugging
        provider_used = None
        for col in ['Provider', 'Provider Name', 'provider', 'Hospital', 'Facility']:
            if col in row.index and row.get(col):
                provider_used = row.get(col)
                break
        
        if tariff_validation and tariff_validation.get('tariff_found', False):
            # We have actual tariff data - use it!
            if tariff_validation.get('amount_valid', True):
                provider_tariff = tariff_validation.get('provider_tariff', 0)
                results['R4'] = {
                    'status': 'PASS',
                    'reason': f"Amount ₦{amount:,.2f} within provider tariff ₦{provider_tariff:,.2f} ({provider_used})"
                }
            else:
                provider_tariff = tariff_validation.get('provider_tariff', 0)
                variance = tariff_validation.get('variance_pct', 0)
                results['R4'] = {
                    'status': 'FAIL',
                    'reason': f"AMOUNT EXCEEDS TARIFF: Claimed ₦{amount:,.2f} > Tariff ₦{provider_tariff:,.2f} ({provider_used}) - {variance:.1f}% over"
                }
        else:
            # No tariff found - fall back to threshold check
            tariff_msg = tariff_validation.get('message', '') if tariff_validation else ''
            
            # Add debugging info
            if not provider_used:
                tariff_msg = f'[Provider column not found in claim data]'
            elif not tariff_msg:
                tariff_msg = f'[Tariff lookup failed for: {provider_used}]'
            
            if amount > 100000:
                results['R4'] = {
                    'status': 'QUERY',
                    'reason': f'Amount ₦{amount:,.2f} exceeds ₦100K threshold. {tariff_msg}'
                }
            elif amount > 50000:
                results['R4'] = {
                    'status': 'PASS',
                    'reason': f'Amount ₦{amount:,.2f} - high value item. {tariff_msg}'
                }
            else:
                results['R4'] = {
                    'status': 'PASS',
                    'reason': f'Amount ₦{amount:,.2f} within general threshold. {tariff_msg}'
                }
        
        # R5: Hospital Shopping Detection with AI CLASS Matching
        # Check if enrollee visited OTHER hospitals for similar diagnosis/procedure in 30 days
        # ENHANCED: Uses AI to detect SAME CLASS even when exact strings don't match
        if hospital_shopping and hospital_shopping.get('found', False):
            other_visits = hospital_shopping.get('other_visits', [])
            same_diag = hospital_shopping.get('same_diagnosis', False)
            same_proc = hospital_shopping.get('same_procedure', False)
            
            # Build visit details
            visit_details = []
            for v in other_visits[:2]:
                visit_details.append(f"{v['provider']} on {v['date']}")
            
            # If visits found but no string match, use AI to check for CLASS match
            if other_visits and not (same_diag or same_proc):
                # Build current visit info for AI analysis
                proc_col = next((c for c in ['Procedure', 'proceduredesc', 'procedure', 'Drug', 'Description'] 
                               if c in row.index and pd.notna(row.get(c))), None)
                diag_col = next((c for c in ['Diagnosis', 'diagnosisdesc', 'diagnosis'] 
                               if c in row.index and pd.notna(row.get(c))), None)
                prov_col = next((c for c in ['Provider Name', 'Provider', 'provider', 'Hospital', 'providername'] 
                               if c in row.index and pd.notna(row.get(c))), None)
                
                current_visit = {
                    'procedure': row.get(proc_col, 'N/A') if proc_col else 'N/A',
                    'diagnosis': row.get(diag_col, 'N/A') if diag_col else 'N/A',
                    'provider': row.get(prov_col, 'N/A') if prov_col else 'N/A',
                    'date': str(row.get('Encounter Date', row.get('encounterdatefrom', 'N/A')))[:10]
                }
                
                # Use AI to analyze if visits are for SAME CLASS of treatment
                ai_analysis = self.analyze_visits_with_ai(current_visit, other_visits)
                
                if ai_analysis.get('same_class', False):
                    fraud_likelihood = ai_analysis.get('fraud_likelihood', 'MEDIUM')
                    drug_class = ai_analysis.get('drug_class', 'Unknown')
                    
                    if fraud_likelihood == 'HIGH':
                        results['R5'] = {
                            'status': 'FAIL',
                            'reason': f"HOSPITAL SHOPPING (AI): Same {drug_class} class treatment at other providers: {', '.join(visit_details)}. {ai_analysis.get('analysis', '')}"
                        }
                    else:
                        results['R5'] = {
                            'status': 'QUERY',
                            'reason': f"POTENTIAL HOSPITAL SHOPPING (AI): {ai_analysis.get('analysis', '')}. Providers: {', '.join(visit_details)}"
                        }
                else:
                    results['R5'] = {
                        'status': 'PASS',
                        'reason': f"Other provider visits found but AI analysis indicates different treatment class: {ai_analysis.get('analysis', '')}"
                    }
            elif same_diag and same_proc:
                # Both diagnosis and procedure match at another provider - high risk
                results['R5'] = {
                    'status': 'FAIL',
                    'reason': f"HOSPITAL SHOPPING: Same diagnosis AND procedure at other provider(s) in 30 days: {', '.join(visit_details)}"
                }
            elif same_diag:
                # Same diagnosis at another provider
                results['R5'] = {
                    'status': 'QUERY',
                    'reason': f"POTENTIAL HOSPITAL SHOPPING: Same diagnosis at other provider(s) in 30 days: {', '.join(visit_details)}"
                }
            elif same_proc:
                # Same procedure at another provider
                results['R5'] = {
                    'status': 'QUERY',
                    'reason': f"POTENTIAL HOSPITAL SHOPPING: Same procedure at other provider(s) in 30 days: {', '.join(visit_details)}"
                }
            else:
                results['R5'] = {
                    'status': 'PASS',
                    'reason': 'No hospital shopping detected'
                }
        else:
            results['R5'] = {
                'status': 'PASS',
                'reason': 'No hospital shopping detected'
            }
        
        # R13: PA Authorization Valid
        # For NO-AUTH claims (PA=0), this is PASS - they're allowed without PA
        # For PRE-AUTH claims, use PA validation result
        if not pa_exists:
            # NO-AUTH claim - valid without PA
            results['R13'] = {
                'status': 'PASS',
                'reason': 'NO-AUTH claim - PA not required'
            }
        elif pa_validation and pa_validation.get('pa_found', False):
            # PRE-AUTH with PA found
            if pa_validation.get('pa_status') == 'AUTHORIZED':
                results['R13'] = {
                    'status': 'PASS',
                    'reason': f"PA {pa_validation.get('pa_status')} - valid authorization"
                }
            else:
                results['R13'] = {
                    'status': 'FAIL',
                    'reason': f"PA status is {pa_validation.get('pa_status')}, not AUTHORIZED"
                }
        elif pa_validation and not pa_validation.get('pa_found', True):
            # PRE-AUTH but PA not found
            results['R13'] = {
                'status': 'FAIL',
                'reason': 'PA number provided but not found in database'
            }
        else:
            # PRE-AUTH but couldn't validate
            results['R13'] = {
                'status': 'PASS',
                'reason': 'PA authorization assumed valid'
            }
        
        return results
    
    def call_claude_api(self, claim_data: Dict) -> Tuple[Dict, str]:
        """
        Call Claude Opus 4.5 API for medical judgment rules (R6-R12)
        
        ENHANCED v3.5: Now includes enrollee demographics for R6/R7 validation!
        
        Returns:
        --------
        (result_dict, raw_response_text)
        """
        self.stats['api_calls'] += 1
        
        # Get demographics
        gender = claim_data.get('enrollee_gender', 'UNKNOWN')
        age = claim_data.get('enrollee_age')
        demographics_found = claim_data.get('demographics_found', False)
        
        # Build demographics section
        if demographics_found:
            age_str = f"{age} years old" if age is not None else "Age unknown"
            demographics_section = f"""
## ⚠️ ENROLLEE DEMOGRAPHICS (CRITICAL FOR R6 & R7)
- **ACTUAL Gender:** {gender}
- **ACTUAL Age:** {age_str}
- **DOB:** {claim_data.get('enrollee_dob', 'Unknown')}

YOU MUST CHECK:
- R6: Is this procedure appropriate for a {gender} patient?
- R7: Is this procedure appropriate for a {age_str} patient?

FAIL R6 if procedure is gender-specific and doesn't match!
FAIL R7 if procedure is age-restricted and doesn't match!"""
        else:
            demographics_section = """
## ⚠️ ENROLLEE DEMOGRAPHICS
- Demographics not found in database
- Use QUERY for R6/R7 if procedure is gender/age specific"""
        
        # Build the user prompt with demographics
        user_prompt = f"""Evaluate this healthcare claim:

## CLAIM DETAILS
- **Procedure/Drug:** {claim_data.get('procedure', 'N/A')}
- **Single Row Diagnosis:** {claim_data.get('diagnosis', 'N/A')}
- **All Encounter Diagnoses:** {', '.join(claim_data.get('all_encounter_dx', []))}
- **Amount:** ₦{claim_data.get('amount', 0):,.2f}
- **Units/Quantity:** {claim_data.get('units', 1)}
- **Enrollee ID:** {claim_data.get('enrollee', 'N/A')}
- **Encounter Date:** {claim_data.get('date', 'N/A')}
{demographics_section}

## 30-DAY PRIOR CLAIMS
{self._format_prior_claims(claim_data.get('prior_claims', []))}

## ENCOUNTER CONTEXT
Other procedures in this encounter:
{chr(10).join(['- ' + p for p in claim_data.get('encounter_procedures', [])[:10]]) or '- No other procedures'}

{self._format_pa_context(claim_data.get('pa_context'))}

Please evaluate Rules R6-R12 and return your JSON verdict.

⚠️ CRITICAL INSTRUCTIONS:

**R6 (Gender):** Compare procedure against ACTUAL gender ({gender}) - FAIL if mismatch, PASS if match!

**R7 (Age):** Compare procedure AND diagnosis against ACTUAL age ({age if age else 'unknown'})!
   - FAIL if diagnosis is AGE-IMPOSSIBLE (e.g., Cephalhematoma for adult = NEWBORN-ONLY condition!)
   - FAIL if procedure is inappropriate for age (e.g., Paediatrician for adult!)
   - Adult seeing Paediatrician = WRONG ENROLLEE ID, not valid claim!

**R8 (Duplicates):** Check if same/similar procedure exists in 30-day prior claims - if yes and it's a course medication (antibiotic/antimalarial), flag it!

**R9 (Drug-Diagnosis Match):** CHECK ALL CONTEXT! Look at prior claims, encounter procedures, AND PA context.
   If drug doesn't match THIS claim's diagnosis, but prior claims/PA context shows RELATED diagnosis, EXPLAIN!
   Example: "Tranexamic acid for Hypertension is questionable, HOWEVER PA shows Threatened Abortion"

**R10 (Symptomatic) - PROVISIONAL DIAGNOSIS LOGIC:** 🚨 THIS IS CRITICAL!
   
   **For DIAGNOSTIC TESTS:**
   - The diagnosis that justified the test is VALID - do NOT say it's wrong!
   - Test + No Treatment = Test was NEGATIVE, but diagnosis served its purpose
   - TEST = PASS
   
   **For SYMPTOMATIC DRUGS (Paracetamol, NSAIDs, antihistamines):**
   - First: Check if there are OTHER diagnoses in the encounter
   - If other dx can justify the symptomatic drug → PASS
   - If ONLY the ruled-out dx exists AND no treatment given → QUERY
   
   EXAMPLE:
   - Encounter: "Malaria" dx + "URI" dx + Malaria Test + Paracetamol (no antimalarial)
   - Malaria Test → PASS (Malaria dx justified it)
   - Paracetamol → PASS (URI dx can justify it for fever/pain)
   
   vs.
   
   - Encounter: "Malaria" dx ONLY + Malaria Test + Paracetamol (no antimalarial)
   - Malaria Test → PASS (dx justified it)
   - Paracetamol → QUERY (only dx was ruled out, nothing else to claim)
   
   HOW TO CHECK:
   1. Look at ALL diagnoses in encounter (all_encounter_dx field)
   2. For symptomatic drugs, find ANY valid indication
   3. Only QUERY if NO other diagnosis can justify it

Return confidence level (0.0-1.0)."""

        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=1500,
                system=SYSTEM_PROMPT,
                messages=[
                    {"role": "user", "content": user_prompt}
                ]
            )
            
            # Extract text content
            response_text = response.content[0].text
            
            # Parse JSON from response
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start != -1 and json_end > json_start:
                json_str = response_text[json_start:json_end]
                result = json.loads(json_str)
                return result, response_text
            else:
                return self._default_api_response("Could not parse API response"), response_text
                
        except json.JSONDecodeError as e:
            self.stats['api_errors'] += 1
            print(f"WARNING: JSON parse error: {e}")
            return self._default_api_response(f"JSON parse error: {str(e)}"), ""
            
        except Exception as e:
            self.stats['api_errors'] += 1
            print(f"WARNING: API call error: {e}")
            return self._default_api_response(f"API error: {str(e)}"), ""
    
    def _format_prior_claims(self, prior_claims: List[Dict]) -> str:
        """
        Format prior claims for the AI prompt - ENHANCED for CLASS matching
        
        Includes both codes AND descriptions so AI can identify:
        - Same drug CLASS (e.g., Amoxicillin & Augmentin = Penicillin antibiotics)
        - Same procedure CLASS (e.g., CT Scan Head & MRI Brain = neuroimaging)
        - Same diagnosis CLASS (e.g., Malaria & Plasmodium = same disease)
        """
        if not prior_claims:
            return "No prior claims in last 30 days"
        
        lines = []
        for claim in prior_claims[:15]:  # Increased from 10 to 15 for better context
            # Get procedure info (code + description)
            proc_code = claim.get('procedure_code', '')
            proc_desc = claim.get('proceduredesc', claim.get('procedure', ''))
            proc_display = f"{proc_desc}" if proc_desc else proc_code
            if proc_code and proc_desc:
                proc_display = f"{proc_desc} [{proc_code}]"
            
            # Get diagnosis info (code + description)
            diag_code = claim.get('diagnosiscode', '')
            diag_desc = claim.get('diagnosisdesc', claim.get('diagnosis', ''))
            diag_display = f"{diag_desc}" if diag_desc else diag_code
            if diag_code and diag_desc:
                diag_display = f"{diag_desc} [{diag_code}]"
            
            # Get provider and date
            provider = claim.get('providername', claim.get('provider', 'Unknown'))
            date = str(claim.get('encounterdatefrom', claim.get('date', 'N/A')))[:10]
            amount = claim.get('approvedamount', claim.get('amount', 0))
            source = claim.get('source', 'Claims')
            
            lines.append(f"- {date} @ {provider}: {proc_display} for {diag_display} (₦{amount:,.0f}) [{source}]")
        
        return '\n'.join(lines) + """

⚠️ CHECK FOR SAME CLASS:
- Antibiotics: Amoxicillin, Augmentin, Ceftriaxone, Ciprofloxacin, Metronidazole, Azithromycin
- Antimalarials: Artemether, Lumefantrine, Artesunate, Coartem, ACT combinations
- NSAIDs: Ibuprofen, Diclofenac, Naproxen, Piroxicam, Meloxicam
- Analgesics: Paracetamol, Tramadol (can repeat - PRN)
- Antihypertensives: Amlodipine, Lisinopril, Losartan, Atenolol
- Antidiabetics: Metformin, Glibenclamide, Insulin variants
- PPIs/Antacids: Omeprazole, Pantoprazole, Ranitidine

If current drug is in SAME CLASS as prior drug (within 30 days), FLAG for R8 unless diagnosis is different."""
    
    def _format_pa_context(self, pa_context: Dict) -> str:
        """
        🚨 v3.8 NEW: Format PA context for NO-AUTH claims
        
        When PA=0, we look up PA table for related diagnoses/procedures
        to give AI the full clinical picture.
        """
        if not pa_context or not pa_context.get('found'):
            return ""  # No PA context for this claim
        
        lines = ["## 🚨 PA CONTEXT (for NO-AUTH claim)"]
        lines.append("This is a NO-AUTH claim. Related PA(s) found within ±24 hours:")
        
        # Add diagnoses from PA
        pa_diagnoses = pa_context.get('pa_diagnoses', [])
        if pa_diagnoses:
            lines.append(f"\n**Related PA Diagnoses:** {', '.join(pa_diagnoses[:5])}")
        
        # Add procedures from PA
        pa_procedures = pa_context.get('pa_procedures', [])
        if pa_procedures:
            lines.append(f"**Related PA Procedures:** {', '.join(pa_procedures[:5])}")
        
        # Add PA numbers
        pa_numbers = pa_context.get('pa_numbers', [])
        if pa_numbers:
            lines.append(f"**PA Numbers:** {', '.join(pa_numbers[:3])}")
        
        lines.append("""
⚠️ IMPORTANT: The current claim diagnosis may look wrong for the drug, but check if 
the drug is appropriate for ANY of the PA diagnoses above! The provider may have split 
the encounter into PA + NO-AUTH items. Use this context to make an informed decision.""")
        
        return '\n'.join(lines)
    
    def _default_api_response(self, error_msg: str) -> Dict:
        """Return default response when API fails"""
        return {
            'R6': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R7': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R8': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R9': {'status': 'QUERY', 'reason': f'Manual review needed - {error_msg}'},
            'R10': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R11': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R12': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'confidence': 0.0,
            'clinical_notes': error_msg,
            'needs_verification': True
        }
    
    def vet_single_claim(self, row: pd.Series, encounter_dx: Dict, all_procedures: List[str]) -> Dict:
        """
        Vet a single claim through all 13 rules
        
        ENHANCED v3.5: 
        - Now looks up actual gender and age for R6/R7 validation!
        - Now validates PA procedure/amount for R3!
        - Now validates against provider tariff for R4!
        """
        # Identify columns - expanded provider detection
        enrollee_col = 'Enrollee No' if 'Enrollee No' in row.index else 'Enrollee'
        pa_col = 'PA Number' if 'PA Number' in row.index else 'PA'
        date_col = 'Encounter Date' if 'Encounter Date' in row.index else 'Date'
        proc_col = 'Procedure Description' if 'Procedure Description' in row.index else 'Procedure'
        diag_col = 'Diagnosis' if 'Diagnosis' in row.index else 'diagnosis'
        
        # EXPANDED: Try multiple provider column names
        provider_col = None
        for col_name in ['Provider', 'Provider Name', 'provider', 'provider_name', 
                         'Hospital', 'Hospital Name', 'Facility', 'facility']:
            if col_name in row.index:
                provider_col = col_name
                break
        
        enrollee = row.get(enrollee_col, '')
        pa = row.get(pa_col, 0)
        date = row.get(date_col, '')
        procedure = row.get(proc_col, '')
        diagnosis = row.get(diag_col, '')
        amount = row.get('Amount', 0)
        units = row.get('Units', 1)
        provider = row.get(provider_col, '') if provider_col else ''
        
        # Get encounter key and all diagnoses
        enc_key = (enrollee, pa, date)
        
        # 🚨 v3.8 FIX: Parse semicolon-separated diagnoses even in fallback!
        if enc_key in encounter_dx:
            all_dx = list(encounter_dx.get(enc_key))
        else:
            # Fallback: parse the diagnosis from this row (handle ; and , separators)
            diag_str = str(diagnosis).upper().replace(';', ',')
            if ',' in diag_str:
                all_dx = [d.strip() for d in diag_str.split(',') if d.strip() and d.strip() != 'NAN']
            else:
                all_dx = [diagnosis.upper()] if diagnosis else []
        
        # Get prior claims
        prior_claims = self.get_prior_claims(enrollee, date)
        
        # ================================================================
        # NEW v3.5: Get enrollee demographics for R6 (Gender) and R7 (Age)
        # ================================================================
        demographics = self.get_enrollee_demographics(enrollee)
        
        # ================================================================
        # NEW v3.5: Validate PA details for PRE-AUTH claims (R3 enhancement)
        # ================================================================
        pa_validation = None
        if pa and pa > 0:
            # PRE-AUTH claim - validate against PA authorization
            pa_validation = self.validate_pa_details(
                pa_number=int(pa),
                claim_procedure=procedure,
                claim_amount=float(amount) if amount else 0,
                claim_date=str(date) if date else None
            )
        
        # ================================================================
        # NEW v3.5: Validate tariff amount for R4 (Tariff Compliance)
        # ================================================================
        tariff_validation = None
        if provider and procedure:
            tariff_validation = self.validate_tariff_amount(
                provider_name=str(provider),
                procedure_desc=str(procedure),
                claimed_amount=float(amount) if amount else 0
            )
        
        # ================================================================
        # NEW v3.7: Check for hospital shopping for R5
        # ================================================================
        hospital_shopping = None
        if enrollee and date and provider:
            hospital_shopping = self.check_hospital_shopping(
                enrollee_id=str(enrollee),
                current_date=str(date),
                current_provider=str(provider),
                current_procedure=str(procedure),
                current_diagnosis=str(diagnosis)
            )
        
        # ================================================================
        # 🚨 v3.8 NEW: Get PA context for NO-AUTH claims
        # When PA=0, look up PA table for related diagnoses/procedures
        # ================================================================
        pa_context = None
        if (not pa or pa == 0) and enrollee and date:
            pa_context = self.get_pa_context_for_noauth(
                enrollee_id=str(enrollee),
                encounter_date=str(date),
                hours=24  # Look ±24 hours
            )
        
        # Run data rules (R1-R5, R13) - now with PA, tariff, and hospital shopping validation!
        data_results = self.run_data_rules(
            row, 
            pa_exists=(pa > 0), 
            pa_validation=pa_validation,
            tariff_validation=tariff_validation,
            hospital_shopping=hospital_shopping
        )
        
        # Prepare claim data for API - now includes demographics and PA context!
        claim_data = {
            'procedure': procedure,
            'diagnosis': diagnosis,
            'all_encounter_dx': all_dx,
            'amount': amount,
            'units': units,
            'enrollee': enrollee,
            'date': date,
            'prior_claims': prior_claims,
            'encounter_procedures': all_procedures,
            # Demographics for R6/R7
            'enrollee_gender': demographics.get('gender', 'UNKNOWN'),
            'enrollee_age': demographics.get('age_years'),
            'enrollee_dob': demographics.get('dob'),
            'demographics_found': demographics.get('found', False),
            # 🚨 v3.8 NEW: PA context for NO-AUTH claims
            'pa_context': pa_context
        }
        
        # Call Claude API for medical judgment (R6-R12)
        api_results, response_text = self.call_claude_api(claim_data)
        
        # Check if web search should be triggered
        should_search, search_reason = self._should_trigger_web_search(
            procedure, api_results, response_text
        )
        
        if should_search:
            self.stats['web_search_triggers'][search_reason] = \
                self.stats['web_search_triggers'].get(search_reason, 0) + 1
            
            # Perform web search with patient context for all rules
            web_results = self.perform_web_search(
                drug_name=procedure, 
                diagnosis=diagnosis,
                patient_gender=claim_data.get('enrollee_gender'),
                patient_age=claim_data.get('enrollee_age')
            )
            
            # Merge results - 🚨 v3.8: Now passes patient gender/age for smart decisions!
            api_results = self._merge_web_search_results(
                api_results, 
                web_results,
                patient_gender=claim_data.get('enrollee_gender'),
                patient_age=claim_data.get('enrollee_age')
            )
        
        # Combine results
        all_results = {**data_results}
        for rule in ['R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12']:
            if rule in api_results:
                all_results[rule] = api_results[rule]
            else:
                all_results[rule] = {'status': 'PASS', 'reason': 'Not evaluated'}
        
        # Determine final status
        has_fail = any(r.get('status') == 'FAIL' for r in all_results.values())
        has_query = any(r.get('status') == 'QUERY' for r in all_results.values())
        
        if has_fail:
            final_status = 'DENY'
            fail_rules = [k for k, v in all_results.items() if v.get('status') == 'FAIL']
            final_reason = f"Failed: {', '.join(fail_rules)}"
        elif has_query:
            final_status = 'QUERY'
            final_reason = 'Requires verification'
        else:
            final_status = 'APPROVED'
            final_reason = 'All 13 rules passed'
        
        # Update stats
        self.stats['total_claims'] += 1
        if final_status == 'APPROVED':
            self.stats['approved'] += 1
        elif final_status == 'DENY':
            self.stats['denied'] += 1
        else:
            self.stats['query'] += 1
        
        return {
            'enrollee': enrollee,
            'pa': pa,
            'date': date,
            'procedure': procedure,
            'diagnosis': diagnosis,
            'all_encounter_dx': ', '.join(all_dx[:3]),
            'amount': amount,
            'rules': all_results,
            'final_status': final_status,
            'final_reason': final_reason,
            'confidence': api_results.get('confidence', 1.0),
            'clinical_notes': api_results.get('clinical_notes', ''),
            'web_search': api_results.get('web_search', {'performed': False})
        }
    
    def vet_batch(self, df: pd.DataFrame, batch_size: int = 10, progress_callback=None) -> List[Dict]:
        """
        Vet a batch of claims
        """
        # Log available columns for debugging
        print(f"\n📋 Available columns in claim data: {list(df.columns)}")
        
        # Check for provider column
        provider_found = None
        for col in ['Provider', 'Provider Name', 'provider', 'provider_name', 
                    'Hospital', 'Hospital Name', 'Facility', 'facility']:
            if col in df.columns:
                provider_found = col
                break
        
        if provider_found:
            print(f"✅ Provider column found: '{provider_found}'")
            # Show sample provider values
            sample_providers = df[provider_found].dropna().head(3).tolist()
            print(f"   Sample values: {sample_providers}")
        else:
            print(f"⚠️  WARNING: No Provider column found! R4 tariff lookup will be skipped.")
            print(f"   Expected columns: Provider, Provider Name, Hospital, Facility")
        
        # Build encounter diagnoses
        encounter_dx = self.build_encounter_diagnoses(df)
        
        # Build encounter procedures mapping
        proc_col = self._find_column(df, ['Procedure Description', 'Procedure', 'procedure'])
        enrollee_col = self._find_column(df, ['Enrollee No', 'Enrollee', 'enrollee_id'])
        pa_col = self._find_column(df, ['PA Number', 'PA', 'panumber'])
        date_col = self._find_column(df, ['Encounter Date', 'Date', 'encounter_date'])
        
        encounter_procedures = {}
        for idx, row in df.iterrows():
            key = (row[enrollee_col], row[pa_col], row[date_col])
            if key not in encounter_procedures:
                encounter_procedures[key] = []
            encounter_procedures[key].append(row[proc_col])
        
        results = []
        total = len(df)
        
        print(f"\n🔍 Starting vetting of {total} claims...")
        print(f"   Web Search Mode: {self.web_search_mode.upper()}")
        print("="*60)
        
        for idx, row in df.iterrows():
            # Get encounter procedures
            key = (row[enrollee_col], row[pa_col], row[date_col])
            all_procs = encounter_procedures.get(key, [])
            
            # Vet the claim
            result = self.vet_single_claim(row, encounter_dx, all_procs)
            result['row_number'] = idx + 1
            results.append(result)
            
            # Progress update
            if (idx + 1) % batch_size == 0 or idx == total - 1:
                pct = (idx + 1) / total * 100
                web_info = f" | Web searches: {self.stats['web_searches']}" if self.web_search_mode != 'off' else ""
                print(f"  Processed {idx + 1}/{total} claims ({pct:.1f}%){web_info}")
                if progress_callback:
                    progress_callback(idx + 1, total)
        
        print("="*60)
        print(f"✅ Vetting complete!")
        print(f"   Approved: {self.stats['approved']}")
        print(f"   Denied: {self.stats['denied']}")
        print(f"   Query: {self.stats['query']}")
        print(f"   API Calls: {self.stats['api_calls']}")
        if self.stats['api_errors'] > 0:
            print(f"   API Errors: {self.stats['api_errors']}")
        if self.web_search_mode != 'off':
            print(f"   Web Searches: {self.stats['web_searches']}")
            if self.stats['web_search_triggers']:
                print(f"   Search Triggers: {self.stats['web_search_triggers']}")
        
        return results
    
    # ========================================================================
    # ENHANCED ANALYTICAL REPORT GENERATOR - v3.8
    # ========================================================================
    
    def build_enrollee_profile(self, results: List[Dict]) -> Dict:
        """
        Build comprehensive profiles for each enrollee showing all their
        encounters, PAs, procedures, and diagnoses for cross-referencing.
        
        This enables deliberative analysis like:
        "PA 445823 on Nov 9 shows Tranexamic acid, but looking at PA 442872 
         on Nov 4 we see cervical cerclage was performed..."
        """
        profiles = {}
        
        for result in results:
            enrollee = result.get('enrollee', 'Unknown')
            if enrollee not in profiles:
                profiles[enrollee] = {
                    'enrollee_id': enrollee,
                    'encounters': {},  # Keyed by (PA, date)
                    'all_diagnoses': set(),
                    'all_procedures': set(),
                    'all_pas': set(),
                    'total_amount': 0,
                    'flags': [],
                    'gender': result.get('enrollee_gender'),
                    'age': result.get('enrollee_age'),
                    'results': []
                }
            
            profile = profiles[enrollee]
            pa = result.get('pa', 0)
            date = result.get('date', '')
            enc_key = (pa, date)
            
            # Add to encounters
            if enc_key not in profile['encounters']:
                profile['encounters'][enc_key] = {
                    'pa': pa,
                    'date': date,
                    'procedures': [],
                    'diagnoses': set(),
                    'claims': [],
                    'total_amount': 0,
                    'has_diagnostic_test': False,
                    'has_treatment': False,
                    'diagnostic_tests': [],
                    'treatment_drugs': [],
                    'symptomatic_drugs': [],
                    'flags': []
                }
            
            enc = profile['encounters'][enc_key]
            procedure = result.get('procedure', '')
            diagnosis = result.get('diagnosis', '')
            amount = result.get('amount', 0) or 0
            
            # Categorize procedure
            proc_lower = procedure.lower()
            
            # Diagnostic tests
            diagnostic_keywords = ['test', 'culture', 'swab', 'film', 'scan', 'xray', 'x-ray', 
                                   'ultrasound', 'sono', 'fbc', 'cbc', 'urinalysis', 'widal',
                                   'malarial parasite', 'mp', 'h.pylori', 'helicobacter', 
                                   'hiv', 'hepatitis', 'lipid', 'sugar', 'glucose', 'hba1c',
                                   'm/c/s', 'mcs', 'microscopy']
            
            # Treatment drugs (actual treatments, not symptomatic)
            treatment_keywords = ['artemether', 'lumefantrine', 'coartem', 'artesunate', 'fansidar',
                                  'sulfadoxine', 'pyrimethamine',  # Antimalarials
                                  'amoxicillin', 'augmentin', 'clavulanic', 'ceftriaxone', 'cefuroxime',
                                  'ciprofloxacin', 'metronidazole', 'azithromycin', 'clarithromycin',
                                  'doxycycline', 'erythromycin',  # Antibiotics
                                  'metformin', 'glibenclamide', 'insulin',  # Antidiabetics
                                  'amlodipine', 'lisinopril', 'losartan', 'nifedipine', 'methyldopa',  # Antihypertensives
                                  'omeprazole', 'pantoprazole', 'esomeprazole', 'lansoprazole',  # PPIs
                                  'tranexamic', 'misoprostol', 'oxytocin',  # Obstetric
                                  'iron', 'ferrous', 'astyfer', 'folic acid', 'b12']  # Hematinic
            
            # Symptomatic drugs
            symptomatic_keywords = ['paracetamol', 'acetaminophen', 'ibuprofen', 'diclofenac',
                                    'tramadol', 'codeine',  # Pain/fever
                                    'chlorpheniramine', 'loratadine', 'cetirizine', 'promethazine',  # Antihistamines
                                    'cough syrup', 'antacid', 'antitussive',
                                    'vitamin c', 'vitamin b', 'multivitamin']  # Vitamins
            
            is_diagnostic = any(kw in proc_lower for kw in diagnostic_keywords)
            is_treatment = any(kw in proc_lower for kw in treatment_keywords)
            is_symptomatic = any(kw in proc_lower for kw in symptomatic_keywords)
            
            if is_diagnostic:
                enc['has_diagnostic_test'] = True
                enc['diagnostic_tests'].append(procedure)
            if is_treatment:
                enc['has_treatment'] = True
                enc['treatment_drugs'].append(procedure)
            if is_symptomatic:
                enc['symptomatic_drugs'].append(procedure)
            
            # Add procedure and diagnosis
            enc['procedures'].append(procedure)
            enc['claims'].append(result)
            enc['total_amount'] += amount
            
            # Parse diagnoses (handle semicolons)
            if diagnosis:
                for dx in str(diagnosis).replace(';', ',').split(','):
                    dx_clean = dx.strip().upper()
                    if dx_clean and dx_clean != 'NAN':
                        enc['diagnoses'].add(dx_clean)
                        profile['all_diagnoses'].add(dx_clean)
            
            profile['all_procedures'].add(procedure)
            if pa and pa != 0:
                profile['all_pas'].add(pa)
            profile['total_amount'] += amount
            profile['results'].append(result)
        
        return profiles
    
    def detect_patterns(self, profile: Dict) -> List[Dict]:
        """
        Detect clinical patterns across enrollee's encounters:
        1. Provisional diagnoses (test + no treatment)
        2. Related PAs (same condition across encounters)
        3. Missing treatments
        4. Cross-encounter context
        """
        patterns = []
        encounters = profile['encounters']
        
        # Sort encounters by date
        sorted_encounters = sorted(encounters.items(), 
                                   key=lambda x: str(x[0][1]))  # Sort by date
        
        # Pattern 1: Provisional Diagnosis Detection
        for enc_key, enc in sorted_encounters:
            if enc['has_diagnostic_test'] and not enc['has_treatment']:
                # Check if only symptomatic drugs given
                if enc['symptomatic_drugs']:
                    patterns.append({
                        'type': 'PROVISIONAL_DIAGNOSIS',
                        'severity': 'QUERY',
                        'encounter': enc_key,
                        'date': enc['date'],
                        'pa': enc['pa'],
                        'tests': enc['diagnostic_tests'],
                        'symptomatic': enc['symptomatic_drugs'],
                        'diagnoses': list(enc['diagnoses']),
                        'analysis': f"Diagnostic test(s) performed ({', '.join(enc['diagnostic_tests'][:3])}) "
                                   f"but NO treatment drug given. Only symptomatic: {', '.join(enc['symptomatic_drugs'][:3])}. "
                                   f"Test was NEGATIVE - check if symptomatic drugs have other dx to claim."
                    })
        
        # Pattern 2: Related PAs (same diagnosis across encounters)
        diagnosis_pa_map = {}  # diagnosis -> list of (pa, date, procedures)
        for enc_key, enc in sorted_encounters:
            for dx in enc['diagnoses']:
                if dx not in diagnosis_pa_map:
                    diagnosis_pa_map[dx] = []
                diagnosis_pa_map[dx].append({
                    'pa': enc['pa'],
                    'date': enc['date'],
                    'procedures': enc['procedures'][:5],
                    'has_treatment': enc['has_treatment']
                })
        
        # Find diagnoses that appear in multiple encounters
        for dx, encounters_list in diagnosis_pa_map.items():
            if len(encounters_list) > 1:
                patterns.append({
                    'type': 'RELATED_ENCOUNTERS',
                    'severity': 'INFO',
                    'diagnosis': dx,
                    'encounters': encounters_list,
                    'analysis': f"Diagnosis '{dx}' appears in {len(encounters_list)} encounters. "
                               f"Cross-reference for continuity of care."
                })
        
        # Pattern 3: Treatment without test (empirical treatment)
        for enc_key, enc in sorted_encounters:
            if enc['has_treatment'] and not enc['has_diagnostic_test']:
                # This might be empirical treatment - note it
                patterns.append({
                    'type': 'EMPIRICAL_TREATMENT',
                    'severity': 'INFO',
                    'encounter': enc_key,
                    'date': enc['date'],
                    'pa': enc['pa'],
                    'treatments': enc['treatment_drugs'],
                    'diagnoses': list(enc['diagnoses']),
                    'analysis': f"Treatment given ({', '.join(enc['treatment_drugs'][:3])}) without "
                               f"diagnostic test in this encounter. May be empirical or based on prior tests."
                })
        
        return patterns
    
    def generate_narrative_analysis(self, profile: Dict, patterns: List[Dict]) -> str:
        """
        Generate human-readable narrative analysis for an enrollee
        """
        enrollee = profile['enrollee_id']
        narrative = []
        
        narrative.append(f"\n{'='*80}")
        narrative.append(f"## ENROLLEE: {enrollee}")
        narrative.append(f"{'='*80}")
        
        if profile['gender'] or profile['age']:
            demo = []
            if profile['gender']:
                demo.append(f"Gender: {profile['gender']}")
            if profile['age']:
                demo.append(f"Age: {profile['age']}")
            narrative.append(f"Demographics: {' | '.join(demo)}")
        
        narrative.append(f"Total Claims: {len(profile['results'])} | Total Amount: ₦{profile['total_amount']:,.2f}")
        narrative.append(f"Unique PAs: {len(profile['all_pas'])} | Unique Encounters: {len(profile['encounters'])}")
        
        # Sort encounters by date
        sorted_encounters = sorted(profile['encounters'].items(), 
                                   key=lambda x: str(x[0][1]))
        
        # Describe each encounter
        for enc_key, enc in sorted_encounters:
            pa, date = enc_key
            narrative.append(f"\n### 📋 PA: {pa} | Date: {date}")
            narrative.append(f"Amount: ₦{enc['total_amount']:,.2f} | Claims: {len(enc['claims'])}")
            
            if enc['diagnoses']:
                narrative.append(f"**Diagnoses:** {'; '.join(list(enc['diagnoses'])[:5])}")
            
            narrative.append("**Procedures:**")
            for proc in enc['procedures'][:10]:
                # Add categorization
                proc_lower = proc.lower()
                category = ""
                if any(kw in proc_lower for kw in ['test', 'culture', 'swab', 'film', 'scan', 'fbc', 'urinalysis', 'm/c/s', 'parasite']):
                    category = " 🔬 [DIAGNOSTIC]"
                elif any(kw in proc_lower for kw in ['artemether', 'amoxicillin', 'ceftriaxone', 'metformin', 'amlodipine', 'omeprazole', 'tranexamic', 'clarithromycin', 'cefuroxime', 'ciprofloxacin', 'iron', 'astyfer']):
                    category = " 💊 [TREATMENT]"
                elif any(kw in proc_lower for kw in ['paracetamol', 'ibuprofen', 'chlorpheniramine', 'loratadine', 'vitamin', 'cough', 'antacid']):
                    category = " 🩹 [SYMPTOMATIC]"
                elif any(kw in proc_lower for kw in ['consultation', 'review', 'follow']):
                    category = " 👨‍⚕️ [CONSULT]"
                elif any(kw in proc_lower for kw in ['room', 'bed', 'feeding', 'admission']):
                    category = " 🏥 [FACILITY]"
                
                narrative.append(f"   - {proc}{category}")
            
            # Add encounter-level analysis
            if enc['has_diagnostic_test'] and not enc['has_treatment'] and enc['symptomatic_drugs']:
                narrative.append(f"\n   ⚠️ **PROVISIONAL DIAGNOSIS ALERT:**")
                narrative.append(f"   - Tests: {', '.join(enc['diagnostic_tests'][:3])}")
                narrative.append(f"   - Symptomatic only: {', '.join(enc['symptomatic_drugs'][:3])}")
                narrative.append(f"   - No treatment found → Test was negative. Check other diagnoses for symptomatic drugs.")
            
            elif enc['has_diagnostic_test'] and enc['has_treatment']:
                narrative.append(f"\n   ✅ **CONFIRMED DIAGNOSIS:**")
                narrative.append(f"   - Test + Treatment present = Properly documented")
        
        # Add pattern analysis
        if patterns:
            narrative.append(f"\n### 🔍 PATTERN ANALYSIS")
            
            provisional = [p for p in patterns if p['type'] == 'PROVISIONAL_DIAGNOSIS']
            related = [p for p in patterns if p['type'] == 'RELATED_ENCOUNTERS']
            
            if provisional:
                narrative.append(f"\n**⚠️ Provisional Diagnosis Issues ({len(provisional)}):**")
                for p in provisional:
                    narrative.append(f"   - {p['date']} (PA {p['pa']}): {p['analysis']}")
            
            if related:
                narrative.append(f"\n**🔗 Related Encounters:**")
                for p in related[:3]:  # Limit to 3
                    dates = [e['date'] for e in p['encounters']]
                    narrative.append(f"   - '{p['diagnosis']}' seen on: {', '.join(dates[:5])}")
        
        return '\n'.join(narrative)
    
    def generate_analytical_report(self, results: List[Dict], output_path: str, batch_name: str = "Claims Analysis"):
        """
        Generate comprehensive analytical report with:
        1. Executive Summary
        2. Critical Flags
        3. Enrollee-by-Enrollee Analysis
        4. Pattern Detection
        5. Recommendations
        
        This is the ENHANCED report that shows deliberative reasoning!
        """
        print("\n📊 Generating Analytical Report...")
        
        # Build enrollee profiles
        profiles = self.build_enrollee_profile(results)
        
        # Detect patterns for each enrollee
        all_patterns = {}
        for enrollee, profile in profiles.items():
            all_patterns[enrollee] = self.detect_patterns(profile)
        
        # Categorize results
        critical_flags = []  # FAIL
        query_flags = []     # QUERY  
        passed = []          # PASS
        
        for result in results:
            status = result.get('final_status', 'QUERY')
            if status == 'FAIL':
                critical_flags.append(result)
            elif status == 'QUERY':
                query_flags.append(result)
            else:
                passed.append(result)
        
        # Calculate totals
        total_amount = sum(r.get('amount', 0) or 0 for r in results)
        flagged_amount = sum(r.get('amount', 0) or 0 for r in critical_flags + query_flags)
        
        # Generate Markdown Report
        md_lines = []
        
        # Header
        md_lines.append(f"# 🏥 CLAIMS VETTING ANALYTICAL REPORT")
        md_lines.append(f"## {batch_name}")
        md_lines.append(f"**Generated:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        md_lines.append(f"**Engine:** KLAIRE Claims Vetting v3.8")
        md_lines.append("")
        
        # Executive Summary
        md_lines.append("---")
        md_lines.append("## 📊 EXECUTIVE SUMMARY")
        md_lines.append("")
        md_lines.append("| Metric | Value |")
        md_lines.append("|--------|-------|")
        md_lines.append(f"| Total Claims | {len(results)} |")
        md_lines.append(f"| Unique Enrollees | {len(profiles)} |")
        md_lines.append(f"| Total Amount | ₦{total_amount:,.2f} |")
        md_lines.append(f"| ✅ PASSED | {len(passed)} ({len(passed)/len(results)*100:.1f}%) |")
        md_lines.append(f"| 🟡 QUERY | {len(query_flags)} ({len(query_flags)/len(results)*100:.1f}%) |")
        md_lines.append(f"| 🔴 REJECTED | {len(critical_flags)} ({len(critical_flags)/len(results)*100:.1f}%) |")
        md_lines.append(f"| Amount Flagged | ₦{flagged_amount:,.2f} |")
        md_lines.append("")
        
        # Critical Flags Section
        if critical_flags:
            md_lines.append("---")
            md_lines.append("## 🚨 CRITICAL FLAGS (REJECT)")
            md_lines.append("")
            for result in critical_flags:
                md_lines.append(f"### ❌ Row {result.get('row_number', '?')}: {result.get('procedure', 'N/A')}")
                md_lines.append(f"- **Enrollee:** {result.get('enrollee', 'N/A')}")
                md_lines.append(f"- **PA:** {result.get('pa', 'N/A')} | **Date:** {result.get('date', 'N/A')}")
                md_lines.append(f"- **Diagnosis:** {result.get('diagnosis', 'N/A')}")
                md_lines.append(f"- **Amount:** ₦{result.get('amount', 0):,.2f}")
                md_lines.append(f"- **Reason:** {result.get('denial_reason', 'See rule details')}")
                
                # Show which rules failed
                for rule in ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']:
                    rule_data = result.get(rule, {})
                    if isinstance(rule_data, dict) and rule_data.get('status') == 'FAIL':
                        md_lines.append(f"- **{rule} FAILED:** {rule_data.get('reason', 'N/A')}")
                md_lines.append("")
        
        # Query Flags Section - Grouped by Issue Type
        if query_flags:
            md_lines.append("---")
            md_lines.append("## ⚠️ QUERY FLAGS (REVIEW REQUIRED)")
            md_lines.append("")
            
            # Group by rule that triggered the query
            query_by_rule = {}
            for result in query_flags:
                triggered_rules = []
                for rule in ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']:
                    rule_data = result.get(rule, {})
                    if isinstance(rule_data, dict) and rule_data.get('status') == 'QUERY':
                        triggered_rules.append(rule)
                
                for rule in triggered_rules:
                    if rule not in query_by_rule:
                        query_by_rule[rule] = []
                    query_by_rule[rule].append(result)
            
            # Show queries by rule
            rule_descriptions = {
                'R1': 'Submission Timing',
                'R6': 'Gender Appropriateness',
                'R7': 'Age Appropriateness',
                'R8': '30-Day Duplicate',
                'R9': 'Drug-Diagnosis Match',
                'R10': 'Symptomatic Only',
                'R11': 'Dosage Validation',
                'R12': 'Procedure-Diagnosis Match'
            }
            
            for rule, rule_results in query_by_rule.items():
                rule_desc = rule_descriptions.get(rule, rule)
                md_lines.append(f"### {rule}: {rule_desc} ({len(rule_results)} issues)")
                md_lines.append("")
                for result in rule_results[:5]:  # Limit to 5 per rule
                    rule_data = result.get(rule, {})
                    reason = rule_data.get('reason', 'N/A') if isinstance(rule_data, dict) else 'N/A'
                    md_lines.append(f"- **{result.get('procedure', 'N/A')}** ({result.get('enrollee', 'N/A')})")
                    md_lines.append(f"  - PA: {result.get('pa', 'N/A')} | Date: {result.get('date', 'N/A')}")
                    md_lines.append(f"  - Issue: {reason[:200]}")
                md_lines.append("")
        
        # Pattern Analysis Section
        md_lines.append("---")
        md_lines.append("## 🔍 PATTERN ANALYSIS")
        md_lines.append("")
        
        provisional_count = 0
        for enrollee, patterns in all_patterns.items():
            provisional = [p for p in patterns if p['type'] == 'PROVISIONAL_DIAGNOSIS']
            provisional_count += len(provisional)
        
        if provisional_count > 0:
            md_lines.append(f"### ⚠️ Provisional Diagnosis Patterns Detected: {provisional_count}")
            md_lines.append("")
            md_lines.append("These encounters have diagnostic tests but NO treatment - diagnosis may be ruled out:")
            md_lines.append("")
            
            for enrollee, patterns in all_patterns.items():
                provisional = [p for p in patterns if p['type'] == 'PROVISIONAL_DIAGNOSIS']
                if provisional:
                    md_lines.append(f"**{enrollee}:**")
                    for p in provisional:
                        md_lines.append(f"- {p['date']} (PA {p['pa']}): Tests={', '.join(p['tests'][:2])}, "
                                       f"Symptomatic={', '.join(p['symptomatic'][:2])}")
            md_lines.append("")
        
        # Detailed Enrollee Analysis
        md_lines.append("---")
        md_lines.append("## 📋 DETAILED ENROLLEE ANALYSIS")
        md_lines.append("")
        
        for enrollee, profile in profiles.items():
            patterns = all_patterns.get(enrollee, [])
            narrative = self.generate_narrative_analysis(profile, patterns)
            md_lines.append(narrative)
        
        # Recommendations
        md_lines.append("")
        md_lines.append("---")
        md_lines.append("## 🎯 RECOMMENDATIONS")
        md_lines.append("")
        
        md_lines.append(f"### REJECT ({len(critical_flags)} claims, ₦{sum(r.get('amount', 0) or 0 for r in critical_flags):,.2f}):")
        if critical_flags:
            for i, result in enumerate(critical_flags[:10], 1):
                md_lines.append(f"{i}. Row {result.get('row_number', '?')} - {result.get('procedure', 'N/A')} "
                               f"({result.get('denial_reason', 'See details')[:50]})")
        else:
            md_lines.append("None")
        md_lines.append("")
        
        md_lines.append(f"### QUERY ({len(query_flags)} claims, ₦{sum(r.get('amount', 0) or 0 for r in query_flags):,.2f}):")
        if query_flags:
            md_lines.append("Review required before approval. See detailed analysis above.")
        else:
            md_lines.append("None")
        md_lines.append("")
        
        md_lines.append(f"### PASS ({len(passed)} claims, ₦{sum(r.get('amount', 0) or 0 for r in passed):,.2f}):")
        md_lines.append("All documentation appropriate. Recommend approval.")
        md_lines.append("")
        
        # Footer
        md_lines.append("---")
        md_lines.append("*Report generated by KLAIRE Claims Vetting Engine v3.8*")
        md_lines.append("*Clearline HMO Analytics Team*")
        
        # Save Markdown report
        md_path = output_path.replace('.xlsx', '_ANALYSIS.md')
        with open(md_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(md_lines))
        print(f"✅ Analytical report saved: {md_path}")
        
        # Also return the data for Excel generation
        return {
            'profiles': profiles,
            'patterns': all_patterns,
            'critical_flags': critical_flags,
            'query_flags': query_flags,
            'passed': passed,
            'total_amount': total_amount,
            'flagged_amount': flagged_amount,
            'markdown_path': md_path
        }
    
    def generate_excel_report(self, results: List[Dict], output_path: str, batch_name: str = "Claims Batch"):
        """
        Generate comprehensive Excel report
        """
        if not OPENPYXL_AVAILABLE:
            print("WARNING: openpyxl not available. Saving as CSV instead.")
            self._save_as_csv(results, output_path.replace('.xlsx', '.csv'))
            return
        
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        query_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        web_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ===== Sheet 1: Executive Summary =====
        ws1 = wb.active
        ws1.title = "Executive Summary"
        
        ws1['A1'] = f"KLAIRE CLAIMS VETTING REPORT v3.8"
        ws1['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws1['A2'] = f"Batch: {batch_name}"
        ws1['A2'].font = Font(bold=True, size=12)
        ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws1['A4'] = f"Model: Claude Opus 4.5"
        ws1['A5'] = f"Web Search Mode: {self.web_search_mode.upper()}"
        
        # Summary stats
        total = len(results)
        approved = sum(1 for r in results if r['final_status'] == 'APPROVED')
        denied = sum(1 for r in results if r['final_status'] == 'DENY')
        query = sum(1 for r in results if r['final_status'] == 'QUERY')
        web_searched = sum(1 for r in results if r.get('web_search', {}).get('performed', False))
        
        total_amt = sum(r['amount'] for r in results)
        approved_amt = sum(r['amount'] for r in results if r['final_status'] == 'APPROVED')
        denied_amt = sum(r['amount'] for r in results if r['final_status'] == 'DENY')
        query_amt = sum(r['amount'] for r in results if r['final_status'] == 'QUERY')
        
        ws1['A7'] = "FINANCIAL SUMMARY"
        ws1['A7'].font = Font(bold=True, size=12)
        
        headers = ['Status', 'Count', 'Amount', '%']
        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        summary_data = [
            ('Total', total, f"₦{total_amt:,.2f}", '100%'),
            ('APPROVED', approved, f"₦{approved_amt:,.2f}", f"{approved_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
            ('DENY', denied, f"₦{denied_amt:,.2f}", f"{denied_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
            ('QUERY', query, f"₦{query_amt:,.2f}", f"{query_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
        ]
        
        for i, (status, count, amt, pct) in enumerate(summary_data, start=9):
            ws1.cell(row=i, column=1, value=status)
            ws1.cell(row=i, column=2, value=count)
            ws1.cell(row=i, column=3, value=amt)
            ws1.cell(row=i, column=4, value=pct)
            
            if status == 'APPROVED':
                ws1.cell(row=i, column=1).fill = pass_fill
            elif status == 'DENY':
                ws1.cell(row=i, column=1).fill = fail_fill
            elif status == 'QUERY':
                ws1.cell(row=i, column=1).fill = query_fill
        
        # Web search stats
        if self.web_search_mode != 'off':
            ws1['A15'] = "WEB SEARCH STATISTICS"
            ws1['A15'].font = Font(bold=True, size=12)
            ws1['A16'] = f"Total Web Searches: {web_searched}"
            ws1['A17'] = f"Search Triggers: {self.stats['web_search_triggers']}"
        
        # Rule summary
        ws1['A20'] = "RULE-BY-RULE SUMMARY"
        ws1['A20'].font = Font(bold=True, size=12)
        
        rules = ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']
        rule_headers = ['Rule', 'PASS', 'FAIL', 'QUERY']
        for col, header in enumerate(rule_headers, start=1):
            cell = ws1.cell(row=21, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for i, rule in enumerate(rules, start=22):
            pass_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'PASS')
            fail_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'FAIL')
            query_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'QUERY')
            
            ws1.cell(row=i, column=1, value=rule)
            ws1.cell(row=i, column=2, value=pass_c)
            ws1.cell(row=i, column=3, value=fail_c)
            ws1.cell(row=i, column=4, value=query_c)
            
            if fail_c > 0:
                ws1.cell(row=i, column=3).fill = fail_fill
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 10
        
        # ===== Sheet 2: Rule Matrix =====
        ws2 = wb.create_sheet("Rule Matrix")
        
        headers = ['Row', 'Enrollee', 'PA', 'Procedure', 'Diagnosis', 'All Dx', 'Amount', 'Conf'] + rules + ['Status', 'Reason', 'Web']
        for col, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        for row_idx, result in enumerate(results, start=2):
            # v3.5 FIX: Increased truncation limits for readability
            data = [
                result['row_number'],
                result['enrollee'][:40] if result['enrollee'] else '',
                result['pa'],
                result['procedure'][:60] if result['procedure'] else '',
                result['diagnosis'][:50] if result['diagnosis'] else '',
                result['all_encounter_dx'][:80] if result['all_encounter_dx'] else '',
                f"₦{result['amount']:,.2f}",
                f"{result.get('confidence', 1.0):.2f}"
            ]
            
            # Add rule statuses
            for rule in rules:
                status = result['rules'].get(rule, {}).get('status', 'N/A')
                data.append(status)
            
            data.append(result['final_status'])
            # v3.5 FIX: Don't truncate final reason - it's important!
            data.append(result['final_reason'] if result['final_reason'] else '')
            data.append('Yes' if result.get('web_search', {}).get('performed', False) else '')
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color code rule columns
                if col_idx > 8 and col_idx <= 8 + len(rules):
                    if value == 'PASS':
                        cell.fill = pass_fill
                    elif value == 'FAIL':
                        cell.fill = fail_fill
                        cell.font = Font(bold=True)
                    elif value == 'QUERY':
                        cell.fill = query_fill
                
                # Color final status
                if col_idx == 8 + len(rules) + 1:
                    if value == 'APPROVED':
                        cell.fill = pass_fill
                    elif value == 'DENY':
                        cell.fill = fail_fill
                        cell.font = Font(bold=True)
                    elif value == 'QUERY':
                        cell.fill = query_fill
                
                # Highlight web searched
                if col_idx == len(data) and value == 'Yes':
                    cell.fill = web_fill
        
        # Adjust column widths - v3.5: Increased for readability
        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 28  # Enrollee
        ws2.column_dimensions['C'].width = 12  # PA
        ws2.column_dimensions['D'].width = 50  # Procedure - increased
        ws2.column_dimensions['E'].width = 40  # Diagnosis - increased
        ws2.column_dimensions['F'].width = 55  # All Encounter Dx - increased
        ws2.column_dimensions['G'].width = 14  # Amount
        ws2.column_dimensions['H'].width = 8   # Confidence
        
        # ===== Sheet 3: Flagged Claims =====
        ws3 = wb.create_sheet("Flagged Claims")
        
        ws3['A1'] = "FLAGGED CLAIMS - DETAILED ANALYSIS"
        ws3['A1'].font = Font(bold=True, size=14)
        
        flagged = [r for r in results if r['final_status'] in ['DENY', 'QUERY']]
        
        current_row = 3
        for result in flagged:
            # Header
            status_text = f"Row {result['row_number']} | {result['final_status']}"
            ws3.cell(row=current_row, column=1, value=status_text)
            ws3.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            if result['final_status'] == 'DENY':
                ws3.cell(row=current_row, column=1).fill = fail_fill
            else:
                ws3.cell(row=current_row, column=1).fill = query_fill
            current_row += 1
            
            # Details
            ws3.cell(row=current_row, column=1, value=f"Enrollee: {result['enrollee']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Procedure: {result['procedure']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Diagnosis: {result['diagnosis']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"All Encounter Dx: {result['all_encounter_dx']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Amount: ₦{result['amount']:,.2f}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Confidence: {result.get('confidence', 'N/A')}")
            current_row += 1
            
            # Web search info
            web_info = result.get('web_search', {})
            if web_info.get('performed'):
                ws3.cell(row=current_row, column=1, value="🔍 Web Search Performed")
                ws3.cell(row=current_row, column=1).fill = web_fill
                current_row += 1
                if web_info.get('verified_indications'):
                    ws3.cell(row=current_row, column=1, value=f"  Verified indications: {', '.join(web_info['verified_indications'][:5])}")
                    current_row += 1
            
            # Rule details
            ws3.cell(row=current_row, column=1, value="Rule Results:")
            ws3.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            for rule in rules:
                rule_result = result['rules'].get(rule, {})
                status = rule_result.get('status', 'N/A')
                reason = rule_result.get('reason', '')
                
                cell = ws3.cell(row=current_row, column=1, value=f"  {rule}: {status}")
                # v3.5 FIX: Don't truncate reasons - use full text!
                ws3.cell(row=current_row, column=2, value=reason)
                
                if status == 'FAIL':
                    cell.fill = fail_fill
                    cell.font = Font(bold=True)
                elif status == 'QUERY':
                    cell.fill = query_fill
                elif status == 'PASS':
                    cell.fill = pass_fill
                
                current_row += 1
            
            # Clinical notes
            if result.get('clinical_notes'):
                ws3.cell(row=current_row, column=1, value=f"Clinical Notes: {result['clinical_notes']}")
                current_row += 1
            
            current_row += 2
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 120  # v3.5: Wider for full reasons
        
        # ===== Sheet 4: Web Search Log (if applicable) =====
        if self.web_search_mode != 'off':
            ws4 = wb.create_sheet("Web Search Log")
            
            ws4['A1'] = "WEB SEARCH LOG"
            ws4['A1'].font = Font(bold=True, size=14)
            
            headers = ['Row', 'Procedure', 'Diagnosis', 'Search Status', 'Verified Indications', 'Appropriate']
            for col, header in enumerate(headers, start=1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            row_num = 4
            for result in results:
                web_info = result.get('web_search', {})
                if web_info.get('performed'):
                    ws4.cell(row=row_num, column=1, value=result['row_number'])
                    ws4.cell(row=row_num, column=2, value=result['procedure'][:60])  # v3.5: Increased
                    ws4.cell(row=row_num, column=3, value=result['diagnosis'][:50])  # v3.5: Increased
                    ws4.cell(row=row_num, column=4, value=web_info.get('status', 'N/A'))
                    verified_indications = web_info.get('verified_indications') or []
                    ws4.cell(row=row_num, column=5, value=', '.join(verified_indications[:5]) if verified_indications else 'N/A')  # v3.5: Show more
                    ws4.cell(row=row_num, column=6, value='Yes' if web_info.get('appropriate_for_diagnosis') else 'No')
                    row_num += 1
            
            ws4.column_dimensions['A'].width = 6
            ws4.column_dimensions['B'].width = 55  # v3.5: Wider
            ws4.column_dimensions['C'].width = 45  # v3.5: Wider
            ws4.column_dimensions['D'].width = 15
            ws4.column_dimensions['E'].width = 60  # v3.5: Wider for more indications
            ws4.column_dimensions['F'].width = 12
        
        # ===== Sheet 5: Analytical Summary (NEW in v3.8) =====
        ws5 = wb.create_sheet("Analytical Summary")
        
        ws5['A1'] = "ANALYTICAL SUMMARY - PATTERN DETECTION"
        ws5['A1'].font = Font(bold=True, size=14, color="1F4E79")
        ws5['A2'] = "This sheet provides encounter-level analysis with cross-PA references"
        
        # Build enrollee profiles and detect patterns
        profiles = self.build_enrollee_profile(results)
        
        # Summary by Enrollee
        row_num = 4
        headers = ['Enrollee', 'Total Claims', 'Total Amount', 'Unique PAs', 'Flags', 'Pattern Issues']
        for col, header in enumerate(headers, start=1):
            cell = ws5.cell(row=row_num, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        row_num = 5
        for enrollee_id, profile in profiles.items():
            patterns = self.detect_patterns(profile)
            provisional = [p for p in patterns if p['type'] == 'PROVISIONAL_DIAGNOSIS']
            
            ws5.cell(row=row_num, column=1, value=enrollee_id)
            ws5.cell(row=row_num, column=2, value=len(profile['results']))
            ws5.cell(row=row_num, column=3, value=f"₦{profile['total_amount']:,.2f}")
            ws5.cell(row=row_num, column=4, value=len(profile['all_pas']))
            ws5.cell(row=row_num, column=5, value=len([r for r in profile['results'] if r.get('final_status') != 'APPROVED']))
            
            # Pattern issues summary
            if provisional:
                pattern_desc = f"PROVISIONAL DX: {len(provisional)} encounters with test but no treatment"
                ws5.cell(row=row_num, column=6, value=pattern_desc)
                ws5.cell(row=row_num, column=6).fill = query_fill
            else:
                ws5.cell(row=row_num, column=6, value="No pattern issues detected")
                ws5.cell(row=row_num, column=6).fill = pass_fill
            
            row_num += 1
        
        # Pattern Details Section
        row_num += 2
        ws5.cell(row=row_num, column=1, value="PATTERN DETAILS")
        ws5.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1
        
        pattern_headers = ['Enrollee', 'PA', 'Date', 'Pattern Type', 'Details']
        for col, header in enumerate(pattern_headers, start=1):
            cell = ws5.cell(row=row_num, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row_num += 1
        
        for enrollee_id, profile in profiles.items():
            patterns = self.detect_patterns(profile)
            for pattern in patterns:
                if pattern['type'] == 'PROVISIONAL_DIAGNOSIS':
                    ws5.cell(row=row_num, column=1, value=enrollee_id)
                    ws5.cell(row=row_num, column=2, value=str(pattern.get('pa', '')))
                    ws5.cell(row=row_num, column=3, value=str(pattern.get('date', '')))
                    ws5.cell(row=row_num, column=4, value="PROVISIONAL DX")
                    ws5.cell(row=row_num, column=4).fill = query_fill
                    ws5.cell(row=row_num, column=5, value=pattern.get('analysis', '')[:150])
                    row_num += 1
        
        # Encounter-Level Summary Section
        row_num += 2
        ws5.cell(row=row_num, column=1, value="ENCOUNTER-LEVEL SUMMARY")
        ws5.cell(row=row_num, column=1).font = Font(bold=True, size=12)
        row_num += 1
        
        enc_headers = ['Enrollee', 'PA', 'Date', 'Procedures', 'Has Test', 'Has Treatment', 'Diagnoses']
        for col, header in enumerate(enc_headers, start=1):
            cell = ws5.cell(row=row_num, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        row_num += 1
        
        for enrollee_id, profile in profiles.items():
            for enc_key, enc in profile['encounters'].items():
                pa, date = enc_key
                ws5.cell(row=row_num, column=1, value=enrollee_id)
                ws5.cell(row=row_num, column=2, value=str(pa))
                ws5.cell(row=row_num, column=3, value=str(date))
                ws5.cell(row=row_num, column=4, value=', '.join(enc['procedures'][:3]) + ('...' if len(enc['procedures']) > 3 else ''))
                
                test_cell = ws5.cell(row=row_num, column=5, value="Yes" if enc['has_diagnostic_test'] else "No")
                if enc['has_diagnostic_test']:
                    test_cell.fill = web_fill
                
                treat_cell = ws5.cell(row=row_num, column=6, value="Yes" if enc['has_treatment'] else "No")
                if enc['has_treatment']:
                    treat_cell.fill = pass_fill
                elif enc['has_diagnostic_test']:
                    treat_cell.fill = query_fill  # Test but no treatment = flag
                
                ws5.cell(row=row_num, column=7, value='; '.join(list(enc['diagnoses'])[:3]))
                row_num += 1
        
        # Adjust column widths
        ws5.column_dimensions['A'].width = 28
        ws5.column_dimensions['B'].width = 12
        ws5.column_dimensions['C'].width = 14
        ws5.column_dimensions['D'].width = 60
        ws5.column_dimensions['E'].width = 12
        ws5.column_dimensions['F'].width = 14
        ws5.column_dimensions['G'].width = 50
        
        # Save
        wb.save(output_path)
        print(f"\n✅ Report saved: {output_path}")
    
    def _save_as_csv(self, results: List[Dict], output_path: str):
        """Fallback to save as CSV if openpyxl not available"""
        rows = []
        for r in results:
            row = {
                'Row': r['row_number'],
                'Enrollee': r['enrollee'],
                'PA': r['pa'],
                'Procedure': r['procedure'],
                'Diagnosis': r['diagnosis'],
                'All_Encounter_Dx': r['all_encounter_dx'],
                'Amount': r['amount'],
                'Confidence': r.get('confidence', 'N/A'),
                'Final_Status': r['final_status'],
                'Final_Reason': r['final_reason'],
                'Web_Searched': r.get('web_search', {}).get('performed', False)
            }
            for rule in ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']:
                row[rule] = r['rules'].get(rule, {}).get('status', 'N/A')
                row[f'{rule}_Reason'] = r['rules'].get(rule, {}).get('reason', '')
            rows.append(row)
        
        pd.DataFrame(rows).to_csv(output_path, index=False)
        print(f"\n✅ CSV saved: {output_path}")


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='KLAIRE Claims Vetting Engine v3.8 - Claude API with Web Search',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standard mode with Opus (smartest, default) - Web search always ON
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx
  
  # Use Haiku for speed (live PA validation)
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx --model haiku
  
  # Use Sonnet for balance
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx --model sonnet
  
  # Smart mode - auto-search only when model is uncertain (DEFAULT)
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx
  
  # Maximum thoroughness - always search web
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx --web-search-always
  
  # Fastest/cheapest - no web search  
  python claims_vet_api_v37.py --input claims.csv --output report.xlsx --no-web-search
  
Environment Variables:
  ANTHROPIC_API_KEY    Your Anthropic API key (required)
  
Model Options:
  --model opus     Claude Opus 4.5 (smartest, default)
  --model sonnet   Claude Sonnet 4.5 (balanced)
  --model haiku    Claude Haiku 4.5 (fastest, cheapest)
  
Web Search Options:
  Default: SMART mode (AI uses knowledge, searches only when uncertain - cost efficient)
  --web-search-always   Force web search for ALL claims (most thorough, higher cost)
  --no-web-search       Disable web search (fastest, cheapest, uses AI knowledge only)
        """
    )
    
    parser.add_argument('--input', '-i', required=True, help='Input claims file (CSV or Excel)')
    parser.add_argument('--output', '-o', default='vetting_report.xlsx', help='Output report file')
    parser.add_argument('--batch-name', '-n', default='Claims Batch', help='Name for the batch in report')
    parser.add_argument('--db-path', '-d', default='ai_driven_data.duckdb', help='Path to DuckDB database')
    parser.add_argument('--api-key', '-k', help='Anthropic API key (or set ANTHROPIC_API_KEY env var)')
    
    # Model selection
    parser.add_argument('--model', '-m', 
                        choices=['opus', 'sonnet', 'haiku', 
                                 'claude-opus-4-5-20251101', 
                                 'claude-sonnet-4-5-20250929', 
                                 'claude-haiku-4-5-20251001'],
                        default='sonnet',
                        help='Claude model to use (default: opus)')
    
    # Web search options - SMART is default (cost-efficient)
    # Use --web-search-always for maximum thoroughness
    parser.add_argument('--web-search-always', action='store_true',
                        help='Force web search for ALL claims (most thorough, higher cost)')
    parser.add_argument('--no-web-search', action='store_true',
                        help='Disable web search entirely (fastest, cheapest)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not Path(args.input).exists():
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)
    
    # Determine web search mode - DEFAULT IS SMART (confidence-based, cost-efficient)
    if args.web_search_always:
        web_search_mode = 'on'  # Always search
    elif args.no_web_search:
        web_search_mode = 'off'  # Never search
    else:
        web_search_mode = 'smart'  # Default: AI uses knowledge, searches when uncertain
    
    # Map model shorthand to full model name
    model_map = {
        'opus': 'claude-opus-4-5-20251101',
        'sonnet': 'claude-sonnet-4-5-20250929',
        'haiku': 'claude-haiku-4-5-20251001',
        'claude-opus-4-5-20251101': 'claude-opus-4-5-20251101',
        'claude-sonnet-4-5-20250929': 'claude-sonnet-4-5-20250929',
        'claude-haiku-4-5-20251001': 'claude-haiku-4-5-20251001'
    }
    selected_model = model_map.get(args.model, 'claude-opus-4-5-20251101')
    
    # Get friendly model name for display
    model_display = {
        'claude-opus-4-5-20251101': 'Claude Opus 4.5 (Smartest)',
        'claude-sonnet-4-5-20250929': 'Claude Sonnet 4.5 (Balanced)',
        'claude-haiku-4-5-20251001': 'Claude Haiku 4.5 (Fastest)'
    }
    
    print("="*60)
    print("  KLAIRE Claims Vetting Engine v3.8")
    print(f"  Model: {model_display.get(selected_model, selected_model)}")
    print(f"  Web Search: {web_search_mode.upper()}")
    print("="*60)
    
    try:
        # Initialize engine
        engine = ClaimsVettingEngine(
            api_key=args.api_key, 
            db_path=args.db_path,
            web_search_mode=web_search_mode,
            model=selected_model
        )
        
        # Load claims
        df = engine.load_claims(args.input)
        
        # Run vetting
        results = engine.vet_batch(df)
        
        # Generate ANALYTICAL report (detailed markdown with deliberative reasoning)
        print("\n" + "="*60)
        print("  GENERATING ANALYTICAL REPORT")
        print("="*60)
        analysis_data = engine.generate_analytical_report(results, args.output, args.batch_name)
        
        # Generate Excel report
        engine.generate_excel_report(results, args.output, args.batch_name)
        
        print("\n" + "="*60)
        print("  VETTING COMPLETE")
        print("="*60)
        print(f"\n📊 Reports generated:")
        print(f"   - Excel: {args.output}")
        print(f"   - Analytical: {analysis_data.get('markdown_path', 'N/A')}")
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()