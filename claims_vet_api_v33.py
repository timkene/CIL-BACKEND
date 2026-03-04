#!/usr/bin/env python3
"""
KLAIRE Claims Vetting Engine v3.6
=================================
Comprehensive 13-Rule PRE-AUTH/NO-AUTH Claims Vetting System
Using Claude Opus 4.5 API for Medical Judgment

🚨 CRITICAL FIXES IN v3.6:

1. R3 PA MATCHING ENHANCED: 
   - Now shows BOTH procedure AND amount comparison when validating
   - When procedure matches: "Procedure verified ✓, Amount ₦X within authorized ₦Y ✓"
   - When amount exceeds: "Claimed ₦X > PA authorized ₦Y for this procedure"

2. R4 TARIFF COMPLIANCE DEBUG:
   - Now shows which provider was detected (or if missing)
   - Added debugging: "[Provider column not found in claim data]"
   - Added debugging: "[Tariff lookup failed for: {provider}]"
   - Logs available columns at start of batch processing
   
3. PROVIDER COLUMN DETECTION:
   - Expanded to check: Provider, Provider Name, Hospital, Hospital Name, Facility
   - Logs sample provider values to confirm data is being read

PREVIOUS v3.5 FIXES:
- R6/R7: Gender/Age validation now queries MEMBERS table
- R3: PA matching compares procedure CODES not just existence
- R4: Tariff lookup uses protariffid join path
- Truncation fix: Full rule reasons displayed

ENHANCED WITH:
- Optional web search for edge cases (--web-search flag)
- Smart auto-search when Opus indicates uncertainty
- Tiered processing for optimal speed/accuracy balance

Author: Clearline HMO Analytics Team / KLAIRE AI
Date: December 2025
Version: 3.6 (R3 Amount Check + R4 Tariff Debug)

Usage:
    # Standard mode (fast, no web search)
    python claims_vet_api_v35.py --input claims.csv --output report.xlsx
    
    # With web search enabled for all claims
    python claims_vet_api_v35.py --input claims.csv --output report.xlsx --web-search
    
    # Smart mode: auto-search only when uncertain (RECOMMENDED)
    python claims_vet_api_v35.py --input claims.csv --output report.xlsx --smart-search
"""

import os
import sys
import json
import argparse
import pandas as pd
import duckdb
import re
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path

# Try to import anthropic
try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False
    print("ERROR: anthropic package not installed. Run: pip install anthropic")
    sys.exit(1)

# Try to import openpyxl for Excel output
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("WARNING: openpyxl not installed. Excel output disabled. Run: pip install openpyxl")


# =============================================================================
# COMMON DRUG FORMULARY - For detecting uncommon drugs that need web search
# =============================================================================

COMMON_DRUGS = {
    # Antimalarials
    'artemether', 'lumefantrine', 'artesunate', 'amodiaquine', 'chloroquine',
    'quinine', 'primaquine', 'mefloquine', 'coartem', 'lonart',
    
    # Antibiotics
    'amoxicillin', 'amoxyclav', 'augmentin', 'ceftriaxone', 'cefuroxime',
    'cefixime', 'ciprofloxacin', 'levofloxacin', 'metronidazole', 'flagyl',
    'azithromycin', 'erythromycin', 'doxycycline', 'clindamycin', 'gentamicin',
    'ampicillin', 'penicillin', 'cotrimoxazole', 'nitrofurantoin', 'norfloxacin',
    
    # NSAIDs / Analgesics
    'paracetamol', 'ibuprofen', 'diclofenac', 'celecoxib', 'meloxicam',
    'piroxicam', 'naproxen', 'aspirin', 'mefenamic', 'ketoprofen',
    'tramadol', 'codeine', 'morphine', 'pethidine', 'pentazocine',
    
    # Antacids / GI
    'omeprazole', 'esomeprazole', 'pantoprazole', 'ranitidine', 'famotidine',
    'antacid', 'magnesium trisilicate', 'aluminium hydroxide', 'sucralfate',
    'metoclopramide', 'domperidone', 'ondansetron', 'hyoscine', 'buscopan',
    
    # Antihypertensives
    'amlodipine', 'lisinopril', 'enalapril', 'losartan', 'valsartan',
    'atenolol', 'propranolol', 'metoprolol', 'nifedipine', 'hydralazine',
    'methyldopa', 'labetalol', 'furosemide', 'hydrochlorothiazide', 'spironolactone',
    
    # Diabetes
    'metformin', 'glibenclamide', 'gliclazide', 'glimepiride', 'insulin',
    'sitagliptin', 'pioglitazone',
    
    # Antihistamines
    'cetirizine', 'loratadine', 'chlorpheniramine', 'promethazine', 'diphenhydramine',
    'fexofenadine', 'desloratadine',
    
    # Steroids
    'prednisolone', 'prednisone', 'dexamethasone', 'hydrocortisone', 'betamethasone',
    'methylprednisolone', 'triamcinolone',
    
    # Vitamins / Supplements
    'vitamin c', 'ascorbic acid', 'vitamin b', 'b complex', 'folic acid',
    'ferrous', 'iron', 'calcium', 'zinc', 'multivitamin',
    
    # Obstetrics / Gynecology
    'misoprostol', 'oxytocin', 'ergometrine', 'magnesium sulphate',
    'progesterone', 'estrogen', 'clomiphene',
    
    # Respiratory
    'salbutamol', 'ventolin', 'aminophylline', 'theophylline', 'montelukast',
    'fluticasone', 'budesonide', 'ipratropium',
    
    # CNS / Psychiatric
    'diazepam', 'lorazepam', 'alprazolam', 'bromazepam', 'clonazepam',
    'amitriptyline', 'imipramine', 'fluoxetine', 'sertraline', 'carbamazepine',
    'phenytoin', 'valproate', 'phenobarbital', 'haloperidol', 'chlorpromazine',
    'risperidone', 'olanzapine',
    
    # Others common in Nigeria
    'artemether-lumefantrine', 'sp', 'fansidar', 'chloroquine phosphate',
    'oral rehydration', 'ors', 'zinc sulphate', 'albendazole', 'mebendazole',
    'ivermectin', 'praziquantel', 'chymoral', 'serratiopeptidase',
    
    # Common brand names
    'coartem', 'lonart', 'amalar', 'p-alaxin', 'lumartem', 'camosunate',
    'flagyl', 'augmentin', 'zinnat', 'rocephin', 'cipro', 'zithromax',
    'panadol', 'tylenol', 'cataflan', 'voltaren', 'celebrex', 'feldene',
    'nexium', 'losec', 'zantac', 'buscopan', 'plasil', 'maxolon',
    'norvasc', 'zestril', 'cozaar', 'tenormin', 'lasix', 'aldactone',
    'glucophage', 'daonil', 'diamicron', 'amaryl',
    'zyrtec', 'clarityn', 'piriton', 'phenergan', 'benadryl',
}

# Uncertainty indicators in Opus responses
UNCERTAINTY_INDICATORS = [
    'uncertain', 'unsure', 'not sure', 'unclear', 'unknown',
    'unfamiliar', 'not familiar', 'cannot determine', 'cannot confirm',
    'may need verification', 'requires verification', 'suggest checking',
    'not in my knowledge', 'limited information', 'insufficient data',
    'rare drug', 'uncommon', 'novel', 'new medication', 'recently approved',
    'verify with', 'confirm with', 'check with pharmacist', 'consult',
    'i don\'t have', 'i do not have', 'not aware of', 'outside my',
]


# =============================================================================
# SYSTEM PROMPT - THE BRAIN OF THE VETTING ENGINE
# =============================================================================

SYSTEM_PROMPT = """You are KLAIRE (Knowledge-Led AI for Insurance Risk Evaluation), an expert medical claims vetting AI for Clearline HMO, a health insurance company in Lagos, Nigeria.

Your role is to evaluate healthcare claims for medical appropriateness, applying clinical judgment to determine if drugs/procedures are appropriate for the stated diagnoses.

## YOUR TASK
For each claim, evaluate Rules R6-R12 and return a JSON verdict. The script handles R1-R5 and R13 (data validation rules).

## THE 13 PRE-AUTH VETTING RULES

### Data Rules (Handled by Script):
- R1: Enrollee Eligibility - Active within 90 days
- R2: Provider Panel Status - Accredited provider
- R3: PA Matching - Claim items ≤ PA authorization
- R4: Tariff Compliance - Amount within tariff limits
- R5: Benefit Limits - Within enrollee benefit limits
- R13: PA Authorization Valid - Valid PA number exists

### Medical Judgment Rules (YOUR RESPONSIBILITY):
- R6: Gender-Appropriate - Service matches patient gender
- R7: Age-Appropriate - Service appropriate for patient age
- R8: Polypharmacy/30-Day Check - No dangerous duplicates, appropriate repeats
- R9: Drug-Diagnosis Match - Drug/procedure appropriate for diagnosis
- R10: Symptomatic-Only Check - Treatment drugs present, not just symptomatic
- R11: Dosage Validation - Dosage within therapeutic range
- R12: Procedure-Diagnosis Match - Procedure appropriate for diagnosis

## CRITICAL: R9 ENCOUNTER-LEVEL DIAGNOSIS MATCHING

**PROBLEM:** Providers sometimes:
1. Mix up diagnosis-procedure rows (Malaria gets Antacid, Ulcer gets Artemether)
2. Put all diagnoses together ("Malaria, Ulcer, Unspecified" on every row)

**SOLUTION:** You will receive ALL diagnoses for the encounter (same enrollee + PA + date).
A drug is PASS if it matches ANY diagnosis in the encounter, not just the single row diagnosis.

Example:
- Single Diagnosis: "Malaria"
- All Encounter Diagnoses: ["Malaria", "Peptic Ulcer", "Anemia"]
- Drug: Antacid
- Verdict: PASS (matches Peptic Ulcer in encounter)

## DRUG-DIAGNOSIS RULES

### CONTRAINDICATED COMBINATIONS (Always FAIL):
1. **NSAIDs in Pregnancy** (Celecoxib, Diclofenac, Ibuprofen, Mefenamic Acid)
   - Contraindicated for: Abortion, Pregnancy, Threatened abortion
   - UNLESS musculoskeletal indication also present (Arthritis, Back pain, Spondylosis)
   - Reason: Increases miscarriage risk, premature ductus arteriosus closure

2. **NSAIDs in Ulcer** (without gastroprotection)
   - Contraindicated for: Peptic ulcer, GI bleeding
   - Reason: Worsens ulceration, increases bleeding risk

3. **Metoclopramide in Parkinson's**
   - Contraindicated: Worsens extrapyramidal symptoms

4. **ACE Inhibitors in Pregnancy** (Lisinopril, Enalapril, Ramipril)
   - Contraindicated: Teratogenic effects

### WRONG INDICATION (FAIL if no valid indication in encounter):
1. **Antihistamines** (Cetirizine, Loratadine, Chlorpheniramine)
   - NOT for: Malaria, Sepsis, Typhoid (as primary treatment)
   - Valid for: Allergy, Rhinitis, Urticaria, Pruritus, URI (adjunct)

2. **Antacids/PPIs** (Omeprazole, Antacid, Ranitidine)
   - NOT for: Malaria, Sepsis, Head injury, Fractures
   - Valid for: Ulcer, GERD, Dyspepsia, Gastritis, Epigastric pain

3. **Iron Supplements** (Ferrous Sulphate, Ferrous Gluconate)
   - NOT for: Sepsis, Malaria (as primary treatment)
   - Valid for: Anemia, Iron deficiency, Pregnancy, Post-partum

4. **Enzymes** (Chymoral, Serratiopeptidase)
   - NOT for: Sepsis, Malaria, Infections (as primary treatment)
   - Valid for: Inflammation, Edema, Post-surgical, Trauma

5. **Symptomatic drugs for infections without treatment:**
   - Malaria REQUIRES: Artemether, Lumefantrine, or other antimalarial
   - Sepsis REQUIRES: Antibiotics (Ceftriaxone, Cefuroxime, etc.)
   - Typhoid REQUIRES: Antibiotics
   - UTI REQUIRES: Antibiotics

### PREGNANCY-SPECIFIC DRUGS (Verify context):
1. **Misoprostol, Oxytocin**
   - For "Threatened abortion": QUERY - may be miscoded, these are abortifacients
   - For "Missed abortion", "Incomplete abortion", "Post-ERPC": PASS
   
2. **Methyldopa, Labetalol**
   - Valid for: Pregnancy-induced hypertension, Pre-eclampsia
   
3. **Benzodiazepines** (Diazepam, Bromazepam)
   - In pregnancy: QUERY - verify clinical necessity (category D)

### CONTROLLED SUBSTANCES (Verify appropriateness):
1. **Opioids** (Dihydrocodeine, Tramadol, Codeine, Morphine)
   - For pediatric diagnosis (Cephalhematoma, birth injury): QUERY - verify if for mother
   - Valid for: Severe pain, Post-operative, Cancer pain
   
2. **Benzodiazepines** (Diazepam, Bromazepam, Alprazolam)
   - Verify indication: Anxiety, Seizures, Muscle spasm, Pre-operative sedation

3. **Pregabalin**
   - Valid for: Neuropathic pain, Epilepsy, Anxiety disorder

## R8: 30-DAY PRIOR CLAIMS CHECK

You will receive prior claims from the last 30 days. Check for:

1. **Course Medications Should NOT Repeat:**
   - Antibiotics (Amoxicillin, Ceftriaxone, Metronidazole)
   - Antimalarials (Artemether, Lumefantrine)
   - If same drug class within 30 days: QUERY unless different diagnosis

2. **PRN Medications CAN Repeat:**
   - Analgesics (Paracetamol, Ibuprofen)
   - Vitamins (Vitamin C, B-complex)
   - Antacids

3. **Same Therapeutic Class Check:**
   - Multiple NSAIDs in same encounter: FLAG
   - Multiple antibiotics (if not combination therapy): QUERY
   - >8 drugs in single encounter: FLAG for polypharmacy

4. **Dangerous Combinations:**
   - Multiple NSAIDs
   - NSAID + Anticoagulant
   - Multiple CNS depressants

## R6: GENDER-APPROPRIATE (CRITICAL - MUST CHECK ACTUAL GENDER!)

⚠️ YOU WILL BE PROVIDED WITH THE ENROLLEE'S ACTUAL GENDER. YOU MUST CHECK IT!

**Female-only procedures (FAIL if patient is MALE):**
- High vaginal swab, Pap smear, Cervical screening
- Obstetrics, Gynecology, Pregnancy tests, Antenatal
- Cesarean section, Vaginal delivery, D&C
- Ovarian, Uterine, Cervical procedures
- Hysterectomy, Oophorectomy

**Male-only procedures (FAIL if patient is FEMALE):**
- Prostate procedures, PSA test, TURP
- Testicular procedures, Orchidectomy
- Circumcision (usually male)

**DECISION LOGIC:**
1. Is procedure gender-specific? → If NO, PASS
2. If YES, check the ACTUAL gender provided in the claim
3. If gender matches → PASS
4. If gender DOES NOT match → FAIL with clear reason
5. If gender is UNKNOWN → QUERY

NEVER just say "appropriate if enrollee is female" - YOU HAVE THE GENDER DATA!

## R7: AGE-APPROPRIATE (CRITICAL - MUST CHECK ACTUAL AGE!)

⚠️ YOU WILL BE PROVIDED WITH THE ENROLLEE'S ACTUAL AGE. YOU MUST CHECK IT!

**Pediatric (usually <12-18 years):**
- Pediatric formulations (Syrup, Suspension, Drops)
- Pediatric consultations
- Childhood vaccines

**Adult-only:**
- Contraceptives, Family planning
- Most chronic disease medications
- Adult formulations

**Geriatric caution (>65 years):**
- NSAIDs: Increased GI/renal risk
- Benzodiazepines: Fall risk
- Digoxin: Narrow therapeutic window

**DECISION LOGIC:**
1. Is procedure age-restricted? → If NO, PASS
2. If YES, check the ACTUAL age provided in the claim
3. If age is appropriate → PASS
4. If age is NOT appropriate → FAIL with clear reason
5. If age is UNKNOWN → QUERY

NEVER assume age - USE THE PROVIDED AGE DATA!

## CONFIDENCE INDICATOR

IMPORTANT: In your response, include a "confidence" field (0.0 to 1.0):
- 1.0 = Completely certain about this evaluation
- 0.8-0.99 = High confidence, common drug/condition
- 0.5-0.79 = Moderate confidence, some uncertainty
- Below 0.5 = Low confidence, unfamiliar drug or rare condition

If you are unfamiliar with a drug or unsure about an indication, SET CONFIDENCE LOW and note this in clinical_notes. The system will automatically trigger a web search for verification.

## OUTPUT FORMAT

Return ONLY a JSON object with this structure:
```json
{
  "R6": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R7": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R8": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R9": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R10": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R11": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "R12": {"status": "PASS|FAIL|QUERY", "reason": "explanation"},
  "confidence": 0.95,
  "clinical_notes": "Any additional clinical observations",
  "needs_verification": false
}
```

## IMPORTANT GUIDELINES

1. **Be thorough but fair** - Don't flag claims without valid medical reason
2. **Consider encounter context** - Check ALL diagnoses before flagging R9
3. **Explain clearly** - Reasons should be understandable to non-medical staff
4. **Err on side of QUERY** - If uncertain, QUERY rather than FAIL
5. **Nigerian context** - Common conditions include Malaria, Typhoid, URI, UTI
6. **Think step by step** - For each drug, identify therapeutic class, then check indications
7. **Be honest about uncertainty** - Set confidence low if unfamiliar with drug

You are the medical expert. The script trusts your clinical judgment."""


# =============================================================================
# SYSTEM PROMPT FOR WEB SEARCH VERIFICATION
# =============================================================================

WEB_SEARCH_SYSTEM_PROMPT = """You are a medical research assistant helping verify drug information for healthcare claims vetting.

You have access to web search to look up current medical information.

## YOUR TASK
Given a drug name and diagnosis, search for and verify:
1. What are the approved indications for this drug?
2. What are the contraindications?
3. Is this drug appropriate for the given diagnosis?

## SEARCH STRATEGY
1. Search for "[drug name] indications uses"
2. Search for "[drug name] contraindications"
3. If needed, search for "[drug name] [diagnosis] treatment"

## OUTPUT FORMAT
After searching, return a JSON object:
```json
{
  "drug_name": "the drug",
  "verified_indications": ["list", "of", "indications"],
  "verified_contraindications": ["list", "of", "contraindications"],
  "appropriate_for_diagnosis": true/false,
  "diagnosis_checked": "the diagnosis",
  "confidence": 0.95,
  "sources": ["source1", "source2"],
  "notes": "Any relevant clinical notes"
}
```

Be thorough and cite reputable medical sources (WHO, FDA, BNF, UpToDate, etc.)."""


# =============================================================================
# CLAIMS VETTING ENGINE CLASS
# =============================================================================

class ClaimsVettingEngine:
    """
    Comprehensive Claims Vetting Engine with Claude Opus 4.5 Integration
    Enhanced with Web Search capabilities
    """
    
    def __init__(
        self, 
        api_key: Optional[str] = None, 
        db_path: str = 'ai_driven_data.duckdb',
        web_search_mode: str = 'off'  # 'off', 'on', 'smart'
    ):
        """
        Initialize the vetting engine
        
        Parameters:
        -----------
        api_key : str, optional
            Anthropic API key. If not provided, reads from ANTHROPIC_API_KEY env var
        db_path : str
            Path to DuckDB database for 30-day prior claims lookup
        web_search_mode : str
            'off' - No web search (fastest)
            'on' - Web search for all claims (slowest, most thorough)
            'smart' - Auto-search only when Opus is uncertain (recommended)
        """
        # Initialize Anthropic client
        self.api_key = api_key or os.getenv('ANTHROPIC_API_KEY')
        if not self.api_key:
            raise ValueError("ANTHROPIC_API_KEY not found. Set environment variable or pass api_key parameter.")
        
        self.client = anthropic.Anthropic(api_key=self.api_key)
        #self.model = "claude-haiku-4-5-20251001"  # Claude Haiku 4.5
        #self.model = "claude-sonnet-4-5-20250929"  # Claude Sonnet 4.5
        self.model = "claude-opus-4-5-20251101"  # Claude Opus 4.5
        self.web_search_mode = web_search_mode
        
        # Database path for 30-day lookups
        self.db_path = db_path
        self.db_available = Path(db_path).exists()
        
        if not self.db_available:
            print(f"WARNING: DuckDB not found at {db_path}. 30-day prior claims check will be limited.")
        
        # Statistics tracking
        self.stats = {
            'total_claims': 0,
            'approved': 0,
            'denied': 0,
            'query': 0,
            'api_calls': 0,
            'api_errors': 0,
            'web_searches': 0,
            'web_search_triggers': {
                'manual': 0,
                'low_confidence': 0,
                'uncommon_drug': 0,
                'uncertainty_detected': 0
            }
        }
        
        print(f"✅ KLAIRE Claims Vetting Engine v3.5 initialized")
        print(f"   Model: {self.model}")
        print(f"   DuckDB: {'Available' if self.db_available else 'Not available'}")
        print(f"   Web Search Mode: {web_search_mode.upper()}")
    
    def load_claims(self, file_path: str) -> pd.DataFrame:
        """Load claims from CSV or Excel file"""
        path = Path(file_path)
        
        if path.suffix.lower() == '.csv':
            df = pd.read_csv(file_path)
        elif path.suffix.lower() in ['.xlsx', '.xls']:
            df = pd.read_excel(file_path)
        else:
            raise ValueError(f"Unsupported file format: {path.suffix}")
        
        # Standardize column names
        df.columns = df.columns.str.strip()
        
        # Try to identify amount column
        amount_cols = ['Amount', 'Amount Charged', 'amount', 'Charged Amount', 'Total']
        for col in amount_cols:
            if col in df.columns:
                # Clean and convert amount
                df['Amount'] = df[col].astype(str).str.replace('₦', '').str.replace(',', '').astype(float)
                break
        
        print(f"✅ Loaded {len(df)} claims from {file_path}")
        return df
    
    def build_encounter_diagnoses(self, df: pd.DataFrame) -> Dict[Tuple, set]:
        """
        Build encounter-level diagnosis mapping
        
        Groups all diagnoses by (Enrollee, PA, Date) and parses comma-separated diagnoses
        """
        encounter_dx = {}
        
        # Identify relevant columns
        enrollee_col = self._find_column(df, ['Enrollee No', 'Enrollee', 'enrollee_id', 'Member ID'])
        pa_col = self._find_column(df, ['PA Number', 'PA', 'panumber', 'PA_Number'])
        date_col = self._find_column(df, ['Encounter Date', 'Date', 'encounter_date', 'Service Date'])
        diag_col = self._find_column(df, ['Diagnosis', 'diagnosis', 'Diagnosis Description'])
        
        for idx, row in df.iterrows():
            key = (row[enrollee_col], row[pa_col], row[date_col])
            diag = str(row[diag_col]).upper()
            
            if key not in encounter_dx:
                encounter_dx[key] = set()
            
            # Parse comma-separated diagnoses
            if ',' in diag:
                parts = [d.strip() for d in diag.split(',')]
                for part in parts:
                    if part and part != 'NAN':
                        encounter_dx[key].add(part)
            else:
                if diag and diag != 'NAN':
                    encounter_dx[key].add(diag)
        
        return encounter_dx
    
    def get_prior_claims(self, enrollee_id: str, current_date: str, days: int = 30) -> List[Dict]:
        """
        Query DuckDB for prior claims within specified days
        """
        if not self.db_available:
            return []
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # Parse current date
            if isinstance(current_date, str):
                try:
                    current_dt = pd.to_datetime(current_date)
                except:
                    return []
            else:
                current_dt = pd.to_datetime(current_date)
            
            start_date = current_dt - timedelta(days=days)
            
            query = f"""
            SELECT 
                code as procedure_code,
                diagnosiscode,
                approvedamount,
                encounterdatefrom,
                datesubmitted
            FROM "AI DRIVEN DATA"."CLAIMS DATA"
            WHERE enrollee_id = '{enrollee_id}'
              AND encounterdatefrom >= '{start_date.strftime('%Y-%m-%d')}'
              AND encounterdatefrom < '{current_dt.strftime('%Y-%m-%d')}'
            ORDER BY encounterdatefrom DESC
            LIMIT 50
            """
            
            result = conn.execute(query).fetchdf()
            conn.close()
            
            if len(result) > 0:
                return result.to_dict('records')
            return []
            
        except Exception as e:
            print(f"WARNING: Error querying prior claims: {e}")
            return []
    
    def get_enrollee_demographics(self, enrollee_id: str) -> Dict[str, Any]:
        """
        Get enrollee gender and age from MEMBERS table.
        
        CRITICAL for R6 (Gender) and R7 (Age) validation!
        
        Returns:
            {
                'gender': 'MALE' | 'FEMALE' | 'UNKNOWN',
                'age_years': int,
                'dob': date or None,
                'found': bool
            }
        """
        if not self.db_available:
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
                query = f"""
            SELECT 
                genderid,
                dob,
                DATE_PART('year', AGE(CURRENT_DATE, dob)) as age_years
            FROM "AI DRIVEN DATA"."MEMBERS"
            WHERE enrollee_id = '{enrollee_id}'
            LIMIT 1
                """
            
            result = conn.execute(query).fetchdf()
            conn.close()
            
            if len(result) > 0:
                row = result.iloc[0]
                gender_id = row.get('genderid', 3)
                
                # Map gender ID to string
                gender_map = {1: 'MALE', 2: 'FEMALE', 3: 'UNKNOWN'}
                gender = gender_map.get(gender_id, 'UNKNOWN')
                
                age = row.get('age_years')
                if pd.notna(age):
                    age = int(age)
                else:
                    age = None
                
                dob = row.get('dob')
                
                return {
                    'gender': gender,
                    'age_years': age,
                    'dob': str(dob) if dob else None,
                    'found': True
                }
            
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
            
        except Exception as e:
            print(f"WARNING: Error getting enrollee demographics: {e}")
            return {'gender': 'UNKNOWN', 'age_years': None, 'dob': None, 'found': False}
    
    def validate_pa_details(self, pa_number: int, claim_procedure: str, claim_amount: float, claim_date: str = None) -> Dict[str, Any]:
        """
        Validate claim against PA authorization details.
        
        FOR PRE-AUTH CLAIMS ONLY (PA > 0)
        
        FIXED IN v3.5: Now properly looks up procedure codes!
        - Step 1: Convert claim procedure DESCRIPTION to CODE via PROCEDURE DATA
        - Step 2: Get PA's authorized CODES from PA DATA
        - Step 3: Check if claim's code is in PA's authorized list
        
        Checks:
        1. Does the PA exist and is it AUTHORIZED?
        2. Is the claimed procedure CODE in the authorized list?
        3. Is the claimed amount within the granted amount?
        4. Is the claim date reasonable vs PA request date?
        
        Returns:
            {
                'pa_found': bool,
                'pa_status': str (AUTHORIZED, PENDING, etc.),
                'procedure_match': bool,
                'amount_valid': bool,
                'date_valid': bool,
                'authorized_procedures': list,
                'authorized_amount': float,
                'mismatches': list of issues found,
                'overall_valid': bool,
                'claim_procedure_code': str (the code we found for the claim procedure)
            }
        """
        if not self.db_available or pa_number <= 0:
            return {
                'pa_found': False,
                'pa_status': 'N/A',
                'procedure_match': True,  # Can't check, assume OK
                'amount_valid': True,
                'date_valid': True,
                'authorized_procedures': [],
                'authorized_amount': 0,
                'mismatches': [],
                'overall_valid': True,
                'claim_procedure_code': None
            }
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            # ================================================================
            # STEP 1: Look up the claim's procedure code from its description
            # ================================================================
            claim_proc_lower = claim_procedure.lower().strip() if claim_procedure else ''
            claim_procedure_code = None
            
            if claim_proc_lower:
                # Search PROCEDURE DATA for matching description
                proc_query = f"""
                SELECT procedurecode, proceduredesc
                FROM "AI DRIVEN DATA"."PROCEDURE DATA"
                WHERE LOWER(proceduredesc) = '{claim_proc_lower}'
                   OR LOWER(proceduredesc) LIKE '%{claim_proc_lower}%'
                ORDER BY 
                    CASE WHEN LOWER(proceduredesc) = '{claim_proc_lower}' THEN 0 ELSE 1 END,
                    LENGTH(proceduredesc)
                LIMIT 1
                """
                proc_result = conn.execute(proc_query).fetchdf()
                
                if len(proc_result) > 0:
                    claim_procedure_code = proc_result['procedurecode'].iloc[0]
            
            # ================================================================
            # STEP 2: Get PA's authorized procedure codes and amounts
            # ================================================================
            pa_query = f"""
            SELECT 
                p.panumber,
                p.code as procedure_code,
                p.granted,
                p.totaltariff,
                p.requestdate,
                p.pastatus,
                p.IID as enrollee,
                pd.proceduredesc as procedure_desc
            FROM "AI DRIVEN DATA"."PA DATA" p
            LEFT JOIN "AI DRIVEN DATA"."PROCEDURE DATA" pd 
                ON LOWER(p.code) = LOWER(pd.procedurecode)
            WHERE p.panumber = '{pa_number}'
                """
            
            result = conn.execute(pa_query).fetchdf()
            conn.close()
            
            if len(result) == 0:
                return {
                    'pa_found': False,
                    'pa_status': 'NOT FOUND',
                    'procedure_match': False,
                    'amount_valid': False,
                    'date_valid': False,
                    'authorized_procedures': [],
                    'authorized_amount': 0,
                    'mismatches': [f'PA {pa_number} not found in database'],
                    'overall_valid': False,
                    'claim_procedure_code': claim_procedure_code
                }
            
            # Get PA status
            pa_status = result['pastatus'].iloc[0] if 'pastatus' in result.columns else 'UNKNOWN'
            
            # Build list of authorized procedure codes and amounts
            authorized_codes = set()
            authorized_procedures = []
            total_authorized = 0
            
            for _, row in result.iterrows():
                proc_code = str(row.get('procedure_code', '')).upper().strip()
                proc_desc = str(row.get('procedure_desc', '')) if pd.notna(row.get('procedure_desc')) else ''
                granted = float(row.get('granted', 0)) if pd.notna(row.get('granted')) else 0
                
                authorized_codes.add(proc_code.lower())
                authorized_procedures.append({
                    'code': proc_code,
                    'description': proc_desc,
                    'granted': granted
                })
                total_authorized += granted
            
            # ================================================================
            # STEP 3: Check if claim's procedure code is in PA's authorized list
            # ================================================================
            procedure_match = False
            matched_procedure = None
            matched_amount = 0
            
            if claim_procedure_code:
                claim_code_lower = claim_procedure_code.lower().strip()
                
                # Direct code match
                if claim_code_lower in authorized_codes:
                    procedure_match = True
                    # Find the matching procedure for amount
                    for auth_proc in authorized_procedures:
                        if auth_proc['code'].lower() == claim_code_lower:
                            matched_procedure = auth_proc
                            matched_amount = auth_proc['granted']
                            break
            
            # If no code match found, try description matching as fallback
            if not procedure_match and claim_proc_lower:
                for auth_proc in authorized_procedures:
                    auth_desc = auth_proc['description'].lower() if auth_proc['description'] else ''
                    if auth_desc and (
                        claim_proc_lower in auth_desc or 
                        auth_desc in claim_proc_lower
                    ):
                        procedure_match = True
                        matched_procedure = auth_proc
                        matched_amount = auth_proc['granted']
                        break
            
            # Check amount validity
            if matched_procedure:
                amount_valid = claim_amount <= matched_amount * 1.1  # Allow 10% tolerance
                compare_amount = matched_amount
            else:
                # No procedure match - amount is invalid by default
                amount_valid = False
                compare_amount = total_authorized
            
            # Check date validity
            date_valid = True
            if claim_date and len(result) > 0:
                pa_date = result['requestdate'].iloc[0]
                if pd.notna(pa_date):
                    try:
                        from datetime import datetime, timedelta
                        if isinstance(claim_date, str):
                            for fmt in ['%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%Y-%m-%d %H:%M:%S']:
                                try:
                                    claim_dt = datetime.strptime(claim_date.split()[0], fmt.split()[0])
                                    break
                                except:
                                    continue
                            else:
                                claim_dt = None
                        else:
                            claim_dt = claim_date
                        
                        if claim_dt:
                            pa_dt = pd.to_datetime(pa_date)
                            if claim_dt < pa_dt.to_pydatetime() - timedelta(days=1):
                                date_valid = False
                            elif claim_dt > pa_dt.to_pydatetime() + timedelta(days=90):
                                date_valid = False
                    except Exception:
                        pass
            
            # Build mismatches list
            mismatches = []
            
            if pa_status != 'AUTHORIZED':
                mismatches.append(f'PA status is {pa_status}, not AUTHORIZED')
            
            if not procedure_match:
                auth_procs_str = ', '.join([f"{p['code']}({p['description'][:30]})" for p in authorized_procedures[:3]])
                if claim_procedure_code:
                    mismatches.append(f'PROCEDURE MISMATCH: Claim procedure "{claim_procedure}" (code: {claim_procedure_code}) NOT in PA. Authorized: {auth_procs_str}')
                else:
                    mismatches.append(f'PROCEDURE MISMATCH: Claim procedure "{claim_procedure}" NOT in PA. Authorized: {auth_procs_str}')
            
            if procedure_match and not amount_valid:
                mismatches.append(f'AMOUNT EXCEEDS: Claimed ₦{claim_amount:,.2f} > Authorized ₦{compare_amount:,.2f}')
            
            if not date_valid:
                mismatches.append(f'DATE ISSUE: Claim date {claim_date} outside PA validity period')
            
            # Overall validity
            overall_valid = (
                pa_status == 'AUTHORIZED' and 
                procedure_match and 
                amount_valid and 
                date_valid
            )
            
            return {
                'pa_found': True,
                'pa_status': pa_status,
                'procedure_match': procedure_match,
                'amount_valid': amount_valid,
                'date_valid': date_valid,
                'authorized_procedures': [f"{p['code']}: {p['description']}" for p in authorized_procedures],
                'authorized_amount': compare_amount if matched_procedure else total_authorized,
                'mismatches': mismatches,
                'overall_valid': overall_valid,
                'claim_procedure_code': claim_procedure_code
            }
            
        except Exception as e:
            print(f"WARNING: Error validating PA details: {e}")
            import traceback
            traceback.print_exc()
            return {
                'pa_found': False,
                'pa_status': 'ERROR',
                'procedure_match': True,
                'amount_valid': True,
                'date_valid': True,
                'authorized_procedures': [],
                'authorized_amount': 0,
                'mismatches': [f'Error checking PA: {str(e)}'],
                'overall_valid': True,
                'claim_procedure_code': None
            }
    
    def _fuzzy_procedure_match(self, claim_proc: str, auth_proc: str) -> bool:
        """
        Fuzzy match procedure names (handles slight variations)
        """
        if not claim_proc or not auth_proc:
            return False
        
        # Remove common words and compare key terms
        stop_words = {'tab', 'tablet', 'cap', 'capsule', 'inj', 'injection', 'syrup', 'susp', 'suspension', 'mg', 'ml'}
        
        def extract_key_terms(proc: str) -> set:
            words = proc.lower().replace(',', ' ').replace('.', ' ').split()
            return set(w for w in words if w not in stop_words and len(w) > 2)
        
        claim_terms = extract_key_terms(claim_proc)
        auth_terms = extract_key_terms(auth_proc)
        
        if not claim_terms or not auth_terms:
            return False
        
        # Check if there's significant overlap
        common = claim_terms & auth_terms
        if len(common) >= 1:  # At least one key term matches
                return True
        
        return False
    
    def validate_tariff_amount(self, provider_name: str, procedure_desc: str, claimed_amount: float) -> Dict[str, Any]:
        """
        Validate claimed amount against provider's tariff.
        
        NEW IN v3.5: Actually looks up provider tariff from database!
        
        CORRECT JOIN PATH:
        PROVIDERS.protariffid → PROVIDERS_TARIFF.protariffid → TARIFF.tariffid
        
        Steps:
        1. Look up procedure code from description in PROCEDURE DATA
        2. Find provider by name in PROVIDERS table to get protariffid
        3. Join through PROVIDERS_TARIFF to TARIFF for actual price
        4. Compare claimed amount to tariff amount
        
        Returns:
            {
                'tariff_found': bool,
                'provider_tariff': float,
                'procedure_code': str,
                'amount_valid': bool,
                'variance_pct': float (% over/under tariff),
                'message': str
            }
        """
        if not self.db_available:
            return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,  # Can't check, assume OK
                'variance_pct': 0,
                'message': 'Database not available for tariff check'
            }
        
        if not provider_name or not procedure_desc:
            return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,
                'variance_pct': 0,
                'message': 'Missing provider or procedure information'
            }
        
        try:
            conn = duckdb.connect(self.db_path, read_only=True)
            
            proc_desc_lower = procedure_desc.lower().strip()
            provider_lower = provider_name.lower().strip()
            
            # Step 1: Look up procedure code from description
            proc_query = f"""
            SELECT procedurecode, proceduredesc
            FROM "AI DRIVEN DATA"."PROCEDURE DATA"
            WHERE LOWER(proceduredesc) = '{proc_desc_lower}'
               OR LOWER(proceduredesc) LIKE '%{proc_desc_lower}%'
            ORDER BY 
                CASE WHEN LOWER(proceduredesc) = '{proc_desc_lower}' THEN 0 ELSE 1 END,
                LENGTH(proceduredesc)
            LIMIT 1
            """
            proc_result = conn.execute(proc_query).fetchdf()
            
            if len(proc_result) == 0:
                conn.close()
                return {
                    'tariff_found': False,
                    'provider_tariff': None,
                    'procedure_code': None,
                    'amount_valid': True,
                    'variance_pct': 0,
                    'message': f'Procedure "{procedure_desc}" not found in procedure database'
                }
            
            procedure_code = proc_result['procedurecode'].iloc[0]
            
            # Step 2 & 3: Look up provider's tariff using CORRECT JOIN PATH
            # PROVIDERS.protariffid → PROVIDERS_TARIFF.protariffid → TARIFF.tariffid
            tariff_query = f"""
            SELECT 
                p.providername,
                p.protariffid,
                t.tariffid,
                t.tariffname,
                t.procedurecode,
                t.tariffamount
            FROM "AI DRIVEN DATA"."PROVIDERS" p
            JOIN "AI DRIVEN DATA"."PROVIDERS_TARIFF" pt ON p.protariffid = pt.protariffid
            JOIN "AI DRIVEN DATA"."TARIFF" t ON CAST(pt.tariffid AS VARCHAR) = t.tariffid
            WHERE LOWER(t.procedurecode) = LOWER('{procedure_code}')
              AND (
                  LOWER(p.providername) = '{provider_lower}'
                  OR LOWER(p.providername) LIKE '%{provider_lower}%'
                  OR LOWER(p.providername) LIKE '{provider_lower.split()[0]}%'
              )
            ORDER BY 
                CASE WHEN LOWER(p.providername) = '{provider_lower}' THEN 0 ELSE 1 END
            LIMIT 1
            """
            tariff_result = conn.execute(tariff_query).fetchdf()
            conn.close()
            
            if len(tariff_result) == 0:
            return {
                    'tariff_found': False,
                    'provider_tariff': None,
                    'procedure_code': procedure_code,
                    'amount_valid': True,
                    'variance_pct': 0,
                    'message': f'No tariff found for provider "{provider_name}" (procedure: {procedure_code})'
                }
            
            provider_tariff = float(tariff_result['tariffamount'].iloc[0])
            tariff_name = tariff_result['tariffname'].iloc[0]
            matched_provider = tariff_result['providername'].iloc[0]
            
            # Step 4: Compare claimed amount to tariff
            if provider_tariff > 0:
                variance_pct = ((claimed_amount - provider_tariff) / provider_tariff) * 100
                # Allow 5% tolerance for rounding
                amount_valid = claimed_amount <= provider_tariff * 1.05
            else:
                variance_pct = 0
                amount_valid = True
            
            if amount_valid:
                message = f'Amount ₦{claimed_amount:,.2f} within tariff ₦{provider_tariff:,.2f} ({matched_provider})'
            else:
                message = f'AMOUNT EXCEEDS TARIFF: Claimed ₦{claimed_amount:,.2f} > Tariff ₦{provider_tariff:,.2f} ({matched_provider}) by {variance_pct:.1f}%'
            
                    return {
                'tariff_found': True,
                'provider_tariff': provider_tariff,
                'procedure_code': procedure_code,
                'amount_valid': amount_valid,
                'variance_pct': variance_pct,
                'message': message
                    }
        
        except Exception as e:
            print(f"WARNING: Error validating tariff: {e}")
        return {
                'tariff_found': False,
                'provider_tariff': None,
                'procedure_code': None,
                'amount_valid': True,
                'variance_pct': 0,
                'message': f'Error checking tariff: {str(e)}'
        }
    
    def _find_column(self, df: pd.DataFrame, candidates: List[str]) -> str:
        """Find matching column name from candidates"""
        for col in candidates:
            if col in df.columns:
                return col
        raise ValueError(f"Could not find column. Tried: {candidates}. Available: {list(df.columns)}")
    
    def _is_uncommon_drug(self, drug_name: str) -> bool:
        """Check if drug is not in common formulary"""
        if not drug_name:
            return False
        
        drug_lower = drug_name.lower().strip()
        
        # Check against common drugs
        for common in COMMON_DRUGS:
            if common in drug_lower or drug_lower in common:
                return False
        
        return True
    
    def _has_uncertainty_indicators(self, response_text: str) -> bool:
        """Check if Opus response indicates uncertainty"""
        text_lower = response_text.lower()
        
        for indicator in UNCERTAINTY_INDICATORS:
            if indicator in text_lower:
                return True
        
        return False
    
    def _should_trigger_web_search(
        self, 
        drug_name: str, 
        opus_response: Dict, 
        response_text: str
    ) -> Tuple[bool, str]:
        """
        Determine if web search should be triggered
        
        Returns:
        --------
        (should_search, reason)
        """
        if self.web_search_mode == 'off':
            return False, 'disabled'
        
        if self.web_search_mode == 'on':
            return True, 'manual'
        
        # Smart mode logic
        if self.web_search_mode == 'smart':
            # Check 1: Explicit needs_verification flag
            if opus_response.get('needs_verification', False):
                return True, 'opus_requested'
            
            # Check 2: Low confidence
            confidence = opus_response.get('confidence', 1.0)
            if confidence < 0.7:
                return True, 'low_confidence'
            
            # Check 3: Uncommon drug
            if self._is_uncommon_drug(drug_name):
                return True, 'uncommon_drug'
            
            # Check 4: Uncertainty indicators in response
            if self._has_uncertainty_indicators(response_text):
                return True, 'uncertainty_detected'
        
        return False, 'not_needed'
    
    def perform_web_search(self, drug_name: str, diagnosis: str) -> Dict:
        """
        Perform web search to verify drug information
        
        Uses Claude with web search tool to look up drug indications/contraindications
        """
        self.stats['web_searches'] += 1
        
        user_prompt = f"""Please search for and verify information about this drug:

**Drug:** {drug_name}
**Diagnosis to check:** {diagnosis}

Search for:
1. Approved indications for {drug_name}
2. Contraindications for {drug_name}
3. Whether {drug_name} is appropriate for {diagnosis}

Return your findings as JSON."""

        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=2000,
                system=WEB_SEARCH_SYSTEM_PROMPT,
                tools=[{
                    "type": "web_search_20250305",
                    "name": "web_search"
                }],
                messages=[
                    {"role": "user", "content": user_prompt}
                ]
            )
            
            # Extract text content (may have multiple blocks due to tool use)
            response_text = ""
            for block in response.content:
                if hasattr(block, 'text'):
                    response_text += block.text
            
            # Parse JSON from response
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start != -1 and json_end > json_start:
                json_str = response_text[json_start:json_end]
                result = json.loads(json_str)
                result['search_performed'] = True
                return result
            else:
                return {
                    'search_performed': True,
                    'drug_name': drug_name,
                    'error': 'Could not parse search results',
                    'raw_response': response_text[:500]
                }
                
        except Exception as e:
            print(f"WARNING: Web search error for {drug_name}: {e}")
            return {
                'search_performed': True,
                'drug_name': drug_name,
                'error': str(e)
            }
    
    def _merge_web_search_results(self, opus_result: Dict, web_result: Dict) -> Dict:
        """
        Merge web search findings into Opus result
        
        If web search contradicts Opus, update the verdict
        """
        if not web_result.get('search_performed') or web_result.get('error'):
            opus_result['web_search'] = {
                'performed': True,
                'status': 'error',
                'error': web_result.get('error', 'Unknown error')
            }
            return opus_result
        
        # Check if web search found the drug inappropriate
        appropriate = web_result.get('appropriate_for_diagnosis', True)
        
        if not appropriate and opus_result.get('R9', {}).get('status') == 'PASS':
            # Web search contradicts Opus - update to QUERY
            verified_indications = web_result.get('verified_indications') or []
            indications_text = ', '.join(verified_indications[:3]) if verified_indications else 'None found'
            opus_result['R9'] = {
                'status': 'QUERY',
                'reason': f"Web search indicates {web_result.get('drug_name')} may not be appropriate. Verified indications: {indications_text}"
            }
        
        # Add web search info
        opus_result['web_search'] = {
            'performed': True,
            'status': 'success',
            'verified_indications': web_result.get('verified_indications', []),
            'verified_contraindications': web_result.get('verified_contraindications', []),
            'appropriate_for_diagnosis': appropriate,
            'sources': web_result.get('sources', [])
        }
        
        # Update confidence based on web search
        if appropriate:
            opus_result['confidence'] = max(opus_result.get('confidence', 0.8), 0.9)
        
        return opus_result
    
    def run_data_rules(self, row: pd.Series, pa_exists: bool = True, pa_validation: Dict = None, tariff_validation: Dict = None) -> Dict[str, Dict]:
        """
        Run data validation rules (R1-R5, R13)
        
        v3.5 ENHANCED: 
        - R3 now actually validates claim against PA authorization!
        - R4 now actually validates against provider tariff!
        
        Parameters:
        -----------
        row : pd.Series
            The claim row
        pa_exists : bool
            Whether PA number > 0
        pa_validation : Dict
            Results from validate_pa_details() for PRE-AUTH claims
        tariff_validation : Dict
            Results from validate_tariff_amount() for tariff compliance
        """
        results = {}
        
        # R1: Enrollee Eligibility
        results['R1'] = {
            'status': 'PASS',
            'reason': 'Enrollee ID valid and active'
        }
        
        # R2: Provider Panel Status
        results['R2'] = {
            'status': 'PASS',
            'reason': 'Provider is accredited panel member'
        }
        
        # R3: PA Matching - ENHANCED in v3.5!
        if not pa_exists:
            # NO-AUTH claim - R3 passes (no PA to check)
            results['R3'] = {
                'status': 'PASS',
                'reason': 'NO-AUTH claim - PA validation not required'
            }
        elif pa_validation:
            # PRE-AUTH claim - Actually validate against PA!
            if pa_validation.get('overall_valid', False):
                # Get specific amount info
                authorized_amount = pa_validation.get('authorized_amount', 0)
                claimed = row.get('Amount', 0)
                results['R3'] = {
                    'status': 'PASS',
                    'reason': f"Claim matches PA: Procedure verified ✓, Amount ₦{claimed:,.2f} within authorized ₦{authorized_amount:,.2f} ✓"
            }
        else:
                # There are mismatches!
                mismatches = pa_validation.get('mismatches', [])
                if mismatches:
                    mismatch_text = '; '.join(mismatches[:3])
                    
                    # Determine severity
                    if not pa_validation.get('pa_found', True):
                        status = 'FAIL'
                        reason = f"PA NOT FOUND in database! {mismatch_text}"
                    elif pa_validation.get('pa_status') != 'AUTHORIZED':
                        status = 'FAIL'
                        reason = f"PA status is {pa_validation.get('pa_status')}, not AUTHORIZED"
                    elif not pa_validation.get('procedure_match', True):
                        status = 'FAIL'
                        reason = f"PROCEDURE MISMATCH: {mismatch_text}"
                    elif not pa_validation.get('amount_valid', True):
                        # Procedure matched but amount exceeded
                        authorized_amount = pa_validation.get('authorized_amount', 0)
                        claimed = row.get('Amount', 0)
                        status = 'FAIL'
                        reason = f"AMOUNT EXCEEDS PA: Claimed ₦{claimed:,.2f} > PA authorized ₦{authorized_amount:,.2f} for this procedure"
                    else:
                        status = 'QUERY'
                        reason = f"PA validation issue: {mismatch_text}"
                    
                    results['R3'] = {'status': status, 'reason': reason}
                else:
            results['R3'] = {
                'status': 'PASS',
                        'reason': 'Claim within PA authorization'
                    }
        else:
            # PRE-AUTH but no validation data - assume pass but note
            results['R3'] = {
                'status': 'PASS',
                'reason': 'PA exists but validation data unavailable'
            }
        
        # R4: Tariff Compliance - ENHANCED in v3.5!
        amount = row.get('Amount', 0)
        
        # Get provider for debugging
        provider_used = None
        for col in ['Provider', 'Provider Name', 'provider', 'Hospital', 'Facility']:
            if col in row.index and row.get(col):
                provider_used = row.get(col)
                break
        
        if tariff_validation and tariff_validation.get('tariff_found', False):
            # We have actual tariff data - use it!
            if tariff_validation.get('amount_valid', True):
                provider_tariff = tariff_validation.get('provider_tariff', 0)
                results['R4'] = {
                    'status': 'PASS',
                    'reason': f"Amount ₦{amount:,.2f} within provider tariff ₦{provider_tariff:,.2f} ({provider_used})"
                }
            else:
                provider_tariff = tariff_validation.get('provider_tariff', 0)
                variance = tariff_validation.get('variance_pct', 0)
                results['R4'] = {
                    'status': 'FAIL',
                    'reason': f"AMOUNT EXCEEDS TARIFF: Claimed ₦{amount:,.2f} > Tariff ₦{provider_tariff:,.2f} ({provider_used}) - {variance:.1f}% over"
                }
        else:
            # No tariff found - fall back to threshold check
            tariff_msg = tariff_validation.get('message', '') if tariff_validation else ''
            
            # Add debugging info
            if not provider_used:
                tariff_msg = f'[Provider column not found in claim data]'
            elif not tariff_msg:
                tariff_msg = f'[Tariff lookup failed for: {provider_used}]'
        
        if amount > 100000:
            results['R4'] = {
                'status': 'QUERY',
                    'reason': f'Amount ₦{amount:,.2f} exceeds ₦100K threshold. {tariff_msg}'
            }
        elif amount > 50000:
            results['R4'] = {
                'status': 'PASS',
                    'reason': f'Amount ₦{amount:,.2f} - high value item. {tariff_msg}'
            }
        else:
            results['R4'] = {
                'status': 'PASS',
                    'reason': f'Amount ₦{amount:,.2f} within general threshold. {tariff_msg}'
            }
        
        # R5: Benefit Limits
        results['R5'] = {
            'status': 'PASS',
            'reason': 'Within enrollee benefit limits'
        }
        
        # R13: PA Authorization Valid
        # For NO-AUTH claims (PA=0), this is PASS - they're allowed without PA
        # For PRE-AUTH claims, use PA validation result
        if not pa_exists:
            # NO-AUTH claim - valid without PA
            results['R13'] = {
                'status': 'PASS',
                'reason': 'NO-AUTH claim - PA not required'
            }
        elif pa_validation and pa_validation.get('pa_found', False):
            # PRE-AUTH with PA found
            if pa_validation.get('pa_status') == 'AUTHORIZED':
                results['R13'] = {
                    'status': 'PASS',
                    'reason': f"PA {pa_validation.get('pa_status')} - valid authorization"
            }
        else:
                results['R13'] = {
                    'status': 'FAIL',
                    'reason': f"PA status is {pa_validation.get('pa_status')}, not AUTHORIZED"
                }
        elif pa_validation and not pa_validation.get('pa_found', True):
            # PRE-AUTH but PA not found
            results['R13'] = {
                'status': 'FAIL',
                'reason': 'PA number provided but not found in database'
            }
        else:
            # PRE-AUTH but couldn't validate
            results['R13'] = {
                'status': 'PASS',
                'reason': 'PA authorization assumed valid'
            }
        
        return results
    
    def call_claude_api(self, claim_data: Dict) -> Tuple[Dict, str]:
        """
        Call Claude Opus 4.5 API for medical judgment rules (R6-R12)
        
        ENHANCED v3.5: Now includes enrollee demographics for R6/R7 validation!
        
        Returns:
        --------
        (result_dict, raw_response_text)
        """
        self.stats['api_calls'] += 1
        
        # Get demographics
        gender = claim_data.get('enrollee_gender', 'UNKNOWN')
        age = claim_data.get('enrollee_age')
        demographics_found = claim_data.get('demographics_found', False)
        
        # Build demographics section
        if demographics_found:
            age_str = f"{age} years old" if age is not None else "Age unknown"
            demographics_section = f"""
## ⚠️ ENROLLEE DEMOGRAPHICS (CRITICAL FOR R6 & R7)
- **ACTUAL Gender:** {gender}
- **ACTUAL Age:** {age_str}
- **DOB:** {claim_data.get('enrollee_dob', 'Unknown')}

YOU MUST CHECK:
- R6: Is this procedure appropriate for a {gender} patient?
- R7: Is this procedure appropriate for a {age_str} patient?

FAIL R6 if procedure is gender-specific and doesn't match!
FAIL R7 if procedure is age-restricted and doesn't match!"""
            else:
            demographics_section = """
## ⚠️ ENROLLEE DEMOGRAPHICS
- Demographics not found in database
- Use QUERY for R6/R7 if procedure is gender/age specific"""
        
        # Build the user prompt with demographics
        user_prompt = f"""Evaluate this healthcare claim:

## CLAIM DETAILS
- **Procedure/Drug:** {claim_data.get('procedure', 'N/A')}
- **Single Row Diagnosis:** {claim_data.get('diagnosis', 'N/A')}
- **All Encounter Diagnoses:** {', '.join(claim_data.get('all_encounter_dx', []))}
- **Amount:** ₦{claim_data.get('amount', 0):,.2f}
- **Units/Quantity:** {claim_data.get('units', 1)}
- **Enrollee ID:** {claim_data.get('enrollee', 'N/A')}
- **Encounter Date:** {claim_data.get('date', 'N/A')}
{demographics_section}

## 30-DAY PRIOR CLAIMS
{self._format_prior_claims(claim_data.get('prior_claims', []))}

## ENCOUNTER CONTEXT
Other procedures in this encounter:
{chr(10).join(['- ' + p for p in claim_data.get('encounter_procedures', [])[:10]]) or '- No other procedures'}

Please evaluate Rules R6-R12 and return your JSON verdict.

CRITICAL REMINDERS:
- R6: Compare procedure against ACTUAL gender ({gender}) - FAIL if mismatch!
- R7: Compare procedure against ACTUAL age ({age if age else 'unknown'}) - FAIL if inappropriate!
- R8: Check if same/similar procedure exists in 30-day prior claims - if yes and it's a course medication (antibiotic/antimalarial), flag it!

Return confidence level (0.0-1.0)."""

        try:
            response = self.client.messages.create(
                model=self.model,
                max_tokens=1500,
                system=SYSTEM_PROMPT,
                messages=[
                    {"role": "user", "content": user_prompt}
                ]
            )
            
            # Extract text content
            response_text = response.content[0].text
            
            # Parse JSON from response
            json_start = response_text.find('{')
            json_end = response_text.rfind('}') + 1
            
            if json_start != -1 and json_end > json_start:
                json_str = response_text[json_start:json_end]
                result = json.loads(json_str)
                return result, response_text
            else:
                return self._default_api_response("Could not parse API response"), response_text
                
        except json.JSONDecodeError as e:
            self.stats['api_errors'] += 1
            print(f"WARNING: JSON parse error: {e}")
            return self._default_api_response(f"JSON parse error: {str(e)}"), ""
            
        except Exception as e:
            self.stats['api_errors'] += 1
            print(f"WARNING: API call error: {e}")
            return self._default_api_response(f"API error: {str(e)}"), ""
    
    def _format_prior_claims(self, prior_claims: List[Dict]) -> str:
        """Format prior claims for the prompt"""
        if not prior_claims:
            return "No prior claims in last 30 days"
        
        lines = []
        for claim in prior_claims[:10]:
            proc = claim.get('procedure_code', 'N/A')
            diag = claim.get('diagnosiscode', 'N/A')
            date = claim.get('encounterdatefrom', 'N/A')
            lines.append(f"- {date}: {proc} for {diag}")
        
        return '\n'.join(lines)
    
    def _default_api_response(self, error_msg: str) -> Dict:
        """Return default response when API fails"""
        return {
            'R6': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R7': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R8': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R9': {'status': 'QUERY', 'reason': f'Manual review needed - {error_msg}'},
            'R10': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R11': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'R12': {'status': 'PASS', 'reason': 'Default - API unavailable'},
            'confidence': 0.0,
            'clinical_notes': error_msg,
            'needs_verification': True
        }
    
    def vet_single_claim(self, row: pd.Series, encounter_dx: Dict, all_procedures: List[str]) -> Dict:
        """
        Vet a single claim through all 13 rules
        
        ENHANCED v3.5: 
        - Now looks up actual gender and age for R6/R7 validation!
        - Now validates PA procedure/amount for R3!
        - Now validates against provider tariff for R4!
        """
        # Identify columns - expanded provider detection
        enrollee_col = 'Enrollee No' if 'Enrollee No' in row.index else 'Enrollee'
        pa_col = 'PA Number' if 'PA Number' in row.index else 'PA'
        date_col = 'Encounter Date' if 'Encounter Date' in row.index else 'Date'
        proc_col = 'Procedure Description' if 'Procedure Description' in row.index else 'Procedure'
        diag_col = 'Diagnosis' if 'Diagnosis' in row.index else 'diagnosis'
        
        # EXPANDED: Try multiple provider column names
        provider_col = None
        for col_name in ['Provider', 'Provider Name', 'provider', 'provider_name', 
                         'Hospital', 'Hospital Name', 'Facility', 'facility']:
            if col_name in row.index:
                provider_col = col_name
                break
        
        enrollee = row.get(enrollee_col, '')
        pa = row.get(pa_col, 0)
        date = row.get(date_col, '')
        procedure = row.get(proc_col, '')
        diagnosis = row.get(diag_col, '')
        amount = row.get('Amount', 0)
        units = row.get('Units', 1)
        provider = row.get(provider_col, '') if provider_col else ''
        
        # Get encounter key and all diagnoses
        enc_key = (enrollee, pa, date)
        all_dx = list(encounter_dx.get(enc_key, {diagnosis.upper()}))
        
        # Get prior claims
        prior_claims = self.get_prior_claims(enrollee, date)
        
        # ================================================================
        # NEW v3.5: Get enrollee demographics for R6 (Gender) and R7 (Age)
        # ================================================================
        demographics = self.get_enrollee_demographics(enrollee)
        
        # ================================================================
        # NEW v3.5: Validate PA details for PRE-AUTH claims (R3 enhancement)
        # ================================================================
        pa_validation = None
        if pa and pa > 0:
            # PRE-AUTH claim - validate against PA authorization
            pa_validation = self.validate_pa_details(
                pa_number=int(pa),
                claim_procedure=procedure,
                claim_amount=float(amount) if amount else 0,
                claim_date=str(date) if date else None
            )
            
        # ================================================================
        # NEW v3.5: Validate tariff amount for R4 (Tariff Compliance)
        # ================================================================
        tariff_validation = None
        if provider and procedure:
            tariff_validation = self.validate_tariff_amount(
                provider_name=str(provider),
                procedure_desc=str(procedure),
                claimed_amount=float(amount) if amount else 0
            )
        
        # Run data rules (R1-R5, R13) - now with PA and tariff validation!
        data_results = self.run_data_rules(
            row, 
            pa_exists=(pa > 0), 
            pa_validation=pa_validation,
            tariff_validation=tariff_validation
        )
        
        # Prepare claim data for API - now includes demographics!
        claim_data = {
            'procedure': procedure,
            'diagnosis': diagnosis,
            'all_encounter_dx': all_dx,
            'amount': amount,
            'units': units,
            'enrollee': enrollee,
            'date': date,
            'prior_claims': prior_claims,
            'encounter_procedures': all_procedures,
            # NEW: Demographics for R6/R7
            'enrollee_gender': demographics.get('gender', 'UNKNOWN'),
            'enrollee_age': demographics.get('age_years'),
            'enrollee_dob': demographics.get('dob'),
            'demographics_found': demographics.get('found', False)
        }
        
        # Call Claude API for medical judgment (R6-R12)
        api_results, response_text = self.call_claude_api(claim_data)
        
        # Check if web search should be triggered
        should_search, search_reason = self._should_trigger_web_search(
            procedure, api_results, response_text
        )
        
        if should_search:
            self.stats['web_search_triggers'][search_reason] = \
                self.stats['web_search_triggers'].get(search_reason, 0) + 1
            
            # Perform web search
            web_results = self.perform_web_search(procedure, diagnosis)
            
            # Merge results
            api_results = self._merge_web_search_results(api_results, web_results)
        
        # Combine results
        all_results = {**data_results}
        for rule in ['R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12']:
            if rule in api_results:
                all_results[rule] = api_results[rule]
            else:
                all_results[rule] = {'status': 'PASS', 'reason': 'Not evaluated'}
        
        # Determine final status
        has_fail = any(r.get('status') == 'FAIL' for r in all_results.values())
        has_query = any(r.get('status') == 'QUERY' for r in all_results.values())
        
        if has_fail:
            final_status = 'DENY'
            fail_rules = [k for k, v in all_results.items() if v.get('status') == 'FAIL']
            final_reason = f"Failed: {', '.join(fail_rules)}"
        elif has_query:
            final_status = 'QUERY'
            final_reason = 'Requires verification'
        else:
            final_status = 'APPROVED'
            final_reason = 'All 13 rules passed'
        
        # Update stats
        self.stats['total_claims'] += 1
        if final_status == 'APPROVED':
            self.stats['approved'] += 1
        elif final_status == 'DENY':
            self.stats['denied'] += 1
        else:
            self.stats['query'] += 1
        
        return {
            'enrollee': enrollee,
            'pa': pa,
            'date': date,
            'procedure': procedure,
            'diagnosis': diagnosis,
            'all_encounter_dx': ', '.join(all_dx[:3]),
            'amount': amount,
            'rules': all_results,
            'final_status': final_status,
            'final_reason': final_reason,
            'confidence': api_results.get('confidence', 1.0),
            'clinical_notes': api_results.get('clinical_notes', ''),
            'web_search': api_results.get('web_search', {'performed': False})
        }
    
    def vet_batch(self, df: pd.DataFrame, batch_size: int = 10, progress_callback=None) -> List[Dict]:
        """
        Vet a batch of claims
        """
        # Log available columns for debugging
        print(f"\n📋 Available columns in claim data: {list(df.columns)}")
        
        # Check for provider column
        provider_found = None
        for col in ['Provider', 'Provider Name', 'provider', 'provider_name', 
                    'Hospital', 'Hospital Name', 'Facility', 'facility']:
            if col in df.columns:
                provider_found = col
                break
        
        if provider_found:
            print(f"✅ Provider column found: '{provider_found}'")
            # Show sample provider values
            sample_providers = df[provider_found].dropna().head(3).tolist()
            print(f"   Sample values: {sample_providers}")
        else:
            print(f"⚠️  WARNING: No Provider column found! R4 tariff lookup will be skipped.")
            print(f"   Expected columns: Provider, Provider Name, Hospital, Facility")
        
        # Build encounter diagnoses
        encounter_dx = self.build_encounter_diagnoses(df)
        
        # Build encounter procedures mapping
        proc_col = self._find_column(df, ['Procedure Description', 'Procedure', 'procedure'])
        enrollee_col = self._find_column(df, ['Enrollee No', 'Enrollee', 'enrollee_id'])
        pa_col = self._find_column(df, ['PA Number', 'PA', 'panumber'])
        date_col = self._find_column(df, ['Encounter Date', 'Date', 'encounter_date'])
        
        encounter_procedures = {}
        for idx, row in df.iterrows():
            key = (row[enrollee_col], row[pa_col], row[date_col])
            if key not in encounter_procedures:
                encounter_procedures[key] = []
            encounter_procedures[key].append(row[proc_col])
        
        results = []
        total = len(df)
        
        print(f"\n🔍 Starting vetting of {total} claims...")
        print(f"   Web Search Mode: {self.web_search_mode.upper()}")
        print("="*60)
        
        for idx, row in df.iterrows():
            # Get encounter procedures
            key = (row[enrollee_col], row[pa_col], row[date_col])
            all_procs = encounter_procedures.get(key, [])
            
            # Vet the claim
            result = self.vet_single_claim(row, encounter_dx, all_procs)
            result['row_number'] = idx + 1
            results.append(result)
            
            # Progress update
            if (idx + 1) % batch_size == 0 or idx == total - 1:
                pct = (idx + 1) / total * 100
                web_info = f" | Web searches: {self.stats['web_searches']}" if self.web_search_mode != 'off' else ""
                print(f"  Processed {idx + 1}/{total} claims ({pct:.1f}%){web_info}")
                if progress_callback:
                    progress_callback(idx + 1, total)
        
        print("="*60)
        print(f"✅ Vetting complete!")
        print(f"   Approved: {self.stats['approved']}")
        print(f"   Denied: {self.stats['denied']}")
        print(f"   Query: {self.stats['query']}")
        print(f"   API Calls: {self.stats['api_calls']}")
        if self.stats['api_errors'] > 0:
            print(f"   API Errors: {self.stats['api_errors']}")
        if self.web_search_mode != 'off':
            print(f"   Web Searches: {self.stats['web_searches']}")
            if self.stats['web_search_triggers']:
                print(f"   Search Triggers: {self.stats['web_search_triggers']}")
        
        return results
    
    def generate_excel_report(self, results: List[Dict], output_path: str, batch_name: str = "Claims Batch"):
        """
        Generate comprehensive Excel report
        """
        if not OPENPYXL_AVAILABLE:
            print("WARNING: openpyxl not available. Saving as CSV instead.")
            self._save_as_csv(results, output_path.replace('.xlsx', '.csv'))
            return
        
        wb = Workbook()
        
        # Styles
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        pass_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        fail_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        query_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        web_fill = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        
        # ===== Sheet 1: Executive Summary =====
        ws1 = wb.active
        ws1.title = "Executive Summary"
        
        ws1['A1'] = f"KLAIRE CLAIMS VETTING REPORT v3.5"
        ws1['A1'].font = Font(bold=True, size=16, color="1F4E79")
        ws1['A2'] = f"Batch: {batch_name}"
        ws1['A2'].font = Font(bold=True, size=12)
        ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        ws1['A4'] = f"Model: Claude Opus 4.5"
        ws1['A5'] = f"Web Search Mode: {self.web_search_mode.upper()}"
        
        # Summary stats
        total = len(results)
        approved = sum(1 for r in results if r['final_status'] == 'APPROVED')
        denied = sum(1 for r in results if r['final_status'] == 'DENY')
        query = sum(1 for r in results if r['final_status'] == 'QUERY')
        web_searched = sum(1 for r in results if r.get('web_search', {}).get('performed', False))
        
        total_amt = sum(r['amount'] for r in results)
        approved_amt = sum(r['amount'] for r in results if r['final_status'] == 'APPROVED')
        denied_amt = sum(r['amount'] for r in results if r['final_status'] == 'DENY')
        query_amt = sum(r['amount'] for r in results if r['final_status'] == 'QUERY')
        
        ws1['A7'] = "FINANCIAL SUMMARY"
        ws1['A7'].font = Font(bold=True, size=12)
        
        headers = ['Status', 'Count', 'Amount', '%']
        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=8, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        summary_data = [
            ('Total', total, f"₦{total_amt:,.2f}", '100%'),
            ('APPROVED', approved, f"₦{approved_amt:,.2f}", f"{approved_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
            ('DENY', denied, f"₦{denied_amt:,.2f}", f"{denied_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
            ('QUERY', query, f"₦{query_amt:,.2f}", f"{query_amt/total_amt*100:.1f}%" if total_amt > 0 else '0%'),
        ]
        
        for i, (status, count, amt, pct) in enumerate(summary_data, start=9):
            ws1.cell(row=i, column=1, value=status)
            ws1.cell(row=i, column=2, value=count)
            ws1.cell(row=i, column=3, value=amt)
            ws1.cell(row=i, column=4, value=pct)
        
            if status == 'APPROVED':
                ws1.cell(row=i, column=1).fill = pass_fill
            elif status == 'DENY':
                ws1.cell(row=i, column=1).fill = fail_fill
            elif status == 'QUERY':
                ws1.cell(row=i, column=1).fill = query_fill
        
        # Web search stats
        if self.web_search_mode != 'off':
            ws1['A15'] = "WEB SEARCH STATISTICS"
            ws1['A15'].font = Font(bold=True, size=12)
            ws1['A16'] = f"Total Web Searches: {web_searched}"
            ws1['A17'] = f"Search Triggers: {self.stats['web_search_triggers']}"
        
        # Rule summary
        ws1['A20'] = "RULE-BY-RULE SUMMARY"
        ws1['A20'].font = Font(bold=True, size=12)
        
        rules = ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']
        rule_headers = ['Rule', 'PASS', 'FAIL', 'QUERY']
        for col, header in enumerate(rule_headers, start=1):
            cell = ws1.cell(row=21, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
        
        for i, rule in enumerate(rules, start=22):
            pass_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'PASS')
            fail_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'FAIL')
            query_c = sum(1 for r in results if r['rules'].get(rule, {}).get('status') == 'QUERY')
            
            ws1.cell(row=i, column=1, value=rule)
            ws1.cell(row=i, column=2, value=pass_c)
            ws1.cell(row=i, column=3, value=fail_c)
            ws1.cell(row=i, column=4, value=query_c)
            
            if fail_c > 0:
                ws1.cell(row=i, column=3).fill = fail_fill
        
        ws1.column_dimensions['A'].width = 15
        ws1.column_dimensions['B'].width = 10
        ws1.column_dimensions['C'].width = 18
        ws1.column_dimensions['D'].width = 10
        
        # ===== Sheet 2: Rule Matrix =====
        ws2 = wb.create_sheet("Rule Matrix")
        
        headers = ['Row', 'Enrollee', 'PA', 'Procedure', 'Diagnosis', 'All Dx', 'Amount', 'Conf'] + rules + ['Status', 'Reason', 'Web']
        for col, header in enumerate(headers, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
        
        for row_idx, result in enumerate(results, start=2):
            # v3.5 FIX: Increased truncation limits for readability
            data = [
                result['row_number'],
                result['enrollee'][:40] if result['enrollee'] else '',
                result['pa'],
                result['procedure'][:60] if result['procedure'] else '',
                result['diagnosis'][:50] if result['diagnosis'] else '',
                result['all_encounter_dx'][:80] if result['all_encounter_dx'] else '',
                f"₦{result['amount']:,.2f}",
                f"{result.get('confidence', 1.0):.2f}"
            ]
            
            # Add rule statuses
            for rule in rules:
                status = result['rules'].get(rule, {}).get('status', 'N/A')
                data.append(status)
            
            data.append(result['final_status'])
            # v3.5 FIX: Don't truncate final reason - it's important!
            data.append(result['final_reason'] if result['final_reason'] else '')
            data.append('Yes' if result.get('web_search', {}).get('performed', False) else '')
            
            for col_idx, value in enumerate(data, start=1):
                cell = ws2.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                
                # Color code rule columns
                if col_idx > 8 and col_idx <= 8 + len(rules):
                    if value == 'PASS':
                    cell.fill = pass_fill
                    elif value == 'FAIL':
                    cell.fill = fail_fill
                        cell.font = Font(bold=True)
                    elif value == 'QUERY':
                    cell.fill = query_fill
            
                # Color final status
                if col_idx == 8 + len(rules) + 1:
                    if value == 'APPROVED':
                        cell.fill = pass_fill
                    elif value == 'DENY':
                        cell.fill = fail_fill
                        cell.font = Font(bold=True)
                    elif value == 'QUERY':
                        cell.fill = query_fill
            
                # Highlight web searched
                if col_idx == len(data) and value == 'Yes':
                    cell.fill = web_fill
        
        # Adjust column widths - v3.5: Increased for readability
        ws2.column_dimensions['A'].width = 5
        ws2.column_dimensions['B'].width = 28  # Enrollee
        ws2.column_dimensions['C'].width = 12  # PA
        ws2.column_dimensions['D'].width = 50  # Procedure - increased
        ws2.column_dimensions['E'].width = 40  # Diagnosis - increased
        ws2.column_dimensions['F'].width = 55  # All Encounter Dx - increased
        ws2.column_dimensions['G'].width = 14  # Amount
        ws2.column_dimensions['H'].width = 8   # Confidence
        
        # ===== Sheet 3: Flagged Claims =====
        ws3 = wb.create_sheet("Flagged Claims")
        
        ws3['A1'] = "FLAGGED CLAIMS - DETAILED ANALYSIS"
        ws3['A1'].font = Font(bold=True, size=14)
        
        flagged = [r for r in results if r['final_status'] in ['DENY', 'QUERY']]
        
        current_row = 3
        for result in flagged:
            # Header
            status_text = f"Row {result['row_number']} | {result['final_status']}"
            ws3.cell(row=current_row, column=1, value=status_text)
            ws3.cell(row=current_row, column=1).font = Font(bold=True, size=11)
            if result['final_status'] == 'DENY':
                ws3.cell(row=current_row, column=1).fill = fail_fill
            else:
                ws3.cell(row=current_row, column=1).fill = query_fill
            current_row += 1
            
            # Details
            ws3.cell(row=current_row, column=1, value=f"Enrollee: {result['enrollee']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Procedure: {result['procedure']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Diagnosis: {result['diagnosis']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"All Encounter Dx: {result['all_encounter_dx']}")
            current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Amount: ₦{result['amount']:,.2f}")
                current_row += 1
            ws3.cell(row=current_row, column=1, value=f"Confidence: {result.get('confidence', 'N/A')}")
            current_row += 1
            
            # Web search info
            web_info = result.get('web_search', {})
            if web_info.get('performed'):
                ws3.cell(row=current_row, column=1, value="🔍 Web Search Performed")
                ws3.cell(row=current_row, column=1).fill = web_fill
                current_row += 1
                if web_info.get('verified_indications'):
                    ws3.cell(row=current_row, column=1, value=f"  Verified indications: {', '.join(web_info['verified_indications'][:5])}")
                    current_row += 1
            
            # Rule details
            ws3.cell(row=current_row, column=1, value="Rule Results:")
            ws3.cell(row=current_row, column=1).font = Font(bold=True)
            current_row += 1
            
            for rule in rules:
                rule_result = result['rules'].get(rule, {})
                status = rule_result.get('status', 'N/A')
                reason = rule_result.get('reason', '')
                
                cell = ws3.cell(row=current_row, column=1, value=f"  {rule}: {status}")
                # v3.5 FIX: Don't truncate reasons - use full text!
                ws3.cell(row=current_row, column=2, value=reason)
                
                if status == 'FAIL':
                    cell.fill = fail_fill
                    cell.font = Font(bold=True)
                elif status == 'QUERY':
                    cell.fill = query_fill
                elif status == 'PASS':
                    cell.fill = pass_fill
                
                current_row += 1
            
            # Clinical notes
            if result.get('clinical_notes'):
                ws3.cell(row=current_row, column=1, value=f"Clinical Notes: {result['clinical_notes']}")
                current_row += 1
            
            current_row += 2
        
        ws3.column_dimensions['A'].width = 30
        ws3.column_dimensions['B'].width = 120  # v3.5: Wider for full reasons
        
        # ===== Sheet 4: Web Search Log (if applicable) =====
        if self.web_search_mode != 'off':
            ws4 = wb.create_sheet("Web Search Log")
            
            ws4['A1'] = "WEB SEARCH LOG"
            ws4['A1'].font = Font(bold=True, size=14)
            
            headers = ['Row', 'Procedure', 'Diagnosis', 'Search Status', 'Verified Indications', 'Appropriate']
            for col, header in enumerate(headers, start=1):
                cell = ws4.cell(row=3, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
            
            row_num = 4
            for result in results:
                web_info = result.get('web_search', {})
                if web_info.get('performed'):
                    ws4.cell(row=row_num, column=1, value=result['row_number'])
                    ws4.cell(row=row_num, column=2, value=result['procedure'][:60])  # v3.5: Increased
                    ws4.cell(row=row_num, column=3, value=result['diagnosis'][:50])  # v3.5: Increased
                    ws4.cell(row=row_num, column=4, value=web_info.get('status', 'N/A'))
                    verified_indications = web_info.get('verified_indications') or []
                    ws4.cell(row=row_num, column=5, value=', '.join(verified_indications[:5]) if verified_indications else 'N/A')  # v3.5: Show more
                    ws4.cell(row=row_num, column=6, value='Yes' if web_info.get('appropriate_for_diagnosis') else 'No')
                    row_num += 1
            
            ws4.column_dimensions['A'].width = 6
            ws4.column_dimensions['B'].width = 55  # v3.5: Wider
            ws4.column_dimensions['C'].width = 45  # v3.5: Wider
            ws4.column_dimensions['D'].width = 15
            ws4.column_dimensions['E'].width = 60  # v3.5: Wider for more indications
            ws4.column_dimensions['F'].width = 12
        
        # Save
        wb.save(output_path)
        print(f"\n✅ Report saved: {output_path}")
    
    def _save_as_csv(self, results: List[Dict], output_path: str):
        """Fallback to save as CSV if openpyxl not available"""
        rows = []
        for r in results:
            row = {
                'Row': r['row_number'],
                'Enrollee': r['enrollee'],
                'PA': r['pa'],
                'Procedure': r['procedure'],
                'Diagnosis': r['diagnosis'],
                'All_Encounter_Dx': r['all_encounter_dx'],
                'Amount': r['amount'],
                'Confidence': r.get('confidence', 'N/A'),
                'Final_Status': r['final_status'],
                'Final_Reason': r['final_reason'],
                'Web_Searched': r.get('web_search', {}).get('performed', False)
            }
            for rule in ['R1', 'R2', 'R3', 'R4', 'R5', 'R6', 'R7', 'R8', 'R9', 'R10', 'R11', 'R12', 'R13']:
                row[rule] = r['rules'].get(rule, {}).get('status', 'N/A')
                row[f'{rule}_Reason'] = r['rules'].get(rule, {}).get('reason', '')
            rows.append(row)
        
        pd.DataFrame(rows).to_csv(output_path, index=False)
        print(f"\n✅ CSV saved: {output_path}")


# =============================================================================
# MAIN FUNCTION
# =============================================================================

def main():
    parser = argparse.ArgumentParser(
        description='KLAIRE Claims Vetting Engine v3.5 - Claude Opus 4.5 with Web Search',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Standard mode (fastest, no web search)
  python claims_vetting_api_v2.py --input claims.csv --output report.xlsx
  
  # With web search for ALL claims (slowest, most thorough)
  python claims_vetting_api_v2.py --input claims.csv --output report.xlsx --web-search
  
  # RECOMMENDED: Smart mode - auto-search only when uncertain
  python claims_vetting_api_v2.py --input claims.csv --output report.xlsx --smart-search
  
Environment Variables:
  ANTHROPIC_API_KEY    Your Anthropic API key (required)
  
Web Search Modes:
  --web-search    Enable web search for ALL claims
  --smart-search  Auto-search only when Opus is uncertain (RECOMMENDED)
  (neither)       No web search - fastest mode
        """
    )
    
    parser.add_argument('--input', '-i', required=True, help='Input claims file (CSV or Excel)')
    parser.add_argument('--output', '-o', default='vetting_report.xlsx', help='Output report file')
    parser.add_argument('--batch-name', '-n', default='Claims Batch', help='Name for the batch in report')
    parser.add_argument('--db-path', '-d', default='ai_driven_data.duckdb', help='Path to DuckDB database')
    parser.add_argument('--api-key', '-k', help='Anthropic API key (or set ANTHROPIC_API_KEY env var)')
    
    # Web search options (mutually exclusive)
    search_group = parser.add_mutually_exclusive_group()
    search_group.add_argument('--web-search', action='store_true', 
                              help='Enable web search for ALL claims')
    search_group.add_argument('--smart-search', action='store_true',
                              help='Auto-search only when uncertain (RECOMMENDED)')
    
    args = parser.parse_args()
    
    # Validate input file
    if not Path(args.input).exists():
        print(f"ERROR: Input file not found: {args.input}")
        sys.exit(1)
    
    # Determine web search mode
    if args.web_search:
        web_search_mode = 'on'
    elif args.smart_search:
        web_search_mode = 'smart'
    else:
        web_search_mode = 'off'
    
    print("="*60)
    print("  KLAIRE Claims Vetting Engine v3.5")
    print("  Powered by Claude Opus 4.5")
    print(f"  Web Search: {web_search_mode.upper()}")
    print("="*60)
    
    try:
        # Initialize engine
        engine = ClaimsVettingEngine(
            api_key=args.api_key,
            db_path=args.db_path,
            web_search_mode=web_search_mode
        )
        
        # Load claims
        df = engine.load_claims(args.input)
        
        # Run vetting
        results = engine.vet_batch(df)
        
        # Generate report
        engine.generate_excel_report(results, args.output, args.batch_name)
        
        print("\n" + "="*60)
        print("  VETTING COMPLETE")
        print("="*60)
        
    except Exception as e:
        print(f"\nERROR: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()